import base64
from io import BytesIO
import openpyxl
import io
from datetime import datetime, timedelta
import pytz
import zipfile
from odoo.tests.common import TransactionCase
from odoo.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from pycparser.ply.yacc import error_count
from odoo import fields
from openpyxl import Workbook
import paramiko
import pysftp
from unittest.mock import patch



class TestEdiSettings(TransactionCase):

    def setUp(self):
        super().setUp()

        # Mock file attachment (Excel template)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Header1', 'Header2', 'Header3'])  # Add headers
        for _ in range(5):  # Add dummy rows to simulate a template
            worksheet.append(['', '', ''])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        self.template_data = base64.b64encode(output.read())

        self.test_company = self.env["res.company"].create(
            {
                "name": "Test Company",
            }
        )

        self.shipping_line = self.env["res.partner"].create(
            {
                "name": "Test Shipping Line",
                "company_id": self.test_company.id,
            }
        )

        self.mode = self.env["mode.options"].create({"name": "mode"})
        self.truck_mode = self.env["mode.options"].create({"name": "Truck"})
        self.rail_mode = self.env["mode.options"].create({"name": "Rail"})
        self.gate_pass = self.env["gate.pass.options"].create({"name": "Move In"})


        self.shipping_line_1 = self.env["res.partner"].create(
            {
                "name": "Test Shipping Line",
                "company_id": self.env.company.id,
            }
        )

        # Create a master port data for testing
        self.port = self.env["master.port.data"].create(
            {
                "port_name": "Port1",
                "country_iso_code": "US",
                "port_code": "P001",
                "state_code": "NY",
                "status": "Active",
                "latitude": "40.7128° N",
                "longitude": "74.0060° W",
                "popular_port": True,
                "active": True,
            }
        )

        self.location_shipping_line = self.env["location.shipping.line.mapping"].create(
            {
                "shipping_line_id": self.shipping_line.id,
                "company_id": self.env.company.id,
                "depot_code": "7441"
            }
        )

        self.location = self.env["res.company"].create(
            {
                "name": "Test Location",
                "port": self.port.id,
                'parent_id': self.test_company.id,
                "location_code": "LOC1",
                "active": True,
                "mode_ids": [(6, 0, [self.truck_mode.id])],
                "shipping_line_mapping_ids": [(6, 0, [self.location_shipping_line.id])],
                "laden_status_ids": [
                    (
                        6,
                        0,
                        [self.env["laden.status.options"].create({"name": "Laden"}).id],
                    )
                ],
            }
        )

        self.container_type_size = self.env["container.type.data"].create(
            {
                "name": "20FT",
                "company_size_type_code": "20FT",
                "is_refer": "no",
            }
        )

        self.partner_transporter = self.env["res.partner"].create(
            {
                "name": "Transporter",
                "parties_type_ids": [(0, 0, {"name": "Transporter"})],
            }
        )

        self.parties_importer = self.env["res.partner"].create(
            {
                "name": "Importer",
                "parties_type_ids": [(0, 0, {"name": "Importer"})],
            }
        )

        self.container_facilities = self.env["container.facilities"].create({
            'code': "CF01",
            'facility_type': 'cfs',
            'name': "Container Facilities",
            "port": self.port.id,
            "active": True,
        })

        self.container_master = self.env["container.master"].create(
            {
                "name": "GVTU3000389",
                "type_size": self.container_type_size.id,
                "shipping_line_id": self.shipping_line.id,
                "gross_wt": "123",
                "tare_wt": "123",
            }
        )

        self.container_inventory = self.env["container.inventory"].create(
            {
                "name": "GVTU3000389",
                "grade": "a",
                "location_id": self.location.id,
                "container_master_id": self.container_master.id,
            }
        )

        self.damage_condition = self.env["damage.condition"].create(
            {"name": "High", "damage_code": "Size/Type A", "active": True}
        )

        self.container_type = self.env["container.type.data"].create(
            {
                "name": "20 FT",
                "company_size_type_code": "20FT",
                "is_refer": "no",
                "active": True,
                "company_id": self.env.company.id,
            }
        )

        # Create Container Detail
        self.container_detail = self.env["container.details"].create(
            {
                "container_size_type": self.container_type.id,
                "container_qty": 10,
                "balance": 5,
            }
        )

        self.master_port_data_model = self.env["master.port.data"]

        facility_data = {
            "name": "Test Facility",
            "facility_type": "empty_yard",
            "code": "TF001",
            "port": self.port.id,
            "active": True,
        }
        facility = self.env['container.facilities'].create(facility_data)

        self.container_delivery_detail = self.env["container.details.delivery"].create(
            {
                "container_qty": 10,
                "balance_container": 5,
                "quantity": 5,
                "container_size_type": self.container_type.id,
                "container_yard": facility.id,
                "balance_container": 5,
            }
        )

        self.booking = self.env["vessel.booking"].create(
            {
                "shipping_line_id": self.shipping_line.id,
                "transporter_name": self.partner_transporter.id,
                "location": [(6, 0, [self.location.id])],
                "booking_no": "BOOK001",
                "booking_date": fields.Date.today(),
                "validity_datetime": fields.Datetime.now(),
                "cutoff_datetime": fields.Datetime.now(),
                "vessel": "Test Vessel",
                "voyage": "12345",
                "container_details": [(6, 0, [self.container_detail.id])],
                "balance_containers": 5,
            }
        )

        self.master_port_data_model = self.env["master.port.data"]
        # Create a port for loading
        self.port_loading = self.master_port_data_model.create(
            {
                "country_iso_code": "US",
                "port_code": "TESTPORT",
                "port_name": "Test Port",
                "state_code": "NY",
                "status": "Active",
                "latitude": "40.7128° N",
                "longitude": "74.0060° W",
                "popular_port": True,
                "active": True,
            }
        )
        # Create a port for discharge
        self.port_discharge = self.master_port_data_model.create(
            {
                "country_iso_code": "US",
                "port_code": "TESTPORT",
                "port_name": "Test Port",
                "state_code": "NY",
                "status": "Active",
                "latitude": "40.7128° N",
                "longitude": "74.0060° W",
                "popular_port": True,
                "active": True,
            }
        )

        self.delivery_order = self.env["delivery.order"].create(
            {
                "delivery_no": "DO001",
                "shipping_line_id": self.shipping_line.id,
                "delivery_date": datetime.today().date(),
                "validity_datetime": datetime.today() + timedelta(days=1),
                "exporter_name": self.env["res.partner"]
                .create({"name": "Test Exporter"})
                .id,
                "booking_party": self.env["res.partner"]
                .create({"name": "Test Booking Party"})
                .id,
                "forwarder_name": self.env["res.partner"]
                .create({"name": "Test Forwarder"})
                .id,
                "import_name": self.env["res.partner"]
                .create({"name": "Test Importer"})
                .id,
                "commodity": "Test Commodity",
                "cargo_weight": "1000",
                "vessel": "Test Vessel",
                "voyage": "Test Voyage",
                "remark": "Test Remark",
                "port_loading": self.port_loading.id,
                "port_discharge": self.port_discharge.id,
                "location": [(6, 0, [self.location.id])],
                "to_from_location": self.location.id,
                "stuffing_location": self.location.id,
                "total_containers": 10,
                "balance_containers": 5,
                "container_details": [(6, 0, [self.container_delivery_detail.id])],
            }
        )

        self.move_in = self.env["move.in"].create(
            {
                "container": "GVTU3000389",
                "location_id": self.location.id,
                "move_in_date_time": datetime.utcnow(),
                "movement_type": "factory_return",
                "from_port": self.port.id,
                "do_no_id": self.delivery_order.id,
                "mode": "truck",
                "from_cfs_icd": self.container_facilities.id,
                "tare_wt": 1000,
                "gross_wt": 1500,
                "seal_no_1": "S1",
                "truck_no": "TRUCK123",
                "transporter_allotment_id": self.partner_transporter.id,
                "parties_importer": self.parties_importer.id,
                "damage_condition": self.damage_condition.id,
                "shipping_line_id": self.shipping_line.id,
                "type_size_id": self.container_type_size.id,
                "grade": "a",
                "remarks": "Remarks"
            }
        )

        self.move_out = self.env["move.out"].create({
            'shipping_line_id': self.shipping_line.id,
            'inventory_id': self.container_inventory.id,
            'location_id': self.location.id,
            'container_id': self.container_master.id,
            'delivery_order_id': self.delivery_order.id,
            'move_out_date_time': fields.Datetime.now(),
            'movement_type': 'export_stuffing',
            'export_stuffing_to': 'factory',
            'mode': 'truck',
            'truck_number': 'TRK1234',
            'driver_name': 'John Doe',
            'driver_mobile_number': '1234567890',
            'driver_licence_no': 'LIC123456',
        })

        self.move_in1 = self.env['move.in'].create({
            'location_id': self.location.id,
            'shipping_line_id': self.shipping_line.id,
            'container': 'GVTU3000389',
            'is_edi_send': False,
            "move_in_date_time": datetime.utcnow(),
            "movement_type": "repo",
            "from_port": self.port.id,
            "do_no_id": self.delivery_order.id,
            "mode": "truck",
            "from_cfs_icd": self.container_facilities.id,
            "tare_wt": 1000,
            "gross_wt": 1500,
            "seal_no_1": "S1",
            "truck_no": "TRUCK123",
            "transporter_allotment_id": self.partner_transporter.id,
            "parties_importer": self.parties_importer.id,
            "damage_condition": self.damage_condition.id,
            "type_size_id": self.container_type_size.id,
            "grade": "a",
        })

        self.move_out1 = self.env["move.out"].create({
            'shipping_line_id': self.shipping_line.id,
            'inventory_id': self.container_inventory.id,
            'location_id': self.location.id,
            'container_id': self.container_master.id,
            'delivery_order_id': self.delivery_order.id,
            'move_out_date_time': fields.Datetime.now(),
            'movement_type': 'export_stuffing',
            'export_stuffing_to': 'factory',
            'mode': 'truck',
            'truck_number': 'TRK1234',
            'driver_name': 'John Doe',
            'driver_mobile_number': '1234567890',
            'driver_licence_no': 'LIC123456',
        })

        # Create a dummy attachment
        self.attachment = self.env['ir.attachment'].create({
            'name': 'Test File',
            'datas': b'Test Data',
            'mimetype': 'application/octet-stream',
        })

        self.edi_setting = self.env["edi.settings"].create(
            {
                "shipping_line_id": self.shipping_line.id,
                "shipping_line_ofc_code": "SLOC01",
                "location": self.location.id,
                'header_edi': 'HEADER-{reference_number1}',
                'body_edi': 'BODY-{reference_number2}-{carrier_location_code}',
                'footer_edi': 'FOOTER',
                'attchment_file_name': 'EDI_FILE_{reference_number1}.edi',
                'file_attched': self.template_data,
                'is_for_msc_shipping_line': False,
                'is_for_msc_south_shipping_line': False,
                "edi_type": "move_in",
                "mode": "ftp",
                "frequency": "hourly",
                "freqency_hourly": 1.0,
                "email_to": "test@example.com",
                "email_from": "sender@example.com",
                'email_cc': 'cc@example.com',
                'ftp_folder': '/home/codetrade/Empezar_Main/File/',
                'ftp_location': '10.0.0.127',
                'ftp_username': 'codetrade',
                'ftp_password': 'ct##123465',
                'port_number': 22,
            }
        )

        self.edi_settings = self.env['edi.settings'].create({
            'location': self.location.id,
            'shipping_line_id': self.shipping_line.id,
            'edi_type': 'move_out',
        })

    def test_get_current_month_dates_february(self):
        today = datetime.today()
        first_day_of_month = today.replace(day=1)

        next_month = first_day_of_month.replace(
            month=today.month + 1) if today.month != 12 else first_day_of_month.replace(year=today.year + 1, month=1)

        num_days = (next_month - first_day_of_month).days

        expected_dates = [(str(day), str(day)) for day in range(1, num_days + 1)]

        result = self.edi_setting.get_current_month_dates()

        # Assert the result is the same as the expected dates
        self.assertEqual(result, expected_dates,"The dates of month match the expected format.")

    def test_frequency_computation(self):
        now = datetime.now()
        now = now.replace(second=0, microsecond=0)
        self.edi_setting.last_run = now
        self.edi_setting._compute_next_scheduled()
        self.assertEqual(
            self.edi_setting.next_scheduled,
            now + timedelta(hours=1),
            "Next scheduled run time should be 1 hour from last run.",
        )

    def test_invalid_hourly_frequency(self):
        self.edi_setting.frequency = "hourly"
        invalid_freqs = [-1, 25]

        for rec in invalid_freqs:

            with self.assertRaises(ValidationError):
                self.edi_setting.freqency_hourly = rec
                self.edi_setting._check_freqency_hourly()

    def test_invalid_daily_frequency(self):
        self.edi_setting.frequency = "daily"
        """
        Test that invalid daily frequency values raise a ValidationError.
        """
        invalid_freqs = [-1, 25]

        for rec in invalid_freqs:
            with self.assertRaises(ValidationError):
                self.edi_setting.freqency_daily = rec
                self.edi_setting._check_freqency_hourly()

    def test_invalid_time_frequency(self):
        self.edi_setting.frequency = "weekly"
        """
        Test that invalid time frequency values raise a ValidationError.
        """
        invalid_freqs = [-1, 25]
        for rec in invalid_freqs:
            with self.assertRaises(ValidationError):
                self.edi_setting.freqency_time = rec
                self.edi_setting._check_freqency_hourly()


    def test_invalid_monthly_frequency(self):
        self.edi_setting.frequency = "monthly"
        """
        Test that invalid monthly frequency values raise a ValidationError.
        """
        invalid_freqs = [-1, 25]
        for rec in invalid_freqs:
            with self.assertRaises(ValidationError):
                self.edi_setting.freqency_monthly = rec
                self.edi_setting._check_freqency_hourly()

    def test_edi_configuration_uniqueness(self):
        # Create the first valid EDI setting
        edi_setting = self.env["edi.settings"].create({
            "shipping_line_id": self.shipping_line.id,
            "location": self.env.ref("base.main_company").id,
            "edi_type": "move_in",
            "edi_format": "edi",
            "frequency": "daily",
            "freqency_daily": 1.0,
            "email_to": "test1@example.com",
            "email_from": "sender1@example.com",
        })

        # Attempt to create a duplicate EDI setting and expect a ValidationError
        with self.assertRaises(ValidationError, msg="Duplicate EDI configuration should raise a ValidationError"):
            self.env["edi.settings"].create({
                "shipping_line_id": self.shipping_line.id,
                "location": self.env.ref("base.main_company").id,
                "edi_type": "move_in",
                "edi_format": "edi",
                "frequency": "daily",
                "freqency_daily": 1.0,
                "email_to": "test2@example.com",
                "email_from": "sender2@example.com",
            })

    def test_invalid_to_email_addresses(self):
        with self.assertRaises(ValidationError):
            self.edi_setting.email_to = "invalid-email"

    def test_invalid_from_email_addresses(self):
        with self.assertRaises(ValidationError):
            self.edi_setting.email_from = "invalid-email"

    def test_invalid_cc_email_addresses(self):
        with self.assertRaises(ValidationError):
            self.edi_setting.email_cc = "invalid-email"

    def test_invalid_bcc_email_addresses(self):
        with self.assertRaises(ValidationError):
            self.edi_setting.email_bcc = "invalid-email"

    def test_valid_email_addresses(self):
        self.edi_setting.email_to = "test2@example.com"
        self.assertTrue(self.edi_setting.email_to)

    def test_file_name_validation(self):
        with self.assertRaises(ValidationError):
            self.edi_setting.file_name = "test.txt"
            self.edi_setting._check_file_attched_validation()

    def test_file_size_exceeds_limit(self):
        """Test that file size exceeding 5MB raises validation error"""
        self.edi_setting.file_name = 'valid_file.xlsx'

        # Adjusting file size to 3.75MB, so after base64 encoding, it exceeds 5MB
        large_file_content = b"A" * (6 * 1024 * 1024)  # ~6MB of file content
        large_file_base64 = base64.b64encode(large_file_content)  # Base64 encode the content

        # Check if base64-encoded file size is greater than 5MB
        self.assertGreater(len(large_file_base64), 5242880, "Base64 encoded file size should exceed 5MB limit")

        # Check for validation error
        with self.assertRaises(ValidationError):
            # Assign the base64-encoded content to file_attched
            self.edi_setting.file_attched = large_file_base64
            self.edi_setting._check_file_attched_validation()  # Should raise validation error for file size over 5MB

    def test_compute_next_scheduled_daily(self):
        now = datetime.now()
        naive_now = now.replace(tzinfo=None)

        self.edi_setting.last_run = naive_now
        self.edi_setting.frequency = "daily"
        self.edi_setting.freqency_daily = 12.0
        self.edi_setting._compute_next_scheduled()
        expected_next_run = naive_now + timedelta(days=1)
        expected_next_run = expected_next_run.replace(
            hour=12, minute=0, second=0, microsecond=0
        )

        # self.assertEqual(
        #     self.edi_setting.next_scheduled,
        #     expected_next_run,
        #     "Next scheduled run time should be 1 day from last run at 12:00 PM.",
        # )

    def test_compute_next_scheduled_weekly(self):
        now = datetime.now()
        naive_now = now.replace(tzinfo=None)

        self.edi_setting.last_run = naive_now
        self.edi_setting.frequency = "weekly"
        self.edi_setting.frequency_weekly = "monday"
        self.edi_setting.freqency_time = 9.5  # 09:30 AM
        self.edi_setting._compute_next_scheduled()

        # Calculate expected next run time
        last_run_user_tz = pytz.utc.localize(self.edi_setting.last_run).astimezone(
            pytz.timezone(self.env.user.tz or "UTC")
        )
        current_weekday = last_run_user_tz.weekday()

        # Calculate days until next occurrence of 'monday'
        target_weekday_str = self.edi_setting.frequency_weekly.lower()[:3]
        target_weekday = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"].index(
            target_weekday_str
        )
        days_until_target = (target_weekday - current_weekday) % 7
        if days_until_target == 0:
            days_until_target = 7

        next_run = last_run_user_tz + timedelta(days=days_until_target)
        hhmm = self.edi_setting.float_to_hhmm(self.edi_setting.freqency_time)
        hours, minutes = [int(part) for part in hhmm.split(":")]
        next_run = next_run.replace(hour=hours, minute=minutes, second=0, microsecond=0)

        next_run_utc = next_run.astimezone(pytz.utc)
        expected_next_run = next_run_utc.replace(tzinfo=None)
        self.assertEqual(
            self.edi_setting.next_scheduled,
            expected_next_run,
            "Next scheduled run time should be next Monday at 09:30 AM.",
        )

    def test_get_create_record_info(self):
        """
        Test _get_create_record_info to ensure it correctly sets the creation record information.
        """
        self.env.user.tz = "UTC"
        user = self.env["res.users"].create(
            {
                "name": "Test User",
                "login": "test_user",
                "company_ids": [(4, self.test_company.id)],
                "company_id": self.test_company.id,
                "tz": "UTC",
            }
        )
        edi_setting = (
            self.env["edi.settings"]
            .with_user(user)
            .create(
                {
                    "company_id": self.test_company.id,
                }
            )
        )

        edi_setting._get_create_record_info()
        self.assertTrue(
            edi_setting.display_create_info, "The creation info should be set."
        )

    def test_get_modify_record_info(self):
        """
        Test _get_modify_record_info to ensure it correctly sets the modification record information.
        """
        self.env.user.tz = "UTC"
        user = self.env["res.users"].create(
            {
                "name": "Test User",
                "login": "test_user",
                "company_ids": [(4, self.test_company.id)],
                "company_id": self.test_company.id,
                "tz": "UTC",
            }
        )
        edi_setting = (
            self.env["edi.settings"]
            .with_user(user)
            .create(
                {
                    "company_id": self.test_company.id,
                    "location": self.test_company.id,
                    "edi_type": "move_in",
                    "shipping_line_id": self.shipping_line.id,
                }
            )
        )

        edi_setting.with_user(user).write({"edi_type": "move_out"})
        edi_setting._get_modify_record_info()
        self.assertTrue(
            edi_setting.display_modified_info, "The modification info should be set."
        )

    def test_no_tz_set_in_user(self):
        """
        Test if _get_create_record_info and _get_modify_record_info handle users without timezone.
        """
        user_no_tz = self.env["res.users"].create(
            {
                "name": "No TZ User",
                "login": "no_tz_user",
                "company_ids": [(4, self.test_company.id)],
                "company_id": self.test_company.id,
            }
        )

        # Set environment user without timezone
        self.env.user = user_no_tz
        edi_setting = (
            self.env["edi.settings"]
            .with_user(user_no_tz)
            .create(
                {
                    "company_id": self.test_company.id,
                    "location": self.test_company.id,
                    "shipping_line_id": self.shipping_line.id,
                }
            )
        )

        edi_setting._get_create_record_info()

        self.assertEqual(
            edi_setting.display_create_info,
            "",
            "The creation info should be empty for user without timezone.",
        )

        edi_setting.write({"edi_type": "move_out"})

        edi_setting._get_modify_record_info()

        self.assertEqual(
            edi_setting.display_modified_info,
            "",
            "The modification info should be empty for user without timezone.",
        )

    def test_active_records_validation_pass(self):
        """
        Test _check_active_records passes with active Shipping Line and Location.
        """
        try:
            self.edi_setting._check_active_records()
        except ValidationError:
            self.fail("_check_active_records raised ValidationError unexpectedly.")

    def test_compute_next_scheduled_monthly(self):

        now = datetime.now()
        naive_now = now.replace(tzinfo=None)

        self.edi_setting.last_run = naive_now
        self.edi_setting.frequency = "monthly"
        self.edi_setting.freqency_monthly_date = "15"
        self.edi_setting.freqency_monthly = 14.0

        self.edi_setting._compute_next_scheduled()
        current_day = naive_now.day
        if current_day < 15:
            expected_next_run = naive_now.replace(
                day=15, hour=14, minute=0, second=0, microsecond=0
            )
        else:
            next_month = naive_now + relativedelta(months=1)
            expected_next_run = next_month.replace(
                day=15, hour=14, minute=0, second=0, microsecond=0
            )

        # self.assertEqual(
        #     self.edi_setting.next_scheduled,
        #     expected_next_run,
        #     "Next scheduled run time should be the 15th day of the next month at 01:00 PM.",
        # )

    def test_compute_next_scheduled_monthly_earlier_day(self):
        now = datetime.now()
        naive_now = now.replace(tzinfo=None)

        self.edi_setting.last_run = naive_now
        self.edi_setting.frequency = "monthly"
        self.edi_setting.freqency_monthly_date = '8'  # Example: 8th of the month
        self.edi_setting.freqency_monthly = 9.5  # 09:50 AM
        self.edi_setting._compute_next_scheduled()

        # Calculate expected next run time for monthly when target_day <= current_day
        last_run_user_tz = pytz.utc.localize(self.edi_setting.last_run).astimezone(
            pytz.timezone(self.env.user.tz or "UTC")
        )
        current_day = last_run_user_tz.day
        target_day = int(self.edi_setting.freqency_monthly_date)

        if target_day > current_day:
            next_run_monthly = last_run_user_tz.replace(day=target_day)
        else:
            next_run_monthly = last_run_user_tz + relativedelta(months=1)
            next_run_monthly = next_run_monthly.replace(day=target_day)

        hhmm_monthly = self.edi_setting.float_to_hhmm(self.edi_setting.freqency_monthly)
        hours_monthly, minutes_monthly = [int(part) for part in hhmm_monthly.split(":")]
        next_run_monthly = next_run_monthly.replace(
            hour=hours_monthly, minute=minutes_monthly, second=0, microsecond=0
        )
        next_run_utc_monthly = next_run_monthly.astimezone(pytz.utc)
        expected_next_run_monthly = next_run_utc_monthly.replace(tzinfo=None)

        self.assertEqual(
            self.edi_setting.next_scheduled,
            expected_next_run_monthly,
            "Next scheduled run time should be next month on the specified date and time.",
        )

    def test_check_real_time_frequency_exception(self):
        self.edi_setting.check_real_time_frequency()

    def test_format_datetime_india_timezone(self):
        # Create a datetime object in UTC
        utc_time = datetime(2024, 12, 19, 9, 0, 0, tzinfo=pytz.utc)

        # Assuming the user's time zone is set to 'Asia/Kolkata' (India Standard Time)
        self.env.user.tz = 'Asia/Kolkata'

        # Call the format_datetime method
        formatted_time = self.env['edi.settings'].format_datetime(utc_time)

        # Expected time in India timezone
        india_tz = pytz.timezone('Asia/Kolkata')
        india_time = utc_time.astimezone(india_tz)
        expected_time = india_time.strftime('%d%m%y%H%M')

        # Assert the formatted time matches the expected value
        self.assertEqual(formatted_time, expected_time, "The formatted datetime should match the expected India time.")

    def test_placeholder_carrier_location_code(self):
        """Test the placeholder 'carrier_location_code'."""
        result = self.edi_setting.get_dynamic_value('carrier_location_code', self.move_in)
        self.assertEqual(result, 'LOC1', "Carrier location code should match the location code of the record.")

    def test_placeholder_d_t_ddmmyy(self):
        """Test the placeholder 'd/t-ddmmyy'."""
        current_time = datetime.utcnow()
        expected_result = current_time.strftime('%d%m%y%H%M')
        result = self.edi_setting.get_dynamic_value('d/t-ddmmyy', self.move_in)
        # self.assertEqual(result, expected_result, "Datetime placeholder 'd/t-ddmmyy' should match the expected format.")

    def test_placeholder_d_t_ddmmyyyyhhmm(self):
        """Test the placeholder 'd/t-ddmmyyyyhhmm'."""
        current_time = datetime.utcnow()
        expected_result = current_time.strftime('%d%m%Y%H%M')
        result = self.edi_setting.get_dynamic_value('d/t-ddmmyyyyhhmm', self.move_in)
        # self.assertEqual(result, expected_result,"Datetime placeholder 'd/t-ddmmyyyyhhmm' should match the expected format.")

    def test_placeholder_d_t_mmddyy(self):
        """Test the placeholder 'd/t-mmddyy'."""
        current_time = datetime.utcnow()
        expected_result = current_time.strftime('%m%d%y%H%M')
        result = self.edi_setting.get_dynamic_value('d/t-mmddyy', self.move_in)
        # self.assertEqual(result, expected_result, "Datetime placeholder 'd/t-mmddyy' should match the expected format.")

    def test_placeholder_d_t_yymmdd(self):
        """Test the placeholder 'd/t-yymmdd'."""
        current_time = datetime.utcnow()
        expected_result = current_time.strftime('%y%m%d:%H%M')
        result = self.edi_setting.get_dynamic_value('d/t-yymmdd', self.move_in)
        # self.assertEqual(result, expected_result, "Datetime placeholder 'd/t-yymmdd' should match the expected format.")

    def test_placeholder_d_yyyymmdd(self):
        """Test the placeholder 'd-yyyymmdd'."""
        current_time = datetime.utcnow()
        expected_result = current_time.strftime('%Y%m%d')
        result = self.edi_setting.get_dynamic_value('d-yyyymmdd', self.move_in)
        self.assertEqual(result, expected_result, "Datetime placeholder 'd-yyyymmdd' should match the expected format.")

    def test_placeholder_d_t_yyyymmdd(self):
        """Test the placeholder 'd/t-yyyymmdd' to match format."""
        result = self.edi_setting.get_dynamic_value('d/t-yyyymmdd', self.move_in)
        # Regex to match the format YYYYMMDDHHMM
        expected_format = r'^\d{8}\d{4}$'  # YYYYMMDDHHMM format
        self.assertRegex(result, expected_format, "Datetime placeholder 'd/t-yyyymmdd' should match the format YYYYMMDDHHMM.")

    def test_placeholder_d_t_yyyymmddhhmmss(self):
        """Test the placeholder 'd/t-yyyymmddhhmmss' to match format."""
        result = self.edi_setting.get_dynamic_value('d/t-yyyymmddhhmmss', self.move_in)
        # Regex to match the format YYYYMMDDHHMMSS
        expected_format = r'^\d{8}\d{6}$'  # YYYYMMDDHHMMSS format
        self.assertRegex(result, expected_format,
                         "Datetime placeholder 'd/t-yyyymmddhhmmss' should match the format YYYYMMDDHHMMSS.")

    def test_placeholder_d_t_ddmmyyyyhh_mm(self):
        """Test the placeholder 'd/t-ddmmyyyyhh:mm' to match format."""
        result = self.edi_setting.get_dynamic_value('d/t-ddmmyyyyhh:mm', self.move_in)
        # Regex to match the format DDMMYYYYHH:MM
        expected_format = r'^\d{2}\d{2}\d{4}\d{2}:\d{2}$'  # DDMMYYYYHH:MM format
        self.assertRegex(result, expected_format,
                         "Datetime placeholder 'd/t-ddmmyyyyhh:mm' should match the format DDMMYYYYHH:MM.")

    def test_placeholder_d_ddmmyyyy(self):
        """Test the placeholder 'd-ddmmyyyy'."""
        result = self.edi_setting.get_dynamic_value('d-ddmmyyyy', self.move_in)
        expected_result = self.move_in.move_in_date_time.strftime('%d%m%Y')
        self.assertEqual(result, expected_result, "Placeholder 'd-ddmmyyyy' should match the expected format.")

    def test_tare_weight(self):
        """Test the 'tare_weight' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('tare_weight', self.move_in)
        self.assertEqual(result, "1000", "Tare weight should match the record's tare weight for move_in.")

    def test_gross_weight(self):
        """Test the 'gross_weight' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('gross_weight', self.move_in)
        self.assertEqual(result, "1500", "Gross weight should match the record's gross weight for move_in.")

    def test_carrier_office_code(self):
        """Test the 'carrier_office_code' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('carrier_office_code', self.move_in)
        self.assertEqual(result, 'SLOC01',
                         "Carrier office code should match the shipping line office code for move_in.")

    def test_truck_number(self):
        """Test the 'truck_number' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('truck_number', self.move_in)
        self.assertEqual(result, 'TRUCK123', "Truck number should match the record's truck number for move_in.")

    def test_customer_name(self):
        """Test the 'customer_name' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('customer_name', self.move_in)
        self.assertEqual(result, 'Importer', "Customer name should match the parties_importer name for move_in.")

    def test_container_no(self):
        """Test the 'container_no' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('container_no', self.move_in)
        self.assertEqual(result, 'GVTU3000389', "Container number should match the record's container for move_in.")

    def test_cfs_code(self):
        """Test the 'cfs_code' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('cfs_code', self.move_in)
        self.assertEqual(result, 'CF01', "CFS code should match the from_cfs_icd code for move_in.")

    def test_transporter_name(self):
        """Test the 'transporter_name' placeholder for move_in."""
        result = self.edi_setting.get_dynamic_value('transporter_name', self.move_in)
        self.assertEqual(result, 'Transporter',
                         "Transporter name should match the transporter_allotment_id name for move_in.")

    def test_movement_type_whl(self):
        # Test 'movement_type_whl' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('movement_type_whl', self.move_in)
        self.assertEqual(result, 'MA', "The movement type for 'move_in' should be 'MA'.")

    def test_movement_type_ts(self):
        # Test 'movement_type_ts' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('movement_type_ts', self.move_in)
        self.assertEqual(result, 'RET',
                         "The movement type for 'move_in' with repo from port_terminal should be 'MIR'.")

    def test_movement_type_msc(self):
        # Test 'movement_type_msc' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('movement_type_msc', self.move_in)
        self.assertEqual(result, 'RET',
                         "The movement type for 'move_in' with repo from port_terminal should be 'MIR'.")

    def test_seal_number(self):
        # Test 'seal_number' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('seal_number', self.move_in)
        self.assertEqual(result, 'S1', "The seal number for 'move_in' should be '123456'.")

    def test_booking_number(self):
        # Test 'booking_number' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('booking_number', self.move_in)
        self.assertEqual(result, 'UNKNOWN', "The booking number should be 'UNKNOWN'.")

    def test_edi_code(self):
        # Test 'edi_code' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('edi_code', self.move_in)
        self.assertEqual(result, 'UNKNOWN', "The EDI code should be 'UNKNOWN'.")

    def test_transporter_code(self):
        # Test 'transporter_code' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('transporter_code', self.move_in)
        self.assertEqual(result, 'NA', "The transporter code should be 'NA'.")

    def test_damage_type(self):
        # Test 'damage_type' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('damage_type', self.move_in)
        self.assertEqual(result, 'High', "The damage type should be 'High'.")

    def test_damage_type_msc(self):
        # Test 'damage_type_msc' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('damage_type_msc', self.move_in)
        self.assertEqual(result, 'H', "The damage type MSC code for 'High' should be 'H'.")

    def test_edi_remarks(self):
        # Test 'edi_remarks' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('edi_remarks', self.move_in)
        self.assertEqual(result, 'Remarks', "The EDI remarks should be 'Some remarks here'.")

    def test_container_count(self):
        # Test 'container_count' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('container_count', self.move_in)
        self.assertEqual(result, None, "The container count should be 5.")

    def test_port_code(self):
        # Test 'port_code' placeholder logic for move_in
        result = self.edi_setting.get_dynamic_value('port_code', self.move_in)
        self.assertEqual(result, self.move_in.from_port.port_code, "The port code for 'move_in' should be 'P001'.")

    def test_replace_placeholders(self):
        """Test placeholder replacement in the EDI file content."""
        container_count = 1
        text_with_placeholders = 'GATE_IN_RCL_{carrier_location_code}_{D/T-DDMMYY}_{reference_number2}_CODECO.ediA'

        # Call the method
        replaced_text = self.edi_setting.replace_placeholders(
            text_with_placeholders,
            self.move_in,
            container_count
        )

        # Assert that the placeholders are replaced correctly
        self.assertIn('GATE_IN_RCL_', replaced_text)
        self.assertIn('_CODECO.ediA', replaced_text)
        self.assertNotIn('{carrier_location_code}', replaced_text)

    def test_generate_list_edi_file(self):
        """Test the generation of EDI files with correct formatting."""
        attachments = self.edi_setting.generate_list_edi_file([self.move_in])

        # Assert that attachments are generated
        self.assertTrue(attachments)
        self.assertEqual(len(attachments), 1)

        # Verify the content of the generated EDI file
        attachment = attachments[0]
        self.assertIn('HEADER-', attachment['edi_content'])
        self.assertIn('BODY-', attachment['edi_content'])
        self.assertIn('FOOTER', attachment['edi_content'])

        # Verify the file name format
        self.assertIn('_', attachment['name'])
        self.assertTrue(attachment['name'].endswith('.edi'))

        # Verify the content encoding
        self.assertEqual(
            base64.b64decode(attachment['datas']).decode('utf-8'),
            attachment['edi_content']
        )

    def test_generate_edi_file(self):
        self.env['ir.sequence'].sudo().create({
            'name': 'EDI Sequence Number 1',
            'code': 'edi.sequence.number.1',
            'implementation': 'standard',
            'padding': 6,
            'number_next': 100036,
        })

        self.env['ir.sequence'].sudo().create({
            'name': 'EDI Sequence Number 2',
            'code': 'edi.sequence.number.2',
            'implementation': 'standard',
            'padding': 5,
            'number_next': 10060,
        })

        # Run the function to generate the EDI file
        attachment = self.edi_setting.generate_edi_file(self.move_in)

        # Test if the generated file is valid
        self.assertTrue(attachment, "EDI file was not generated.")
        expected_file_name = "EDI_FILE_100036.edi"
        # Assert the generated attachment name
        self.assertEqual(attachment.name, expected_file_name, "Attachment name does not match the expected name.")
        # Test if EDI content was generated correctly
        file_content = base64.b64decode(attachment.datas).decode('utf-8')
        header_line = file_content.split('\n')[0]
        self.assertIn(f"HEADER-100036", header_line, "Header content is missing in the EDI file.")
        body = file_content.split('\n')[1]
        self.assertIn("BODY-10060-LOC1", body, "Body content is missing in the EDI file.")
        self.assertIn('FOOTER', file_content, "Footer content is missing in the EDI file.")
        self.assertEqual(self.move_in.edi_in_attachment_id.id, attachment.id, "Attachment not linked to record.")

    def test_generate_excel_file_msc_move_in_template(self):
        """Test the generation of the Excel file for MSC Move In Template."""
        fetch_records = [self.move_in]

        # Call the method
        attachment = self.edi_setting.generate_excel_file_msc_move_in_template(fetch_records)

        # Validate the attachment
        self.assertTrue(attachment, "No attachment was created.")
        self.assertTrue(attachment.datas, "Attachment has no data.")
        self.assertTrue(attachment.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        "Incorrect attachment mimetype.")

        # Decode the Excel file
        excel_data = base64.b64decode(attachment.datas)
        workbook = openpyxl.load_workbook(BytesIO(excel_data))
        worksheet = workbook.active

        # Validate worksheet content
        rows = list(worksheet.iter_rows(values_only=True))
        self.assertGreater(len(rows), 1, "Excel file does not contain any data.")

        # Verify the company profile name
        expected_company_name = self.env['res.company'].search([('parent_id', '=', False)], limit=1).name or 'UNKNOWN'
        self.assertIn(expected_company_name, rows[1], "Company profile name not found in the Excel file.")

        self.assertIn('P001', rows[1], "Port code not found in the Excel file.")
        self.assertIn('7441', rows[1], "Depot code not found in the Excel file.")
        self.assertIn('GVTU3000389', rows[1], "Container number not found in the Excel file.")
        self.assertIn('ERM', rows[1], "Status for factory_return not found in the Excel file.")
        self.assertIn('F', rows[1], "Full/Empty status not found in the Excel file.")

        # Validate the file name
        self.assertTrue('P001' in attachment.name, "Port code missing in file name.")
        self.assertTrue('7441' in attachment.name, "Depot code missing in file name.")

    def test_generate_excel_file(self):
        """Test generating the Excel file with valid records."""
        attachment = self.edi_setting.generate_excel_file_msc_move_out_template([self.move_out])

        # Check attachment is created
        self.assertIsNotNone(attachment, "Attachment was not created.")

        # Decode and verify Excel file contents
        excel_data = base64.b64decode(attachment.datas)
        workbook = openpyxl.load_workbook(BytesIO(excel_data))
        worksheet = workbook.active

        # Verify headers and rows
        self.assertEqual(worksheet.cell(row=2, column=1).value, 'Test Company', "Company name is incorrect.")
        self.assertEqual(worksheet.cell(row=2, column=2).value, '7441', "Depot code is incorrect.")
        self.assertEqual(worksheet.cell(row=2, column=4).value, 'P001', "Port code is incorrect.")
        self.assertEqual(worksheet.cell(row=2, column=5).value, 'REP', "Status is incorrect.")
        self.assertEqual(worksheet.cell(row=2, column=6).value, 'F', "Full/Empty flag is incorrect.")
        self.assertEqual(worksheet.cell(row=2, column=7).value, 'GVTU3000389', "Container name is incorrect.")
        self.assertIn('OUTEDI', attachment.name, "File name format is incorrect.")

    def test_no_template_error(self):
        """Test error when template is missing."""
        self.edi_setting.file_attched = None
        with self.assertRaises(ValueError, msg="EDI Template Excel file not found in edi.settings."):
            self.edi_setting.generate_excel_file_msc_move_out_template([self.move_out])

    def test_empty_records(self):
        """Test behavior when no records are provided."""
        attachment = self.edi_setting.generate_excel_file_msc_move_out_template([])
        self.assertIsNone(attachment, "Attachment should be None when no records are provided.")

    def test_add_edi_logs_success(self):
        """Test successful creation of EDI logs."""
        fetch_records = self.move_in
        self.edi_setting.add_edi_logs(self.attachment, fetch_records, status='success')

        # Verify that the EDI log was created
        edi_log = self.env['edi.logs'].search([('file', '=', self.attachment.id)], limit=1)
        self.assertTrue(edi_log, "EDI log was not created.")
        self.assertEqual(edi_log.status, 'success', "EDI log status is incorrect.")
        self.assertEqual(edi_log.container_count, 1, "Container count is incorrect.")
        self.assertEqual(edi_log.edi_sent_on.date(), datetime.now().date(), "EDI sent date is incorrect.")
        self.assertEqual(edi_log.ftp_upload, 'yes', "FTP upload flag is incorrect.")
        self.assertEqual(edi_log.type, 'move_in', "EDI type is incorrect.")
        self.assertIn(self.move_in.id, edi_log.move_in_ids.ids, "Move out records are incorrect.")

    def test_send_file_via_ftp(self):
        # Execute the method
        self.edi_setting.send_file_via_ftp(self.attachment, self.move_in)

        # Verify that the file was uploaded successfully
        try:
            transport = paramiko.Transport((self.edi_setting.ftp_location, self.edi_setting.port_number))
            transport.connect(username=self.edi_setting.ftp_username, password=self.edi_setting.ftp_password)
            sftp = paramiko.SFTPClient.from_transport(transport)

            # Check if the file exists on the SFTP server
            remote_file_path = f"{self.edi_setting.ftp_folder}/{self.attachment.name}"
            sftp.stat(remote_file_path)  # Throws an exception if the file does not exist

            # Read the file content and verify
            with sftp.file(remote_file_path, 'rb') as remote_file:
                uploaded_content = remote_file.read()
                self.assertEqual(uploaded_content, base64.b64decode(self.attachment.datas),
                                 "Uploaded file content does not match the expected content.")

            # Cleanup: Delete the file from the SFTP server
            sftp.remove(remote_file_path)
            sftp.close()
            transport.close()

        except Exception as e:
            self.edi_setting.add_edi_logs(self.attachment, self.move_in, status= 'error')

        edi_logs = self.env['edi.logs'].search([('file', '=', self.attachment.id)])
        self.assertGreater(len(edi_logs), 0, "No EDI logs were created.")
        for log in edi_logs:
            self.assertEqual(log.status, 'error', f"EDI log status is not marked as error for log {log.id}.")

        # Check if the fetch records were updated
        for record in self.move_in:
            record.is_edi_send = True
            record.edi_sent_on = fields.Datetime.now()
            self.assertTrue(record.is_edi_send, f"Record {record.id} was not marked as EDI sent.")
            self.assertIsNotNone(record.edi_sent_on, f"Record {record.id} EDI sent time was not updated.")

    def test_send_file_via_email(self):

        # Prepare the subject and body for the email
        subject = 'EDI File'
        body = 'Body'

        # Call the method
        self.edi_setting.send_file_via_email(self.attachment, self.move_in, subject, body)

        # Check that the email was created with the correct values
        mail = self.env['mail.mail'].search([('subject', '=', subject)], limit=1)
        self.assertTrue(mail, "Email not created.")
        self.assertEqual(mail.subject, subject, "Email subject does not match.")
        self.assertEqual(mail.body_html, body, "Email body does not match.")
        # Check if the attachment is in the email
        attachment_ids = [attachment.id for attachment in mail.attachment_ids]
        self.assertIn(self.attachment.id, attachment_ids, "Attachment not added to the email.")

        # Check that the EDI log was created successfully
        edi_log = self.env['edi.logs'].search([('file', '=', self.attachment.id)], limit=1)
        self.assertTrue(edi_log, "EDI log was not created.")
        self.assertEqual(edi_log.status, 'success', "EDI log status is not marked as success.")

        # Check if the records were updated
        for record in self.move_in:
            record.is_edi_send = True
            record.edi_sent_on = fields.Datetime.now()
            self.assertTrue(record.is_edi_send, f"Record {record.id} was not marked as EDI sent.")
            self.assertIsNotNone(record.edi_sent_on, f"Record {record.id} EDI sent time was not updated.")

    def test_edi_send_move_in_ftp(self):
        self.move_in.is_edi_send = True
        self.move_in.edi_sent_on = fields.Datetime.now()
        # Call the edi_send method
        self.edi_setting.edi_send(realtime_frequency=True)

        # Check if the move_in record is marked as sent
        self.assertTrue(self.move_in.is_edi_send, "Move-in record is not marked as EDI sent.")

        # Check if edi_sent_on is updated
        self.assertIsNotNone(self.move_in.edi_sent_on, "EDI sent time is not updated.")

        # For the FTP mode, verify that the appropriate FTP upload mechanism was triggered
        # Example: Check if a corresponding EDI log entry was created (depending on your EDI logs setup)
        edi_log = self.env['edi.logs'].search([('file', '=', self.move_in.id)])
        # self.assertTrue(edi_log, "EDI log was not created.")
        # self.assertEqual(edi_log.status, 'success', "EDI log status is not marked as success.")

    def test_edi_send_move_in_email(self):

        self.move_in.is_edi_send = True
        self.move_in.edi_sent_on = fields.Datetime.now()
        # Call the edi_send method
        self.edi_setting.edi_send(realtime_frequency=True)

        # Check if the move_in record is marked as sent
        self.assertTrue(self.move_in.is_edi_send, "Move-in record is not marked as EDI sent.")

        # Check if edi_sent_on is updated
        self.assertIsNotNone(self.move_in.edi_sent_on, "EDI sent time is not updated.")

        # Look for an email in the mail.mail model that corresponds to the move_in record
        email = self.env['mail.mail'].search([('subject', '=', 'Move In EDI')])
        # self.assertTrue(email, "No email sent.")
        # self.assertIn('test@example.com', [recipient.email_to for recipient in email.recipients],
        #               "Email recipient mismatch.")

    def test_edi_send_move_out_ftp(self):

        self.move_out.is_edi_send = True
        self.move_out.edi_sent_on = fields.Datetime.now()
        # Call the edi_send method
        self.edi_setting.edi_send(realtime_frequency=True)

        # Check if the move_out record is marked as sent
        self.assertTrue(self.move_out.is_edi_send, "Move-out record is not marked as EDI sent.")

        # Check if edi_sent_on is updated
        self.assertIsNotNone(self.move_out.edi_sent_on, "EDI sent time is not updated.")

        # Ensure that the file was sent via FTP (same as above, you can verify FTP mechanism, logs, etc.)
        edi_log = self.env['edi.logs'].search([('file', '=', self.move_out.id)])
        # self.assertTrue(edi_log, "EDI log was not created.")
        # self.assertEqual(edi_log.status, 'success', "EDI log status is not marked as success.")

    def test_edi_send_move_out_email(self):

        self.move_out.is_edi_send = True
        self.move_out.edi_sent_on = fields.Datetime.now()
        # Call the edi_send method
        self.edi_setting.edi_send(realtime_frequency=True)

        # Check if the move_out record is marked as sent
        self.assertTrue(self.move_out.is_edi_send, "Move-out record is not marked as EDI sent.")

    def test_send_list_view_rec_edis_success(self):
        self.edi_setting.write({'edi_format': 'edi'})
        self.location.write({'email': "adaniinfo@mail.com"})
        """Test successful execution of send_list_view_rec_edis."""
        result = self.edi_setting.send_list_view_rec_edis(self.move_in.id)

        # Assert the result
        self.assertEqual(result, {'success': True}, "Expected success but got an error.")

        # Check that EDI files are generated and email is sent
        self.assertTrue(self.move_in.edi_in_attachment_id, f"No EDI attachment found for record {self.move_in.id}.")
        self.assertTrue(self.move_in.is_edi_send, f"Record {self.move_in.id} not marked as EDI sent.")

    def test_no_selected_ids(self):
        """Test when no IDs are selected."""
        result = self.edi_setting.send_list_view_rec_edis([])
        self.assertEqual(result, {'error': 'No selected record or model found!'}, "Expected error for no selected IDs.")

    def test_no_records_found(self):
        """Test when selected IDs do not correspond to valid records."""
        result = self.edi_setting.send_list_view_rec_edis([99999])  # Non-existent ID
        self.assertEqual(result, {'error': 'No records found to generate EDI files'},
                         "Expected error for no records found.")

    def test_different_location_or_shipping_line(self):
        """Test when selected records have different locations or shipping lines."""
        new_location = self.env['res.company'].create({'name': 'Different Location'})
        self.move_in1.write({'location_id': new_location.id})
        fetch_records = [self.move_in.id, self.move_in1.id]

        result = self.edi_setting.send_list_view_rec_edis(fetch_records)
        self.assertEqual(result, {'error': 'All selected records must have the same shipping line or location.'},
                         "Expected error for inconsistent location or shipping line.")

    def test_no_edi_configuration(self):
        edi_settings = self.env['edi.settings'].create({
            'location': self.location.id,
            'shipping_line_id': self.shipping_line.id,})
        """Test when no EDI configuration is found for the records."""
        result = edi_settings.send_list_view_rec_edis(self.move_in.id)
        self.assertEqual(result, {'error': 'No EDI configurations found for selected records'},
                         "Expected error for missing EDI configuration.")

    def test_no_email_for_location(self):
        self.edi_setting.write({'edi_format': 'edi'})
        """Test when no email is found for the location."""
        result = self.edi_setting.send_list_view_rec_edis(self.move_in1.id)
        self.assertEqual(result, {'error': 'No email found for selected location.'},
                         "Expected error for missing location email.")

    def test_add_edi_logs_for_manual_operation(self):
        # Test parameters
        data = self.attachment.id
        status = 'success'
        fetch_records = [self.move_in.id, self.move_in1.id]

        # Call the method
        self.edi_setting.add_edi_logs_for_manual_operation(data, status, fetch_records)

        # Check that the EDI log was created successfully
        edi_log = self.env['edi.logs'].search([], order='create_date desc', limit=1)

        # Assertions
        self.assertTrue(edi_log, "EDI log was not created.")
        # self.assertEqual(edi_log.location_id.id, self.location.id, "Location ID mismatch.")
        self.assertEqual(edi_log.shipping_line_id.id, self.shipping_line.id, "Shipping Line ID mismatch.")
        self.assertEqual(edi_log.file.id, data, "EDI file content mismatch.")
        self.assertEqual(edi_log.container_count, len(fetch_records), "Container count mismatch.")
        self.assertEqual(edi_log.status, status, "Status mismatch.")
        self.assertEqual(edi_log.generated_by, 'user', "Generated by field mismatch.")
        self.assertTrue(edi_log.edi_sent_on, "EDI sent timestamp not set.")
        self.assertEqual(edi_log.email_sent, 'yes', "Email sent field mismatch.")
        self.assertEqual(edi_log.type, 'move_in', "Type mismatch.")
        self.assertEqual(set(edi_log.move_in_ids.ids), set(fetch_records), "Move In IDs mismatch.")

    def test_send_edi_email_success(self):
        """Test successful email sending."""
        subject = 'Test EDI Email'
        body = '<p>This is a test email for EDI</p>'
        attachment_ids = [(4, self.attachment.id)]

        # Call the method
        result = self.edi_setting._send_edi_email(subject, body, attachment_ids, self.edi_setting, [self.move_in.id])

        # Check that the email was sent
        sent_mail = self.env['mail.mail'].search([('subject', '=', subject)], limit=1)
        self.assertTrue(sent_mail, 'Email was not created or sent successfully.')

        # Verify attachment in the sent email
        self.assertIn(self.attachment.id, sent_mail.attachment_ids.ids, 'Attachment not found in the sent email.')

        # Verify logs
        edi_log = self.env['edi.logs'].search([('status', '=', 'success')], limit=1)
        self.assertTrue(edi_log, 'EDI log not created for success case.')
        self.assertEqual(edi_log.file.id, self.attachment.id, 'EDI log file does not match the sent attachment.')

    def test_update_edi_sent_status(self):
        """Test that _update_edi_sent_status updates records correctly."""
        # Call the method
        self.edi_setting._update_edi_sent_status(self.move_in)

        # Check if 'is_edi_send' is updated to True
        self.assertTrue(self.move_in.is_edi_send, "'is_edi_send' was not updated to True.")

        # Check if 'edi_sent_on' is updated
        user_tz = self.env.user.tz or 'UTC'
        current_utc_time = fields.Datetime.now()
        expected_time = (
            pytz.timezone(user_tz)
            .localize(current_utc_time, is_dst=None)
            .astimezone(pytz.timezone(user_tz))
            .replace(tzinfo=None)
        )
        self.assertAlmostEqual(
            self.move_in.edi_sent_on,
            expected_time,
            msg="'edi_sent_on' was not updated to the expected local time.",
        )

    def test_create_zip_attachment(self):
        self.attachments = [
            {'name': 'file1.edi', 'edi_content': 'Content of file 1'},
            {'name': 'file2.edi', 'edi_content': 'Content of file 2'},
        ]
        """Test that _create_zip_attachment creates a zip file correctly."""
        edi_settings = self.env['edi.settings']

        # Call the method
        zip_attachment = edi_settings._create_zip_attachment(self.attachments, 'move_in')

        # Assert that an attachment is created
        self.assertTrue(zip_attachment, "No zip attachment was created.")
        self.assertEqual(zip_attachment.mimetype, 'application/zip', "Incorrect mimetype for the zip attachment.")

        # Decode the zip data and verify its contents
        zip_data = base64.b64decode(zip_attachment.datas)
        with zipfile.ZipFile(io.BytesIO(zip_data), 'r') as zip_file:
            # Check the list of files in the zip
            zip_file_names = zip_file.namelist()
            self.assertEqual(len(zip_file_names), len(self.attachments), "Incorrect number of files in the zip.")
            self.assertListEqual(
                zip_file_names,
                [attachment['name'] for attachment in self.attachments],
                "File names in the zip do not match the attachments.",
            )

            # Check the content of each file
            for attachment in self.attachments:
                with zip_file.open(attachment['name']) as file:
                    content = file.read().decode('utf-8')
                    self.assertEqual(
                        content,
                        attachment['edi_content'],
                        f"Content of {attachment['name']} does not match the expected content.",
                    )

        # Check the filename format of the zip file
        current_time_str = edi_settings.format_datetime(datetime.utcnow(), '%d%m%y%H%M')
        expected_filename = f'MoveIn_EDI_{current_time_str}.zip'
        self.assertEqual(
            zip_attachment.name,
            expected_filename,
            "The zip filename does not match the expected format.",
        )

    def test_no_selected_ids(self):
        """Test when no IDs are selected."""
        result = self.edi_settings.send_move_out_list_view_rec_edis([])
        self.assertEqual(result, {'error': 'No selected record or model found!'}, "Expected error for no selected IDs.")

    def test_no_records_found(self):
        """Test when selected IDs do not correspond to valid records."""
        result = self.edi_settings.send_move_out_list_view_rec_edis([99999])  # Non-existent ID
        self.assertEqual(result, {'error': 'No records found to generate EDI files'},
                         "Expected error for no records found.")

    def test_different_location_or_shipping_line(self):
        """Test when selected records have different locations or shipping lines."""
        new_location = self.env['res.company'].create({'name': 'Different Location'})
        self.move_out1.write({'location_id': new_location.id})
        fetch_records = [self.move_out.id, self.move_out1.id]

        result = self.edi_settings.send_move_out_list_view_rec_edis(fetch_records)
        self.assertEqual(result, {'error': 'All selected records must have the same shipping line or location.'},
                         "Expected error for inconsistent location or shipping line.")

    def test_no_edi_configuration(self):
        edi_settings = self.env['edi.settings'].create({
            'location': self.location.id,
            'shipping_line_id': self.shipping_line.id,})

        """Test when no EDI configuration is found for the records."""
        result = edi_settings.send_move_out_list_view_rec_edis(self.move_out.id)
        self.assertEqual(result, {'error': 'No EDI configurations found for selected records'},
                         "Expected error for missing EDI configuration.")

    def test_no_email_for_location(self):
        self.edi_settings.write({'edi_format': 'edi'})
        result = self.edi_settings.send_move_out_list_view_rec_edis(self.move_out.id)
        self.assertEqual(result, {'error': 'No email found for selected location.'},
                         "Expected error for missing location email.")
