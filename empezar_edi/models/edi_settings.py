# -*- coding: utf-8 -*-
import paramiko
import pytz
import zipfile
import logging
import re
import io
import base64
import openpyxl
from io import BytesIO
from odoo import fields, models, _, api
from odoo.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from ...empezar_base.models.res_users import ResUsers
from datetime import datetime, timedelta
from  ...empezar_base.models.contrainer_type_edi import ContainerTypeEdi

_logger = logging.getLogger(__name__)


class EdiSetting(models.Model):
    _name = "edi.settings"
    _rec_name = "edi_type"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = "Edi Settings"

    def get_current_month_dates(self):
        """
        Generate a list of tuples containing Dates Of Current Month.

        Returns:
            list: A list of tuples where each tuple contains a Date as a string in the format (date, date).

        Example:
            >> get_current_month_dates
            [('1', '1'), ('2', '2'), ..., ('31', '31')]
        """
        today = datetime.today()
        first_day_of_month = today.replace(day=1)

        # Calculate the first day of the next month
        if today.month == 12:
            next_month = first_day_of_month.replace(year=today.year + 1, month=1)
        else:
            next_month = first_day_of_month.replace(month=today.month + 1)

        # Calculate the number of days in the current month
        num_days = (next_month - first_day_of_month).days

        date_list = []
        for day in range(1, num_days + 1):
            date_str = str(day)
            date_list.append((date_str, date_str))

        return date_list

    shipping_line_id = fields.Many2one('res.partner', domain="[('is_shipping_line', '=', True)]",
                                       string="Shipping Line", tracking=True)
    location_domain = fields.Many2many(
        "res.company", compute="_compute_location_domain"
    )
    location = fields.Many2one('res.company', string="Location", domain="[('parent_id','!=', False)]", tracking=True)
    edi_type = fields.Selection([
        ('move_in', 'Move In'),
        ('move_out', 'Move Out'),
        ('repair', 'Repair'),
        ('damaged', 'Damage'),
    ], string="EDI Type", tracking=True)
    shipping_line_ofc_code = fields.Char(string='Shipping Line Office Code', size=10)
    frequency = fields.Selection([
        ('hourly', 'Hourly'),
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('real_time', 'Real Time'),
    ], string="Frequency", tracking=True)
    freqency_hourly = fields.Float('Frequency Setting (Hourly)')
    freqency_daily = fields.Float('Frequency Setting (Daily)')
    frequency_weekly = fields.Selection([
        ('sunday', 'Sunday'),
        ('monday', 'Monday'),
        ('tuesday', 'Tuesday'),
        ('wednesday', 'Wednesday'),
        ('thursday', 'Thursday'),
        ('friday', 'Friday'),
        ('saturday', 'Saturday'),
    ], string="Frequency Week", default="sunday")
    freqency_time = fields.Float('Frequency Time')
    freqency_monthly = fields.Float('Frequency Setting (Monthly)')
    freqency_real_time = fields.Datetime('Frequency Setting (Real Time)')
    mode = fields.Selection([('email', 'Email'), ('ftp', 'FTP')])
    email_to = fields.Char(string="Email (To)")
    email_from = fields.Char(string="Email (From)")
    email_cc = fields.Char(string="Email (CC)")
    email_bcc = fields.Char(string="Email (BCC)")
    ftp_location = fields.Char(string="FTP Location")
    ftp_username = fields.Char(string="FTP Username")
    ftp_password = fields.Char(string="FTP Password")
    port_number = fields.Integer(size='10', string="Port Number")
    secure_connection = fields.Selection([('yes', 'Yes'),
                                          ('no', 'No')], string="Secure Connection")
    ftp_folder = fields.Char(string="FTP Folder Name")
    edi_format = fields.Selection([('excel', 'Excel'),
                                   ('edi', 'EDI')], string="EDI Format")
    header_edi = fields.Text(string='Header')
    body_edi = fields.Text(string='Body')
    footer_edi = fields.Text(string='Footer')
    file_attched = fields.Binary(String="Upload File")
    file_name = fields.Char(string="File-Name")
    display_create_info = fields.Char(compute="_get_create_record_info")
    display_modified_info = fields.Char(compute="_get_modify_record_info")
    last_run = fields.Datetime(string="Last Run")
    next_scheduled = fields.Datetime(string="Next Scheduled Run", compute="_compute_next_scheduled", store=True)
    attchment_file_name = fields.Char(string="File Name")
    freqency_monthly_date = fields.Selection(get_current_month_dates)
    company_id = fields.Many2one(
        "res.company",
        required=True,
        readonly=True,
        default=lambda self: self.env.company,
    )
    active = fields.Boolean("Active", default=True)
    rec_status = fields.Selection(
        [
            ("active", "Active"),
            ("disable", "Disable"),
        ],
        default="active",
        compute="_check_active_records",
        string="Status",
    )
    is_for_msc_shipping_line = fields.Boolean(string="Is For MSC In Shipping Line ?")
    is_for_msc_south_shipping_line = fields.Boolean(string="Is For MSC In South Shipping Line ?")

    @api.onchange('shipping_line_id')
    @api.depends("location")
    def _compute_location_domain(self):
        for record in self:
            company = self.env['res.company'].search([('parent_id', '!=', False)])
            if record.shipping_line_id:
                locations = []
                for rec in company:
                    shipping_line_ids = rec.shipping_line_mapping_ids.mapped(
                        "shipping_line_id"
                    ).ids
                    if record.shipping_line_id.id in shipping_line_ids:
                        locations.append(rec.id)
                record.location_domain = locations
            else:
                record.location_domain = []

    @api.model
    def create(self, vals):
        try:
            vals["last_run"] = fields.Datetime.now()  # Store current datetime
            vals['freqency_real_time'] = vals.get('freqency_real_time', fields.Datetime.now())
        except Exception as e:
            print(f"An error occurred during record creation: {e}")
        return super(EdiSetting, self).create(vals)

    def write(self, vals):
        try:
            vals['freqency_real_time'] = vals.get('freqency_real_time', fields.Datetime.now())
        except Exception as e:
            print(f"An error occurred during record creation: {e}")
        return super(EdiSetting, self).write(vals)

    def float_to_hhmm(self, float_time):
        """
        Convert a float representing time (hours) into a string formatted as HH:MM.

        Args:
            float_time (float): Time in float format (e.g., 3.5 for 3 hours and 30 minutes).

        Returns:
            str: Time formatted as HH:MM (e.g., '03:30').
        """
        hours = int(float_time)
        minutes = int((float_time - hours) * 60)
        return "{:02d}:{:02d}".format(hours, minutes)

    @api.depends('last_run', 'frequency_weekly', 'frequency', 'freqency_monthly_date', 'freqency_hourly',
                  'freqency_daily', 'freqency_time', 'freqency_monthly')
    def _compute_next_scheduled(self):
        """
        Compute the next scheduled datetime for the record based on the last run time
        and frequency . The next scheduled time is calculated differently
        depending on whether the frequency is set to hourly, daily, weekly, or monthly.

        - For 'hourly' frequency, the next scheduled time is the last run time plus
          the number of specified hours.
        - For 'daily' frequency, the next scheduled time is the next day at the time.
        - For 'weekly' frequency, the next scheduled time is the next occurrence of
          the specified weekday at the specified time`.
        - For 'monthly' frequency, the next scheduled time is the next occurrence of
          the specified day of the month and specified time.
        """
        for record in self:
            if record.last_run != False:
                last_run = fields.Datetime.from_string(record.last_run)
                user_tz = self.env.user.tz or 'UTC'
                last_run_user_tz = last_run.astimezone(pytz.timezone(user_tz))

                if record.frequency == 'hourly':
                    next_run = last_run + timedelta(hours=record.freqency_hourly)
                    next_run = next_run.replace(second=0, microsecond=0)
                    record.next_scheduled = next_run

                elif record.frequency == 'daily':
                    next_run = last_run_user_tz + timedelta(days=1)
                    hhmm = record.float_to_hhmm(record.freqency_daily)
                    hours, minutes = map(int, hhmm.split(':'))
                    next_run = next_run.replace(hour=hours, minute=minutes, second=0, microsecond=0)
                    next_run_utc = next_run.astimezone(pytz.utc)
                    naive_next_run = next_run_utc.replace(tzinfo=None)
                    record.next_scheduled = naive_next_run

                elif record.frequency == 'weekly':
                    target_weekday_str = record.frequency_weekly.lower()[:3]
                    target_weekday = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"].index(target_weekday_str)
                    current_weekday = last_run_user_tz.weekday()
                    days_until_target = (target_weekday - current_weekday) % 7
                    if days_until_target == 0:
                        days_until_target = 7
                    next_run = last_run_user_tz + timedelta(days=days_until_target)
                    hhmm = record.float_to_hhmm(record.freqency_time)
                    hours, minutes = map(int, hhmm.split(':'))
                    next_run = next_run.replace(hour=hours, minute=minutes, second=0, microsecond=0)
                    next_run_utc = next_run.astimezone(pytz.utc)
                    naive_next_run = next_run_utc.replace(tzinfo=None)
                    record.next_scheduled = naive_next_run

                elif record.frequency == 'monthly':
                    current_day = last_run_user_tz.day
                    target_day = int(record.freqency_monthly_date) if record.freqency_monthly_date else 1
                    record.freqency_monthly_date = record.freqency_monthly_date if record.freqency_monthly_date else '1'
                    if target_day > current_day:
                        next_run = last_run_user_tz.replace(day=target_day)
                        hhmm = record.float_to_hhmm(record.freqency_monthly)
                        hours, minutes = map(int, hhmm.split(':'))
                        next_run = next_run.replace(hour=hours, minute=minutes, second=0, microsecond=0)
                        next_run_utc = next_run.astimezone(pytz.utc)
                        naive_next_run = next_run_utc.replace(tzinfo=None)
                        record.next_scheduled = naive_next_run
                    else:
                        next_run = last_run_user_tz + relativedelta(months=1)
                        next_run = next_run.replace(day=target_day)
                        hhmm = record.float_to_hhmm(record.freqency_monthly)
                        hours, minutes = map(int, hhmm.split(':'))
                        next_run = next_run.replace(hour=hours, minute=minutes, second=0, microsecond=0)
                        next_run_utc = next_run.astimezone(pytz.utc)
                        naive_next_run = next_run_utc.replace(tzinfo=None)
                        record.next_scheduled = naive_next_run
                else:
                    record.next_scheduled = None

            else:
                record.next_scheduled = None

    @api.onchange('freqency_hourly', 'freqency_daily', 'freqency_time', 'freqency_monthly')
    @api.constrains('freqency_hourly', 'freqency_daily', 'freqency_time', 'freqency_monthly')
    def _check_freqency_hourly(self):
        """
        Validates the frequency fields to ensure they are within the allowed range of 0 to 24
        """
        if self.frequency == 'hourly':
            if self.freqency_hourly <= 0 or self.freqency_hourly > 24:
                raise ValidationError(_("Hourly frequency should be between 1 to 24."))
        if self.frequency == 'daily':
            if self.freqency_daily < 0 or self.freqency_daily > 24:
                raise ValidationError(_("Daily frequency should be between 0 to 24."))
        if self.frequency == 'weekly':
            if self.freqency_time < 0 or self.freqency_time > 24:
                raise ValidationError(_("Weekly frequency should be between 0 to 24."))
        if self.frequency == 'monthly':
            if self.freqency_monthly < 0 or self.freqency_monthly > 24:
                raise ValidationError(_("Monthly frequency should be between 0 to 24."))

    @api.onchange('file_attched')
    @api.constrains('file_attched')
    def _check_file_attched_validation(self):
        """
        ValidationError: If the file name does not end with .xls or .xlsx.
        ValidationError: If the file size exceeds 5MB.
        """
        if self.file_name and not self.file_name.lower().endswith(('.xls', '.xlsx')):
            raise ValidationError("Only .xls and .xlsx files are allowed.")
        if self.file_attched and len(self.file_attched) > 5242880:  # 5MB in bytes
            raise ValidationError("File size cannot exceed 5MB.")

    @api.constrains('shipping_line_id', 'location', 'edi_type', 'edi_format')
    def uniq_edi_configuration(self):
        """
            ValidationError: If a record with the same combination of Shipping Line, Location, 
            EDI Type, and EDI Format already exists.
        """
        for rec in self:
            if rec.shipping_line_id.active != True:
                raise ValidationError(_("Shipping Line Selected is Disabled"))
            if rec.location.active != True:
                raise ValidationError(_("Shipping Line Selected is Disabled"))
            existing_record = rec.env['edi.settings'].search([('id', '!=', rec.id)])
            for record in existing_record:
                if record.shipping_line_id.id == rec.shipping_line_id.id and record.location.id == rec.location.id and record.edi_type == rec.edi_type and record.edi_format == rec.edi_format:
                    raise ValidationError(
                        _("EDI configuration for the Shipping Line, Location, EDI Type, and EDI Format already exists!"))

    @api.constrains('email_from', 'email_to', 'email_cc', 'email_bcc')
    def is_valid_email(self):
        """
        This function checks if the provided email addresses are valid using a regular expression.
        It validates multiple email addresses separated by commas.

        Returns:
            True if all emails are valid, raises ValidationError if any email is invalid.
        """
        email_regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'

        def validate_emails(email_string):
            if not email_string:
                return True  # If empty, consider valid
            emails = [email.strip() for email in email_string.split(',') if email.strip()]
            for email in emails:
                if not re.fullmatch(email_regex, email):
                    return False
            return True

        if not validate_emails(self.email_to):
            raise ValidationError(_("Invalid Email (To) address. Please enter correct email address."))
        if self.email_from and not bool(re.fullmatch(email_regex, self.email_from)):
            raise ValidationError(_("Invalid Email (From) address. Please enter correct email address."))
        if not validate_emails(self.email_cc):
            raise ValidationError(_("Invalid Email (CC) address. Please enter correct email address."))
        if not validate_emails(self.email_bcc):
            raise ValidationError(_("Invalid Email (BCC) address. Please enter correct email address."))

    def _get_create_record_info(self):
        """
        Assign create record log string to the appropriate field.
        :return: none
        """
        for rec in self:
            if rec.env.user.tz and rec.create_uid:
                tz_create_date = ResUsers.convert_datetime_to_user_timezone(
                    rec.env.user, rec.create_date
                )
                create_uid_name = rec.create_uid.name
                if tz_create_date:
                    rec.display_create_info = ResUsers.get_user_log_data(
                        rec, tz_create_date, create_uid_name
                    )
            else:
                rec.display_create_info = ""

    def _check_active_records(self):
        ContainerTypeEdi.check_active_records(self)

    def _get_modify_record_info(self):
        """
        Assign update record log string to the appropriate field.
        :return: none
        """
        for rec in self:
            if rec.env.user.tz and rec.write_uid:
                tz_write_date = ResUsers.convert_datetime_to_user_timezone(
                    rec.env.user, rec.write_date
                )
                write_uid_name = rec.write_uid.name
                if tz_write_date:
                    rec.display_modified_info = ResUsers.get_user_log_data(
                        rec, tz_write_date, write_uid_name
                    )
            else:
                rec.display_modified_info = ""

    @api.constrains('frequency')
    def check_real_time_frequency(self):
        try:
            if self.frequency == 'real_time':
                self.edi_send('real_time_frequency')
        except Exception as e:
            _logger.info("There is some issue while sending edi from real time",e)

    # Helper function to convert datetime to user timezone and format it
    def format_datetime(self, dt, format_str='%d%m%y%H%M'):
        user_tz = self.env.user.tz or 'UTC'
        user_timezone = pytz.timezone(user_tz)
        user_time = dt.astimezone(user_timezone)
        return user_time.strftime(format_str)

    # Placeholder mapping for dynamic data
    def get_dynamic_value(self, placeholder_name, rec, container_count=None):
        """
        Fetches the value of a dynamic placeholder from the record.
        Handles special placeholders or maps fields correctly.
        """
        # Define dictionary to map placeholders to functions or lambda expressions
        placeholder_map = {
            'carrier_location_code': lambda: rec.location_id.location_code or 'UNKNOWN',
            'd/t-ddmmyy': lambda: self.format_datetime(datetime.utcnow(), '%d%m%y%H%M'),
            'd/t-ddmmyyyyhhmm': lambda: self.format_datetime(datetime.utcnow(), '%d%m%Y%H%M'),
            'd/t-mmddyy': lambda: self.format_datetime(datetime.utcnow(), '%m%d%y%H%M'),
            'd/t-yymmdd': lambda: self.format_datetime(datetime.utcnow(), '%y%m%d:%H%M'),
            'd-yyyymmdd': lambda: self.format_datetime(datetime.utcnow(), '%Y%m%d'),
            'd/t-yyyymmdd': lambda: (
                self.format_datetime(rec.move_in_date_time, '%Y%m%d%H%M') if self.edi_type == 'move_in' else
                self.format_datetime(rec.move_out_date_time, '%Y%m%d%H%M') if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'd/t-yyyymmddhhmmss': lambda: (
                self.format_datetime(rec.move_in_date_time, '%Y%m%d%H%M%S') if self.edi_type == 'move_in' else
                self.format_datetime(rec.move_out_date_time, '%Y%m%d%H%M%S') if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'd/t-ddmmyyyyhh:mm': lambda: (
                self.format_datetime(rec.move_in_date_time, '%d%m%Y%H:%M') if self.edi_type == 'move_in' else
                self.format_datetime(rec.move_out_date_time, '%d%m%Y%H:%M') if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'd-ddmmyyyy': lambda: (
                self.format_datetime(rec.move_in_date_time, '%d%m%Y') if self.edi_type == 'move_in' else
                self.format_datetime(rec.move_out_date_time, '%d%m%Y') if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'tare_weight': lambda: rec.tare_wt or 'UNKNOWN',
            'gross_weight': lambda: rec.gross_wt or 'UNKNOWN',
            'carrier_office_code': lambda : self.shipping_line_ofc_code or 'UNKNOWN',
            'truck_number': lambda: (
                rec.truck_no if self.edi_type == 'move_in' else
                rec.truck_number if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'customer_name': lambda: (
                rec.parties_importer.name if self.edi_type == 'move_in' and rec.parties_importer else
                rec.exporter_name_id.name if self.edi_type == 'move_out' and rec.exporter_name_id else
                'UNKNOWN'
            ),
            'container_no': lambda: (
                rec.container if self.edi_type == 'move_in' else
                rec.inventory_id.name if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'cfs_code': lambda: (
                rec.from_cfs_icd.code if self.edi_type == 'move_in' and rec.from_cfs_icd else
                rec.to_cfs_icd_repo_id.code if self.edi_type == 'move_out' and rec.to_cfs_icd_repo_id else
                'UNKNOWN'
            ),
            'transporter_name': lambda: (
                rec.transporter_allotment_id.name if self.edi_type == 'move_in' and rec.transporter_allotment_id else
                rec.transporter_allocated_id.name if self.edi_type == 'move_out' and rec.transporter_allocated_id else
                'UNKNOWN'
            ),
            'movement_type_whl': lambda: (
                'MA' if self.edi_type == 'move_in' else
                'MI' if self.edi_type == 'move_out' and rec.movement_type == 'repo' and rec.repo_to == 'port_terminal' else
                'MS'
            ),
            'movement_type_ts': lambda: (
                'MIR' if self.edi_type == 'move_in' and rec.movement_type == 'repo' and rec.repo_from == 'port_terminal' else
                'DEV' if self.edi_type == 'move_in' and rec.movement_type == 'import_destuffing' and rec.import_destuffing_from == 'factory' else
                'RET' if self.edi_type == 'move_in' and rec.movement_type == 'factory_return' else
                'MRO' if self.edi_type == 'move_out' and rec.movement_type == 'repo' and rec.repo_to == 'port_terminal' else
                'FST' if self.edi_type == 'move_out' and rec.movement_type == 'export_stuffing' and rec.export_stuffing_to == 'factory' else
                'NA'
            ),
            'movement_type_msc': lambda: (
                'MIR' if self.edi_type == 'move_in' and rec.movement_type == 'repo' and rec.repo_from == 'port_terminal' else
                'MTIN' if self.edi_type == 'move_in' and rec.movement_type == 'import_destuffing' and rec.import_destuffing_from == 'factory' else
                'RET' if self.edi_type == 'move_in' and rec.movement_type == 'factory_return' else
                # 'MRO' if self.edi_type == 'move_out' and rec.movement_type == 'repo' and rec.repo_to == 'port_terminal' else
                # 'FST' if self.edi_type == 'move_out' and rec.movement_type == 'export_stuffing' and rec.export_stuffing_to == 'factory' else
                'NA'
            ),
            'seal_number': lambda: (
                rec.seal_no_1 if self.edi_type == 'move_in' else
                rec.seal_no_1.seal_number if self.edi_type == 'move_out' else
                'UNKNOWN'
            ),
            'vent_seal': lambda: (
                rec.vent_seal_no if self.edi_type == 'move_out' else
                ''
            ),
            'booking_number': lambda: rec.booking_no_id.booking_no or 'UNKNOWN',
            'edi_code': lambda: (
                    rec.shipping_line_id.container_type_edi_ids.filtered(
                        lambda a: a.container_type_data_id == rec.type_size_id
                    )[:1].edi_code or 'UNKNOWN'
            ),
            'transporter_code': lambda: (
                    rec.shipping_line_id.shipping_line_transporters_ids.filtered(
                        lambda a: a.transporter_id == rec.transporter_allotment_id
                    )[:1].code or 'NA'
            ),
            'damage_type': lambda : rec.damage_condition.name or 'NA',
            'damage_type_msc': lambda: (
                'H' if rec.damage_condition.name == 'High' else
                'M' if rec.damage_condition.name == 'Medium' else
                'L' if rec.damage_condition.name == 'Low' else
                '' if rec.damage_condition.name == 'No Damage' else
                'N' if rec.damage_condition.name == 'Normal Wash' else
                'NA'
            ),
            'edi_remarks': lambda : rec.remarks or 'UNKNOWN',
            'container_count': lambda : container_count,
            'reference_number1': lambda: '{reference_number1}',
            'reference_number2': lambda : '{reference_number2}',
            'port_code': lambda: (
                rec.from_port.port_code if self.edi_type == 'move_in' and rec.from_port else
                rec.to_port_id.port_code if self.edi_type == 'move_out' and rec.to_port_id else
                'UNKNOWN'
            ),
        }

        # Get and execute the appropriate function or return 'UNKNOWN' if not found
        return placeholder_map.get(placeholder_name.lower(), lambda: 'UNKNOWN')()

    def replace_placeholders(self, text, rec, container_count):
        """
        Replace dynamic placeholders in the given text using the record's data.
        """
        # Find all placeholders in the format {placeholder}
        placeholders = re.findall(r'{(.*?)}', text)

        # Replace each placeholder with the corresponding record value
        for placeholder in placeholders:
            value = self.get_dynamic_value(placeholder, rec, container_count)
            text = text.replace(f'{{{placeholder}}}', str(value))  # Replace with actual value

        return text.strip()  # No need to strip HTML now

    def generate_edi_file(self, records):
        """
        Generates an EDI file based on Header, Body, Footer formats and
        dynamically replaces placeholders with record data.
        """
        if self and records:
            # Extract EDI format parts
            header_text = self.header_edi
            body_text = self.body_edi
            footer_text = self.footer_edi
            container_count = len(records)
            file_name = self.attchment_file_name
            new_file_name = self.replace_placeholders(file_name, records[0], container_count)

            # Prepare the header section (static, appears once)
            if header_text:
                header_formatted = self.replace_placeholders(header_text, records[0], container_count)
            else:
                header_formatted = ''

            # Initialize the body content, which will be repeated for each record
            body_content = ''
            last_generated_reference_number = ''
            for rec in records:
                modified_body_text = body_text
                lines = body_text.splitlines()
                # If neither is_for_msc_shipping_line nor is_for_msc_south_shipping_line is True, skip modification
                if not self.is_for_msc_shipping_line and not self.is_for_msc_south_shipping_line:
                    modified_body_text = body_text
                else:
                    # Logic for factory return for MSC shipping line
                    if self.is_for_msc_shipping_line and rec.movement_type == 'factory_return':
                        if lines:
                            modified_body_text = '\n'.join(lines[:1])  # Include only Line 1

                    # Logic for factory return for MSC South shipping line
                    elif self.is_for_msc_south_shipping_line and rec.movement_type == 'factory_return':
                        if lines:
                            modified_body_text = '\n'.join(lines[1:2])  # Include only Line 2

                    # Logic for non-factory return for MSC shipping line
                    elif self.is_for_msc_south_shipping_line and rec.movement_type != 'factory_return':
                        # Add only Line 1 if the condition is met for import_destuffing or repo
                        if (rec.movement_type == 'import_destuffing' and rec.import_destuffing_from == 'CFS/ICD') or \
                                (rec.movement_type == 'repo' and rec.repo_from == 'CFS/ICD'):
                            modified_body_text = '\n'.join([lines[0]] + lines[2:])  # add Line 1,3,4
                        else:
                            # Only add Lines 3 and 4
                            modified_body_text = '\n'.join(lines[2:])
                # Dynamically replace placeholders for each record
                body_formatted = self.replace_placeholders(modified_body_text, rec, container_count)
                reference_number_2 = re.findall(r'{reference_number2}', body_formatted)
                if reference_number_2:
                    ref_number_2 = self.env['ir.sequence'].next_by_code('edi.sequence.number.2')
                    body_formatted = body_formatted.replace('{reference_number2}',ref_number_2)
                    last_generated_reference_number = ref_number_2
                body_content += body_formatted + '\n'

            # Prepare the footer section (static, appears once)
            if footer_text:
                footer_formatted = self.replace_placeholders(footer_text, records[0], container_count)
            else:
                footer_formatted = ''

            # Check if the "Reference Number 2" placeholder exists in any part
            if ('{reference_number2}' in header_formatted or '{reference_number2}' in footer_formatted or
                    '{reference_number2}' in new_file_name):
                # Replace the "Reference Number 1" placeholder in all relevant sections
                header_formatted = header_formatted.replace('{reference_number2}', last_generated_reference_number)
                footer_formatted = footer_formatted.replace('{reference_number2}', last_generated_reference_number)
                new_file_name = new_file_name.replace('{reference_number2}', last_generated_reference_number)

            # Check if the "Reference Number 1" placeholder exists in any part
            if ('{reference_number1}' in header_formatted or '{reference_number1}' in body_content or '{reference_number1}' in footer_formatted or
                    '{reference_number1}' in new_file_name):
                # Generate the sequence number only once if needed
                ref_number_1 = self.env['ir.sequence'].next_by_code('edi.sequence.number.1')

                # Replace the "Reference Number 1" placeholder in all relevant sections
                header_formatted = header_formatted.replace('{reference_number1}', ref_number_1)
                body_content = body_content.replace('{reference_number1}', ref_number_1)
                footer_formatted = footer_formatted.replace('{reference_number1}', ref_number_1)
                new_file_name = new_file_name.replace('{reference_number1}', ref_number_1)

            # Combine all parts into the final EDI content
            if header_formatted:
                edi_content = f"{header_formatted}\n{body_content}{footer_formatted}"
            else:
                edi_content = f"{body_content}{footer_formatted}"
            # Remove any existing attachments with the same name
            existing_attachments = self.env['ir.attachment'].sudo().search([('name', '=', file_name)])
            for attach in existing_attachments:
                attach.sudo().unlink()

            # Save the EDI content as an attachment
            attachment = self.env['ir.attachment'].create({
                'name': new_file_name,
                'datas': base64.b64encode(edi_content.encode('utf-8')),
                'public': True,
                'type': 'binary',
                'mimetype': 'text/plain',  # Ensure this is set to plain text
            })
            if attachment and self.edi_type == 'move_in':
                for rec in records:
                    rec.edi_in_attachment_id = attachment.id
            elif attachment and self.edi_type == 'move_out':
                for rec in records:
                    rec.edi_out_attachment_id = attachment.id
            return attachment
        return False

    def generate_excel_file_msc_move_in_template(self, fetch_records):
        """
        Generates an Excel file based on the record data and the placeholder template.
        Each record generates 4 lines for 'factory_return' and 3 lines for other movement types.
        """
        # Load the template
        template_attachment = self.file_attched
        if not template_attachment:
            raise ValueError("EDI Template Excel file not found in edi.settings.")

        # Decode and load workbook
        excel_data = base64.b64decode(template_attachment)
        workbook = openpyxl.load_workbook(BytesIO(excel_data))
        worksheet = workbook.active

        # Get company profile name
        company_profile_name = self.env['res.company'].search([('parent_id', '=', False)], limit=1).name or 'UNKNOWN'

        # Get the number of rows
        row_count = worksheet.max_row

        if row_count == 5:
            # Define statuses
            factory_return_statuses = ["ERM"]
            regular_statuses = ["MCY", "DAM", "TBR"]
        else:
            # Define statuses
            factory_return_statuses = ["ERM"]
            regular_statuses = ["ICO","MCY", "DAM", "TBR"]

        # If no records, return an empty template
        if not fetch_records:
            return None

        # Fetch depot_code and laden_status for the first record to use in the filename
        first_record = fetch_records[0]
        depot_code_first = (
                first_record.location_id.shipping_line_mapping_ids.filtered(
                    lambda a: a.shipping_line_id == first_record.shipping_line_id
                )[:1].depot_code or 'UNKNOWN'
        )

        # Start filling data from the second row
        filled_rows = []
        for record in fetch_records:
            # Pre-fetch depot code, laden status, and full/empty flag outside loops
            laden_status = record.location_id.laden_status_ids.mapped('name')
            depot_code = (
                    record.location_id.shipping_line_mapping_ids.filtered(
                        lambda a: a.shipping_line_id == record.shipping_line_id
                    )[:1].depot_code or 'UNKNOWN'
            )
            full_or_empty = 'F' if 'Laden' in laden_status else 'E' if 'Empty' in laden_status else 'UNKNOWN'

            # Set status list based on movement type
            status_list = factory_return_statuses if record.movement_type == 'factory_return' else regular_statuses

            # Append row data, adjusting time for "DAM" and "TBR" statuses
            for status in status_list:
                # Adjust move_in_date_time based on status
                adjusted_move_in_date_time = record.move_in_date_time
                if status == "DAM":
                    adjusted_move_in_date_time += timedelta(minutes=15)
                elif status == "TBR":
                    adjusted_move_in_date_time += timedelta(minutes=30)
                elif status == "ICO":
                    # Subtract 3 hours for "IOC" status
                    adjusted_move_in_date_time -= timedelta(hours=3)

                filled_rows.append([
                    company_profile_name, depot_code, '', record.location_id.port.port_code or 'UNKNOWN', status,
                    full_or_empty, record.container or 'UNKNOWN', self.format_datetime(adjusted_move_in_date_time,
                                                                                       '%d/%m/%Y %H:%M') or 'UNKNOWN',
                    '', '',
                    '', '', '', '', record.gross_wt or 0, '', '', '', '', '', '', '', '', '', '',
                    '', '', '', '','','',''
                ])

        # Clear existing rows from the second row
        worksheet.delete_rows(2, worksheet.max_row)

        # Write filled rows to worksheet
        for row_data in filled_rows:
            worksheet.append(row_data)

        # Save the workbook to a buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Generate the file name
        current_datetime = self.format_datetime(datetime.now(), '%d%m%Y%H%M')
        file_name = f"{first_record.location_id.port.port_code}_{depot_code_first}_{company_profile_name}INEDI{current_datetime}.xlsx"

        # Remove existing attachments
        self.env['ir.attachment'].sudo().search([('name', '=', file_name)]).unlink()

        # Save as attachment
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'datas': base64.b64encode(output.read()),
            'public': True,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        for rec in fetch_records:
            rec.edi_in_attachment_id = attachment.id

        return attachment

    def generate_excel_file_msc_move_out_template(self, fetch_records):
        """
        Generates an Excel file based on the record data and the placeholder template.
        Each record generates 4 lines for 'factory_return' and 3 lines for other movement types.
        """
        # Load the template
        template_attachment = self.file_attched
        if not template_attachment:
            raise ValueError("EDI Template Excel file not found in edi.settings.")

        # Decode and load workbook
        excel_data = base64.b64decode(template_attachment)
        workbook = openpyxl.load_workbook(BytesIO(excel_data))
        worksheet = workbook.active

        # Get company profile name
        company_profile_name = self.env['res.company'].search([('parent_id', '=', False)], limit=1).name or 'UNKNOWN'

        regular_statuses = ["REP", "MSH"]

        # If no records, return an empty template
        if not fetch_records:
            return None

        # Fetch depot_code and laden_status for the first record to use in the filename
        first_record = fetch_records[0]
        depot_code_first = (
                first_record.location_id.shipping_line_mapping_ids.filtered(
                    lambda a: a.shipping_line_id == first_record.shipping_line_id
                )[:1].depot_code or 'UNKNOWN'
        )

        # Start filling data from the second row
        filled_rows = []
        for record in fetch_records:
            # Pre-fetch depot code, laden status, and full/empty flag outside loops
            laden_status = record.location_id.laden_status_ids.mapped('name')
            depot_code = (
                    record.location_id.shipping_line_mapping_ids.filtered(
                        lambda a: a.shipping_line_id == record.shipping_line_id
                    )[:1].depot_code or 'UNKNOWN'
            )
            full_or_empty = 'F' if 'Laden' in laden_status else 'E' if 'Empty' in laden_status else 'UNKNOWN'
            booking_number = 'NA'
            if record.movement_type == 'repo' and record.booking_no_id:
                booking_number = record.booking_no_id.booking_no
            truck_number = 'NA'
            if record.mode == 'truck' and record.truck_number:
                truck_number = record.truck_number

            # Append row data, adjusting time for "DAM" and "TBR" statuses
            for status in regular_statuses:
                # Adjust move_in_date_time based on status
                adjusted_move_out_date_time = record.move_out_date_time
                if status == "REP":
                    adjusted_move_out_date_time -= timedelta(minutes=15)
                    filled_rows.append([
                        company_profile_name, depot_code, '', record.location_id.port.port_code or 'UNKNOWN', status,
                        full_or_empty, record.inventory_id.name or 'UNKNOWN', self.format_datetime(adjusted_move_out_date_time,
                                                                                           '%d/%m/%Y %H:%M') or 'UNKNOWN',
                        '', '',
                        '', '', '', '', record.gross_wt or 0, '', '', '', '', '', '', '', '', '', '',
                        '', '', '', '','','',''
                    ])
                else:
                    filled_rows.append([
                        company_profile_name, depot_code, '', record.location_id.port.port_code or 'UNKNOWN', status,
                        full_or_empty, record.inventory_id.name or 'UNKNOWN', self.format_datetime(adjusted_move_out_date_time,
                                                                                           '%d/%m/%Y %H:%M') or 'UNKNOWN',
                        booking_number, '',
                        '', '', '', '', record.gross_wt or 0, 'L', record.seal_no_1.seal_number or 'NA', '', '', '', '', '', '', '', '','','','',
                        record.transporter_allocated_id.name or 'NA', truck_number, '', ''
                    ])

        # Clear existing rows from the second row
        worksheet.delete_rows(2, worksheet.max_row)

        # Write filled rows to worksheet
        for row_data in filled_rows:
            worksheet.append(row_data)

        # Save the workbook to a buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Generate the file name
        current_datetime = self.format_datetime(datetime.now(), '%d%m%Y%H%M')
        file_name = f"{first_record.location_id.port.port_code}_{depot_code_first}_{company_profile_name}OUTEDI{current_datetime}.xlsx"

        # Remove existing attachments
        self.env['ir.attachment'].sudo().search([('name', '=', file_name)]).unlink()

        # Save as attachment
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'datas': base64.b64encode(output.read()),
            'public': True,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        for rec in fetch_records:
            rec.edi_out_attachment_id = attachment.id

        return attachment

    def add_edi_logs(self, data, fetch_records, status):
        try:
            # Prepare common values
            vals = {
                'location_id': self.location.id or False,
                'shipping_line_id': self.shipping_line_id.id or False,
                'file': data.id,
                'container_count': len(fetch_records),
                'status': status,
                'edi_sent_on': fields.Datetime.now() if status == 'success' else False,
                'ftp_upload': 'yes' if self.mode == 'ftp' and status == 'success' else '',
                'email_sent': 'yes' if self.mode != 'ftp' and status == 'success' else '',
                'type': 'move_in' if self.edi_type == 'move_in' else 'move_out',
            }
            # Add record IDs based on edi_type
            record_ids_field = 'move_in_ids' if self.edi_type == 'move_in' else 'move_out_ids'
            vals[record_ids_field] = fetch_records.mapped('id')
            # Create the EDI log
            self.env['edi.logs'].create(vals)
        except Exception as e:
            print(e)

    def send_file_via_ftp(self, data, fetch_records):
        """
        Send edi file via sftp.
        """
        # Decode the base64 data to get the actual file content
        file_content = base64.b64decode(data.datas)
        file_name = data.name

        # FTP/SFTP connection details
        ftp_directory = self.ftp_folder
        ftp_ip = self.ftp_location
        username = self.ftp_username
        password = self.ftp_password

        try:
            # Establish SFTP connection using paramiko
            transport = paramiko.Transport((ftp_ip, self.port_number))
            transport.connect(username=username, password=password)

            # Start SFTP session
            sftp = paramiko.SFTPClient.from_transport(transport)

            # Create a full path for the file to be uploaded
            remote_file_path = f'{ftp_directory}/{file_name}'

            # Write the file to the FTP server
            with sftp.file(remote_file_path, 'wb') as remote_file:
                remote_file.write(file_content)

            # Close the connection
            sftp.close()
            transport.close()

            # Update the record to send edis.
            self.last_run = fields.Datetime.now()
            # Get the current time in UTC and convert it to the user's timezone
            current_utc_time = fields.Datetime.now()
            user_tz = self.env.user.tz or 'UTC'

            # Add Edi logs
            self.add_edi_logs(data, fetch_records, 'success')

            # Convert UTC to the user's timezone and make it naive (strip timezone info)
            local_time = pytz.timezone(user_tz).localize(current_utc_time, is_dst=None).astimezone(
                pytz.timezone(user_tz)).replace(tzinfo=None)

            for record in fetch_records:
                    record.is_edi_send = True
                    record.edi_sent_on = local_time

        except Exception as e:
            self.add_edi_logs(data, fetch_records, 'error')
            print(f"Failed to upload file via SFTP: {str(e)}")

    def send_file_via_email(self, data, fetch_records, subject, body):
        """
        Send edi file via email.
        """
        try:
            mail_values = {
                'email_to': self.email_to,
                'email_from': self.email_from,
                'email_cc': self.email_cc,
                'subject': subject,
                'body_html': body,
                'attachment_ids': [(4, data.id)],
            }
            self.env['mail.mail'].create(mail_values).send()
            self.last_run = datetime.now()
            self.add_edi_logs(data,fetch_records,'success')
            # Get the current time in UTC and convert it to the user's timezone
            current_utc_time = fields.Datetime.now()
            user_tz = self.env.user.tz or 'UTC'

            # Convert UTC to the user's timezone and make it naive (strip timezone info)
            local_time = pytz.timezone(user_tz).localize(current_utc_time, is_dst=None).astimezone(
                pytz.timezone(user_tz)).replace(tzinfo=None)
            for record in fetch_records:
                record.is_edi_send = True
                record.edi_sent_on = local_time
        except Exception as e:
            self.add_edi_logs(data, fetch_records, 'error')
            print(f"Failed to send file via email: {str(e)}")

    def edi_send(self, realtime_frequency=None):
        """
        Searches for move.in records and sends an EDI file to an FTP server
        if the records meet the criteria.
        """
        if realtime_frequency:
            records = self
        else:
            # current_date_time = datetime.now().replace(microsecond=0, second=0)
            # records = self.env['edi.settings'].search([('next_scheduled','=',current_date_time)])
            records = self.env['edi.settings'].search([('active','=',True)])
        for rec in records:
            rec.last_run = datetime.now()
            edi_setting_location_id = rec.location.id
            edi_setting_shipping_line_id = rec.shipping_line_id.id
            edi_type = rec.edi_type
            edi_mode = rec.mode
            edi_format = rec.edi_format
            if edi_type == 'move_in':
                fetch_records = self.env['move.in'].search([
                    ('location_id', '=', edi_setting_location_id),
                    ('shipping_line_id', '=', edi_setting_shipping_line_id),
                    ('is_edi_send', '=', False)
                ])
                container_list = fetch_records.mapped('container')
                if edi_format == 'edi':
                    # Generate the EDI file content
                    data = rec.generate_edi_file(fetch_records)
                    if data and edi_mode == 'ftp':
                        rec.send_file_via_ftp(data, fetch_records)
                    elif data and edi_mode == 'email':
                        subject = 'Move In EDI'
                        formatted_container_list = '<br/>'.join(
                            [f"{i + 1}. {container}" for i, container in enumerate(container_list)])
                        body = f"""
                                Dear Sir/Madam,<br/><br/>
                                Please find the attached EDI files of the below mentioned containers:<br/>
                                {formatted_container_list}<br/><br/>
                                Regards,<br/>
                                CMS Team
                                """
                        rec.send_file_via_email(data, fetch_records, subject, body)
                    else:
                        _logger.info("No data found to generate IN EDI file for FTP upload.")
                else:
                    data = rec.generate_excel_file_msc_move_in_template(fetch_records)
                    if data and edi_mode == 'ftp':
                        rec.send_file_via_ftp(data, fetch_records)
                    elif data and edi_mode == 'email':
                        subject = 'Move In EDI'
                        formatted_container_list = '<br/>'.join(
                            [f"{i + 1}. {container}" for i, container in enumerate(container_list)])
                        body = f"""
                                                        Dear Sir/Madam,<br/><br/>
                                                        Please find the attached EDI files of the below mentioned containers:<br/>
                                                        {formatted_container_list}<br/><br/>
                                                        Regards,<br/>
                                                        CMS Team
                                                        """
                        rec.send_file_via_email(data, fetch_records, subject, body)
                    else:
                        _logger.info("No data found to generate IN EDI file for Email upload.")
            if edi_type == 'move_out':
                fetch_records = self.env['move.out'].search([
                    ('location_id', '=', edi_setting_location_id),
                    ('shipping_line_id', '=', edi_setting_shipping_line_id),
                    ('is_edi_send', '=', False)
                ])
                container_list = fetch_records.mapped('inventory_id').mapped('name')
                if edi_format == 'edi':
                    # Generate the EDI file content
                    data = rec.generate_edi_file(fetch_records)
                    if data and edi_mode == 'ftp':
                        rec.send_file_via_ftp(data, fetch_records)
                    elif data and edi_mode == 'email':
                        formatted_container_list = '<br/>'.join(
                            [f"{i + 1}. {container}" for i, container in enumerate(container_list)])
                        subject = 'Move Out EDI'
                        body = f"""
                            Dear Sir/Madam,<br/><br/>
                            Please find the attached EDI files of the below mentioned containers:<br/>
                            {formatted_container_list}<br/><br/>
                            Regards,<br/>
                            CMS Team
                        """
                        rec.send_file_via_email(data, fetch_records, subject, body)
                    else:
                        _logger.info("No data found to generate OUT EDI file for EDI upload.")
                else:
                    data = rec.generate_excel_file_msc_move_out_template(fetch_records)
                    if data and edi_mode == 'ftp':
                        rec.send_file_via_ftp(data, fetch_records)
                    elif data and edi_mode == 'email':
                        subject = 'Move OUT EDI'
                        formatted_container_list = '<br/>'.join(
                            [f"{i + 1}. {container}" for i, container in enumerate(container_list)])
                        body = f"""
                                                                            Dear Sir/Madam,<br/><br/>
                                                                            Please find the attached EDI files of the below mentioned containers:<br/>
                                                                            {formatted_container_list}<br/><br/>
                                                                            Regards,<br/>
                                                                            CMS Team
                                                                            """
                        rec.send_file_via_email(data, fetch_records, subject, body)
                    else:
                        _logger.info("No data found to generate OUT EDI file for Email upload.")

    def generate_list_edi_file(self, records):
        """
        Generates an EDI file for each record and returns them as a list of attachments.
        """
        if not records:
            return False

        attachments = []
        container_count = 1

        for rec in records:
            # Extract EDI format parts
            header_text = self.header_edi
            body_text = self.body_edi
            footer_text = self.footer_edi

            # Generate a dynamic filename using placeholders and append the record ID before the .edi extension
            file_name = self.replace_placeholders(self.attchment_file_name, rec, container_count)

            # Prepare the header section for each record
            if header_text:
                header_formatted = self.replace_placeholders(header_text, rec, container_count)
            else:
                header_formatted = ''

            # Prepare the body content for the record
            modified_body_text = body_text
            lines = body_text.splitlines()
            if not self.is_for_msc_shipping_line and not self.is_for_msc_south_shipping_line:
                modified_body_text = body_text
            else:
                # Logic for factory return for MSC shipping line
                if self.is_for_msc_shipping_line and rec.movement_type == 'factory_return':
                    if lines:
                        modified_body_text = '\n'.join(lines[:1])  # Include only Line 1

                # Logic for factory return for MSC South shipping line
                elif self.is_for_msc_south_shipping_line and rec.movement_type == 'factory_return':
                    if lines:
                        modified_body_text = '\n'.join(lines[1:2])  # Include only Line 2

                # Logic for non-factory return for MSC shipping line
                elif self.is_for_msc_south_shipping_line and rec.movement_type != 'factory_return':
                    # Add only Line 1 if the condition is met for import_destuffing or repo
                    if (rec.movement_type == 'import_destuffing' and rec.import_destuffing_from == 'CFS/ICD') or \
                            (rec.movement_type == 'repo' and rec.repo_from == 'CFS/ICD'):
                        modified_body_text = '\n'.join([lines[0]] + lines[2:])  # add Line 1,3,4
                    else:
                        # Only add Lines 3 and 4
                        modified_body_text = '\n'.join(lines[2:])
            body_formatted = self.replace_placeholders(modified_body_text, rec, container_count)

            ref_number_2 = ''
            # If the "reference_number2" placeholder exists, replace it with a generated sequence
            if '{reference_number2}' in body_formatted:
                ref_number_2 = self.env['ir.sequence'].next_by_code('edi.sequence.number.2')
                body_formatted = body_formatted.replace('{reference_number2}', ref_number_2)

            # Prepare the footer section for each record
            if footer_text:
                footer_formatted = self.replace_placeholders(footer_text, rec, container_count)
            else:
                footer_formatted = ''
            # Generate "reference_number1" if needed
            if ('{reference_number1}' in header_formatted or '{reference_number1}' in body_formatted or '{reference_number1}' in footer_formatted
                    or '{reference_number1}' in file_name):
                ref_number_1 = self.env['ir.sequence'].next_by_code('edi.sequence.number.1')
                header_formatted = header_formatted.replace('{reference_number1}', ref_number_1)
                body_formatted = body_formatted.replace('{reference_number1}', ref_number_1)
                footer_formatted = footer_formatted.replace('{reference_number1}', ref_number_1)
                file_name = file_name.replace('{reference_number1}', ref_number_1)

            # Check if the "Reference Number 2" placeholder exists in any part
            if ('{reference_number2}' in header_formatted or '{reference_number2}' in footer_formatted or
                    '{reference_number2}' in file_name):
                # Replace the "Reference Number 1" placeholder in all relevant sections
                header_formatted = header_formatted.replace('{reference_number2}', ref_number_2)
                footer_formatted = footer_formatted.replace('{reference_number2}', ref_number_2)
                file_name = file_name.replace('{reference_number2}', ref_number_2)

            # Combine all parts into the final EDI content for the record
            if header_formatted:
                edi_content = f"{header_formatted}\n{body_formatted}\n{footer_formatted}"
            else:
                edi_content = f"{body_formatted}\n{footer_formatted}"

            # Split the file name to insert the record ID
            base_name, extension = file_name.rsplit('.', 1)  # Split on the last dot
            unique_file_name = f"{base_name}_{rec.id}.{extension}"  # Append the record ID before the extension

            # Create an attachment for each EDI file
            attachment = {
                'name': unique_file_name,
                'datas': base64.b64encode(edi_content.encode('utf-8')),
                'mimetype': 'text/plain',
                'edi_content': edi_content,  # Keep plain text content to use in zip file
            }
            attachments.append(attachment)

        return attachments

    @api.model
    def send_list_view_rec_edis(self, selected_ids):
        """
        This method is called from the JS file.
        :param selected_ids: List of selected record IDs
        :return: Dictionary with success or error message
        """
        if not selected_ids:
            return {'error': 'No selected record or model found!'}

        # Fetch move.in records that exist for the given selected IDs
        move_in_obj = self.env['move.in'].browse(selected_ids).exists()

        if not move_in_obj:
            return {'error': 'No records found to generate EDI files'}

        # Get unique 'location_id' and 'shipping_line_id' values
        location_ids = move_in_obj.mapped('location_id').ids
        shipping_line_ids = move_in_obj.mapped('shipping_line_id').ids

        # Ensure all selected records have the same location and shipping line
        if len(set(location_ids)) > 1 or len(set(shipping_line_ids)) > 1:
            return {'error': 'All selected records must have the same shipping line or location.'}

        # Fetch EDI setting record
        edi_setting_rec = self.env['edi.settings'].search([
            ('shipping_line_id', '=', shipping_line_ids[0]),
            ('location', '=', location_ids[0]),
            ('edi_type', '=', 'move_in'),
            ('edi_format', '=', 'edi')
        ], limit=1)

        if not edi_setting_rec:
            return {'error': 'No EDI configurations found for selected records'}

        if not edi_setting_rec.location.email:
            return {'error': 'No email found for selected location.'}

        # Generate individual EDI files for each record
        attachments = edi_setting_rec.generate_list_edi_file(move_in_obj)
        container_list = move_in_obj.mapped('container')

        if not attachments:
            return {'error': 'Failed to generate EDI files'}

        # Format the container list for the email body
        formatted_container_list = '<br/>'.join(f"{i + 1}. {container}" for i, container in enumerate(container_list))
        subject = 'Move In EDI'
        body = f"""
            Dear Sir/Madam,<br/><br/>
            Please find the attached EDI files of the below-mentioned containers:<br/>
            {formatted_container_list}<br/><br/>
            Regards,<br/>
            CMS Team
        """

        # Initialize a list for the attachments to be sent via email
        email_attachments = []

        # Loop over each attachment and associate it with its respective `move_in_obj` record
        for move_in, attachment in zip(move_in_obj, attachments):
            # Create an attachment for the `move.in` record
            edi_attachment = self.env['ir.attachment'].create({
                'name': attachment['name'],
                'datas': attachment['datas'],
                'res_model': 'move.in',
                'res_id': move_in.id,
                'public': True,
                'type': 'binary',
                'mimetype': 'text/plain',
            })

            # Set the attachment on the move_in record
            move_in.edi_in_attachment_id = edi_attachment.id

            # Append the attachment to the email attachments list
            email_attachments.append((4, edi_attachment.id))

        # If there's only one attachment, send it directly
        if len(email_attachments) == 1:
            self._send_edi_email(subject, body, email_attachments, edi_setting_rec, selected_ids)
        else:
            # Create a zip file for multiple EDI files
            zip_attachment = self._create_zip_attachment(attachments, 'move_in')
            # Send zip file as email attachment
            self._send_edi_email(subject, body, [(4, zip_attachment.id)], edi_setting_rec, selected_ids)

        # Mark records as EDI sent and update sent time
        self._update_edi_sent_status(move_in_obj)

        return {'success': True}

    @api.model
    def send_move_out_list_view_rec_edis(self, selected_ids):
        """
        This method is called from the JS file.
        :param selected_ids: List of selected record IDs
        :return: Dictionary with success or error message
        """
        if not selected_ids:
            return {'error': 'No selected record or model found!'}

        # Fetch move.in records that exist for the given selected IDs
        move_out_obj = self.env['move.out'].browse(selected_ids).exists()

        if not move_out_obj:
            return {'error': 'No records found to generate EDI files'}

        # Get unique 'location_id' and 'shipping_line_id' values
        location_ids = move_out_obj.mapped('location_id').ids
        shipping_line_ids = move_out_obj.mapped('shipping_line_id').ids

        # Ensure all selected records have the same location and shipping line
        if len(set(location_ids)) > 1 or len(set(shipping_line_ids)) > 1:
            return {'error': 'All selected records must have the same shipping line or location.'}

        # Fetch EDI setting record
        edi_setting_rec = self.env['edi.settings'].search([
            ('shipping_line_id', '=', shipping_line_ids[0]),
            ('location', '=', location_ids[0]),
            ('edi_type', '=', 'move_out'),
            ('edi_format', '=', 'edi')
        ], limit=1)

        if not edi_setting_rec:
            return {'error': 'No EDI configurations found for selected records'}

        if not edi_setting_rec.location.email:
            return {'error': 'No email found for selected location.'}

        # Generate individual EDI files for each record
        attachments = edi_setting_rec.generate_list_edi_file(move_out_obj)
        container_list = move_out_obj.mapped('inventory_id').mapped('name')

        if not attachments:
            return {'error': 'Failed to generate EDI files'}

        # Format the container list for the email body
        formatted_container_list = '<br/>'.join(f"{i + 1}. {container}" for i, container in enumerate(container_list))
        subject = 'Move OUT EDI'
        body = f"""
               Dear Sir/Madam,<br/><br/>
               Please find the attached EDI files of the below-mentioned containers:<br/>
               {formatted_container_list}<br/><br/>
               Regards,<br/>
               CMS Team
           """

        # Initialize a list for the attachments to be sent via email
        email_attachments = []

        # Loop over each attachment and associate it with its respective `move_in_obj` record
        for move_out, attachment in zip(move_out_obj, attachments):
            # Create an attachment for the `move.in` record
            edi_attachment = self.env['ir.attachment'].create({
                'name': attachment['name'],
                'datas': attachment['datas'],
                'res_model': 'move.out',
                'res_id': move_out.id,
                'public': True,
                'type': 'binary',
                'mimetype': 'text/plain',
            })

            # Set the attachment on the move_in record
            move_out.edi_out_attachment_id = edi_attachment.id  # Assuming `attachment_field` is the many2one field for ir.attachment

            # Append the attachment to the email attachments list
            email_attachments.append((4, edi_attachment.id))

        # If there's only one attachment, send it directly
        if len(email_attachments) == 1:
            self._send_edi_email(subject, body, email_attachments, edi_setting_rec,selected_ids)
        else:
            # Create a zip file for multiple EDI files
            zip_attachment = self._create_zip_attachment(attachments, 'move_out')
            # Send zip file as email attachment
            self._send_edi_email(subject, body, [(4, zip_attachment.id)], edi_setting_rec, selected_ids)

        # Mark records as EDI sent and update sent time
        self._update_edi_sent_status(move_out_obj)

        return {'success': True}

    def add_edi_logs_for_manual_operation(self,data,status,fetch_records):
        try:
            # Prepare common values
            vals = {
                'location_id': self.location.id or False,
                'shipping_line_id': self.shipping_line_id.id or False,
                'file': data,
                'container_count': len(fetch_records),
                'status': status,
                'generated_by': 'user',
                'edi_sent_on': fields.Datetime.now() if status == 'success' else False,
                'email_sent': 'yes' if status == 'success' else '',
                'type': 'move_in' if self.edi_type == 'move_in' else 'move_out',
            }
            # Add record IDs based on edi_type
            record_ids_field = 'move_in_ids' if self.edi_type == 'move_in' else 'move_out_ids'
            vals[record_ids_field] = fetch_records
            # Create the EDI log
            self.env['edi.logs'].create(vals)
        except Exception as e:
            print(e)

    def _send_edi_email(self, subject, body, attachment_ids, edi_setting_rec, selected_ids):
        """ Send email with specified subject, body, and attachments. """
        try:
            mail_values = {
                'email_to': edi_setting_rec.location.email,
                'subject': subject,
                'body_html': body,
                'attachment_ids': attachment_ids,
            }
            self.env['mail.mail'].create(mail_values).send()
            # Add logs.
            edi_setting_rec.add_edi_logs_for_manual_operation(attachment_ids[0][1],'success',selected_ids)
        except Exception as e:
            # Add logs.
            edi_setting_rec.add_edi_logs_for_manual_operation(attachment_ids[0][1], 'error', selected_ids)
            print(e)

    def _update_edi_sent_status(self, records):
        """ Mark records as EDI sent and set the sent time. """
        current_utc_time = fields.Datetime.now()
        user_tz = self.env.user.tz or 'UTC'
        # Convert UTC to the user's timezone and make it naive (strip timezone info)
        local_time = pytz.timezone(user_tz).localize(current_utc_time, is_dst=None).astimezone(
            pytz.timezone(user_tz)).replace(tzinfo=None)
        records.write({'is_edi_send': True, 'edi_sent_on': local_time})

    def _create_zip_attachment(self, attachments, model):
        """ Create a zip file from multiple attachments and return it as an Odoo attachment. """
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for attachment in attachments:
                zip_file.writestr(attachment['name'], attachment['edi_content'])

        zip_buffer.seek(0)
        zip_data = base64.b64encode(zip_buffer.read())

        current_time_str = self.format_datetime(datetime.utcnow(), '%d%m%y%H%M')
        if model == 'move_in':
            zip_filename = f'MoveIn_EDI_{current_time_str}.zip'
        else:
            zip_filename = f'MoveOut_EDI_{current_time_str}.zip'

        # Save the zip file as an attachment in Odoo
        return self.env['ir.attachment'].create({
            'name': zip_filename,
            'datas': zip_data,
            'public': True,
            'type': 'binary',
            'mimetype': 'application/zip',
        })
