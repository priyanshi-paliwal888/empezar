from odoo.tests.common import TransactionCase
import io
import base64
import openpyxl

class TestUpdateTariffModel(TransactionCase):

    def setUp(self):
        super().setUp()

        self.excel_data = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Repair Code", "Some Data", "Error"])
        ws.append(["RC001", "Data1", None])
        ws.append(["RC002", "Data2", "Error found"])
        wb.save(self.excel_data)
        self.excel_data.seek(0)

        """ Encode the Excel file in base64 """
        self.uploaded_file = base64.b64encode(self.excel_data.read())

        self.update_tariff = self.env["update.tariff"].create({
            "name": "Tariff1",
            "rec_status": 'in_progress',
            'upload_tariff_file': self.uploaded_file,
        })

    def test_set_upload_id_when_none(self):
        """ Test setting upload_id when it is not already set. """
        self.update_tariff.upload_id = False
        self.update_tariff.set_upload_id()
        self.assertEqual(int(self.update_tariff.upload_id),self.update_tariff.id, "upload_id should be set to the record id.")

    def test_set_upload_id_when_already_set(self):
        """ Test that upload_id is not overwritten when it is already set. """
        self.update_tariff.upload_id = "999"
        self.update_tariff.set_upload_id()
        self.assertEqual(self.update_tariff.upload_id, "999", "upload_id should remain unchanged if already set.")

    def test_download_xlsx_file(self):
        result = self.update_tariff.download_xlsx_file()

        """ Check the returned value """
        self.assertEqual(result['type'], 'ir.actions.act_url')
        self.assertEqual(result['url'], '/empezar_repair/static/src/document/Tariff Sample.xlsx')
        self.assertEqual(result['target'], 'new')

    def test_action_submit(self):
        self.update_tariff.action_submit()

        """ Here you can check if the expected changes happened"""
        self.assertTrue(self.update_tariff.rec_status, 'Expected status should be updated')
        file_content = base64.b64decode(self.update_tariff.upload_tariff_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = wb.active

        """ Check for the "Error" column"""
        error_column_index = 3
        error_header = sheet.cell(row=1, column=error_column_index).value
        self.assertEqual(error_header, "Error", "Expected 'Error' header should be present.")

    def test_action_submit_no_in_progress_records(self):
        """ Test the action_submit method when there are no in-progress records """
        self.update_tariff.write({'rec_status': 'success'})
        self.update_tariff.action_submit()

        self.assertEqual(self.update_tariff.rec_status, 'success', "Record status should remain 'success'.")

    def test_update_tariff_records_with_errors(self):
        """ Test the update_tariff_records method with non-empty Error column """
        file_content = base64.b64decode(self.update_tariff.upload_tariff_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = wb.active
        error_column_index = 3
        self.update_tariff.update_tariff_records(wb, sheet, error_column_index)

        """ Assertions to verify the record status and file update """
        self.assertEqual(self.update_tariff.rec_status, 'error', "Expected record status to be 'error'.")

        """ Check if the uploaded file has been updated """
        updated_file_content = base64.b64decode(self.update_tariff.upload_tariff_file)
        updated_wb = openpyxl.load_workbook(io.BytesIO(updated_file_content))
        updated_sheet = updated_wb.active

        """ Check if the Error column width was updated """
        self.assertEqual(updated_sheet.column_dimensions['S'].width, 45, "Expected 'Error' column width to be 45.")

