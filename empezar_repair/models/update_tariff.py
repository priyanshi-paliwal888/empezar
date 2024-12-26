# -*- coding: utf-8 -*-
import io
import base64
import logging
from odoo import fields, models, api,_
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)

class UpdateTariff(models.Model):

    _name = "update.tariff"
    _description = "Update Tariff"

    name = fields.Char(string="File Name", required=True)
    upload_tariff_file = fields.Binary(string="Upload File",
                                          help="Only XLS or XLSX files with max size of 5 MB and"
                                                                        " having max 200 entries.",
                                          required=True, copy=False, exportable=False)
    upload_id = fields.Char(string="Upload ID", compute='set_upload_id')
    uploaded_by = fields.Char(string="Uploaded By")
    uploaded_on = fields.Datetime(string="Uploaded On")
    rec_status = fields.Selection([
        ('success', 'Success'),
        ('error', 'Error'),
        ('in_progress', 'In Progress')
    ], string="Status")

    def set_upload_id(self):
        """
        Set upload id values.
        :return:
        """
        for rec in self:
            if not rec.upload_id:
                rec.upload_id = rec.id
            else:
                pass

    @api.model
    def download_xlsx_file(self):
        """ Download the xlsx file uploaded
                """
        return {
            'type': 'ir.actions.act_url',
            'url': '/empezar_repair/static/src/document/Tariff Sample.xlsx',
            'target': 'new',
        }

    def action_submit(self):
        """  Validate the actions on button submit
                """
        active_records = self.search([('rec_status', '=', 'in_progress')])
        if active_records:
            for rec in active_records:
                # Read the uploaded Excel file
                try:
                    file_content = base64.b64decode(rec.upload_tariff_file)
                    df = pd.read_excel(io.BytesIO(file_content))
                    env = self.env

                    file_data = io.BytesIO(base64.b64decode(rec.upload_tariff_file))
                    wb = openpyxl.load_workbook(file_data)
                    sheet = wb.active  # Assuming data is in the first sheet

                    remarks_column_index = None
                    for col in sheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == "Repair Code":
                                remarks_column_index = cell.column

                    # Delete any column after "Repair Code" if it is empty
                    next_column_index = remarks_column_index + 1
                    sheet.delete_cols(next_column_index)

                    # Insert a column after "Repair Code" for "Error"
                    error_column_index = remarks_column_index + 1
                    sheet.insert_cols(error_column_index)

                    # Set the header for the newly inserted column ("Error")
                    sheet.cell(row=1, column=error_column_index, value="Error")\
                        .font = Font(bold=True)

                    # Save the workbook back to BytesIO or to a file
                    # If saving to BytesIO:
                    file_output = io.BytesIO()
                    wb.save(file_output)
                    file_output.seek(0)

                    UpdateTariff.check_damage_location_validations(df, env, sheet, error_column_index)
                    UpdateTariff.check_damage_type_validations(df, env, sheet, error_column_index)
                    UpdateTariff.check_component_type_validations(df, env, sheet, error_column_index)
                    UpdateTariff.check_repair_type_validations(df, env, sheet, error_column_index)
                    UpdateTariff.check_measurement_validations(df, sheet, error_column_index)
                    UpdateTariff.check_key_value_validation(df, sheet, error_column_index)
                    UpdateTariff.check_numeric_values_validation(df, sheet, error_column_index)
                    UpdateTariff.check_size_type_validations(df, env, sheet, error_column_index)
                    rec.record_creation_for_repair_tariff(df, sheet, error_column_index)
                    rec.update_tariff_records(wb, sheet, error_column_index)
                except Exception as e:
                    _logger.error('There is some issue while update a inventory in %s\n%s', rec.upload_id, e)
        else:
            _logger.info("There is no any in progress records")



    def update_tariff_records(self, wb, sheet, error_column_index):
        """ Update the tariff records when the sheet is uploaded."""
        columns_to_widths = {'S': 45}

        # Iterate over the columns_to_widths dictionary
        for column, width in columns_to_widths.items():
            sheet.column_dimensions[column].width = width

        # Save workbook to BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)  # Move to the beginning of the BytesIO object

        # Encode the output to base64
        encoded_error_df = base64.b64encode(output.read())

        error_column_values_empty = True

        # Iterate through all rows in the specified column 'Error'
        for row in sheet.iter_rows(min_col=error_column_index, max_col=error_column_index, min_row=2):
            for cell in row:
                if cell.value is not None:
                    # If any cell in the column has a value, mark as not all empty
                    error_column_values_empty = False
                    break  # No need to check further if we found a non-empty cell
            if not error_column_values_empty:
                break

        if not error_column_values_empty:
            self.write({
                'rec_status': 'error',
                'upload_tariff_file': encoded_error_df
            })
        else:
            self.write({
                'rec_status': 'success'
            })

    @staticmethod
    def validate_column(column_name, valid_values, error_message, df, sheet, error_column_index):
        """
        helper method to validate the column.
        :param column_name:
        :param valid_values:
        :param error_message:
        :param df:
        :param sheet:
        :param error_column_index:
        :return:
        """

        # Get column data without NaN values
        column_data = df[column_name].dropna()

        # Find invalid values
        invalid_values = column_data[~column_data.isin(valid_values)]

        if not invalid_values.empty:
            # Convert all invalid values to strings and join them
            invalid_value_index = invalid_values.index.tolist()
            for rec in invalid_value_index:
                get_data = sheet.cell(row=rec + 2, column=error_column_index).value
                values = get_data + f'\n>{error_message}.' if get_data else f'>{error_message}.'
                sheet.cell(row=rec + 2, column=error_column_index, value=values)

    @staticmethod
    def check_damage_location_validations(df, env, sheet, error_column_index):
        """
        Validate 'Damage Location' columns from the uploaded Excel sheet.
        """

        # Fetch valid "Damage Location" codes from the `damage.locations` model
        valid_damage_location_codes = set(env['damage.locations'].search([]).mapped('code'))

        # Validate the 'Damage Location' column using the helper method
        UpdateTariff.validate_column('Damage location', valid_damage_location_codes,
                                     'Value not available in Damage location seeder', df, sheet, error_column_index)

    @staticmethod
    def check_damage_type_validations(df, env, sheet, error_column_index):
        """
        Validate 'Damage Type' columns from the uploaded Excel sheet.
        """
        # Fetch valid "Damage Type" codes from the `damage.type` model
        valid_damage_type_codes = set(env['damage.type'].search([]).mapped('damage_type_code'))

        # Validate the 'Damage Type' column using the helper method
        UpdateTariff.validate_column('Damage Type', valid_damage_type_codes,
                                     'Value not available in Damage type seeder', df, sheet, error_column_index)

    @staticmethod
    def check_component_type_validations(df, env, sheet, error_column_index):
        """
        Validate 'Component' columns from the uploaded Excel sheet.
        """
        # Fetch valid "Component" codes from the `master.component` model
        valid_component_codes = set(env['master.component'].search([]).mapped('code'))

        # Validate the 'Component' column using the helper method
        UpdateTariff.validate_column('Component', valid_component_codes,
                                     'Value not available in component seeder', df, sheet, error_column_index)

    @staticmethod
    def check_repair_type_validations(df, env, sheet, error_column_index):
        """
        Validate 'Repair Type' columns from the uploaded Excel sheet.
        """

        # Fetch valid "Repair Type" codes from the `repair.types` model
        valid_repair_type_codes = set(env['repair.types'].search([]).mapped('repair_type_code'))

        # Validate the 'Repair Type' column using the helper method
        UpdateTariff.validate_column('Repair Type', valid_repair_type_codes,
                                     'Value not available in Repair Type seeder', df, sheet, error_column_index)

    @staticmethod
    def check_size_type_validations(df, env, sheet, error_column_index):
        """
        Validate 'Size/Type' columns from the uploaded Excel sheet.
        """
        # Fetch valid "Size/Type" codes from the `container.type.data` model
        valid_repair_type_codes = set(env['container.type.data'].search([]).mapped('company_size_type_code'))

        # Call the validation helper method
        UpdateTariff.validate_column('Size/Type', valid_repair_type_codes,
                                     'Invalid container type size', df, sheet, error_column_index)
    @staticmethod
    def check_measurement_validations(df, sheet, error_column_index):
        """
        Check status column validations.
        :return:
        """

        allowed_measurements = {'FOT', 'INH', 'QTY'}
        error_message = 'Only "FOT", "INH", and "QTY" accepted as measurement values'
        UpdateTariff.validate_column('Measurement', allowed_measurements, error_message, df, sheet, error_column_index)

    @staticmethod
    def check_key_value_validation(df, sheet, error_column_index):
        """
        Validate that the values under the 'Key Value' column are only 'LN', 'LN*W', or 'Q'.
        :param df: DataFrame containing the uploaded file data
        """
        allowed_key_values = {'LN', 'LN*W', 'Q'}
        error_message = 'Only "LN", "LN*W", and "Q" accepted as Key values'
        UpdateTariff.validate_column('Key Value', allowed_key_values, error_message, df, sheet, error_column_index)

    @staticmethod
    def check_numeric_values_validation(df, sheet, error_column_index):
        """
        Validate that the values under specified columns are numeric.
        """
        numeric_columns = ['Material Cost', 'Labour Hours']
        allowed_numeric_types = (int, float)

        for column in numeric_columns:
            if column in df.columns:
                # Identify rows where the value is non-numeric
                non_numeric_rows = df[~df[column].apply(lambda x: isinstance(x, allowed_numeric_types))].dropna()

                # Apply validation only on those rows where the issue exists
                for index in non_numeric_rows.index:
                    # Update the specific row with the error
                    UpdateTariff.validate_column(
                        column,
                        allowed_numeric_types,
                        f'Only numeric values accepted for {column}',
                        df.loc[[index]],
                        sheet,
                        error_column_index
                    )

    @api.model
    def record_creation_for_repair_tariff(self, df, sheet, error_column_index):
        """Create records in update.repair.tariff if all validations are passed."""
        try:
            error_column_values_empty = True

            # Check if the error column is empty
            for row in sheet.iter_rows(min_col=error_column_index, max_col=error_column_index, min_row=2):
                for cell in row:
                    if cell.value is not None:
                        error_column_values_empty = False
                        break
                if not error_column_values_empty:
                    break

            if error_column_values_empty:
                # Iterate through each row of the DataFrame
                for index, row in df.iterrows():
                    # Prepare values for update.repair.tariff
                    vals_repair_tariff = {
                        'damage_location': row['Damage location'],
                        'component': row['Component'],
                        'damage_type': row['Damage Type'],
                        'repair_type': row['Repair Type'],
                        'measurement': row['Measurement'],
                        'size_type': row['Size/Type'],
                        'material_cost':row['Material Cost'],
                        'labour_hour': row['Labour Hours'],
                        'key_value': row['Key Value'],
                        'limit': row['Limit'],
                        'repair_code': row['Repair Code'],
                    }

                    # Create the new record
                    self.env['update.repair.tariff'].create(vals_repair_tariff)

        except Exception as e:
            raise ValidationError(_('Some issues occurred while creating a repair tariff record: %s', e))

        return True
