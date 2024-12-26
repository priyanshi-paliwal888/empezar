# -*- coding: utf-8 -*-
import io
from io import BytesIO
import base64
from datetime import date
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from odoo import fields, models, api,_
from odoo.addons.empezar_base.models.res_users import ResUsers
from odoo.exceptions import ValidationError
import pytz


class VesselBooking(models.Model):
    _name = "vessel.booking"
    _description = "Vessel Booking"
    _rec_name = 'booking_no'

    shipping_line_logo = fields.Binary(
        string="Shipping Line", related="shipping_line_id.logo")
    shipping_line_id = fields.Many2one(
        "res.partner",
        string="Shipping Line",
        domain="[('is_shipping_line', '=', True)]", required=True
    )
    is_shipping_line = fields.Boolean("Is Shipping Line", default=False)
    location = fields.Many2many('res.company',string="Location",required=True)
    transporter_name = fields.Many2one(
        "res.partner", string="Transporter",
        domain="[('parties_type_ids.name', '=', 'Transporter')]",required=True)
    total_containers = fields.Integer(string="Total Containers", compute='_get_total_containers')
    balance_containers = fields.Integer(string="Balance Containers", compute='_get_total_balance_container')
    rec_status = fields.Selection(
        [
            ("active", "Active"),
            ("disable", "Disable"),
        ],
        default="active",
        string="Status",compute="_check_active_records")
    booking_no= fields.Char(string="Booking No.",size=32,required=True)
    booking_date = fields.Date("Booking Date",required=True)
    validity_datetime = fields.Datetime("Valid Till",required=True)
    cutoff_datetime = fields.Datetime("Cut-Off Date/Time",required=True)
    vessel= fields.Char(string="Vessel",size=32)
    voyage= fields.Char(string="Voyage",size=32)
    via= fields.Char(string="Via",size=32)
    service= fields.Char(string="Service",size=32)
    remarks= fields.Char(string="Remarks",size=128)
    port_loading = fields.Many2one(
        'master.port.data',
        string='Port of Loading')
    port_discharge = fields.Many2one(
        'master.port.data',
        string='Port of Discharge')
    container_qty = fields.Integer(string="Quantity")
    actions = fields.Selection(
        [("move_in", "Move In"), ("move_out", "Move Out")],
        string="Movement Type")

    container_details = fields.One2many("container.details", "booking_id")
    container_numbers = fields.One2many('container.number', 'vessel_booking_id', string='Container Numbers')
    display_create_info = fields.Char(compute="_get_create_record_info")
    display_modified_info = fields.Char(compute="_get_modify_record_info")
    display_sources = fields.Char(string="Sources",readonly=True,default="User Entry")
    active = fields.Boolean(string="Active", default=True)
    display_info = fields.Char(
        string='Booking No./Valid Till Date',
        compute='_compute_display_info',
    )
    refer_container_selection = fields.Char("Refer Container Selection",
                                            compute='_compute_location_shipping_line', store=True)


    @api.depends('location', 'shipping_line_id','container_details.container_size_type')
    def _compute_location_shipping_line(self):
        for record in self:
            if not record.location:
                raise ValidationError(_("Please select at least one Location."))
            elif not record.shipping_line_id:
                raise ValidationError(_("Please select at least one Shipping Line."))

            # Fetch location IDs for the current record
            location_ids = record.location.ids
            shipping_mappings = self.env['location.shipping.line.mapping'].search([
                ('company_id', 'in', location_ids),
                ('shipping_line_id', '=', record.shipping_line_id.id)
            ])
            refer_container_values = set()
            # Iterate through mappings to collect refer_container values
            for mapping in shipping_mappings:
                refer_container_values.add(mapping.refer_container)

            if 'yes' in refer_container_values and 'no' in refer_container_values:
                # Both refer and non-refer containers are allowed
                domain = [('is_refer', '=', 'yes')]
            elif 'yes' in refer_container_values:
                # Only refer containers are allowed
                domain = [('is_refer', '=', 'yes')]
            elif 'no' in refer_container_values:
                # Only non-refer containers are allowed
                domain = [('is_refer', '=', 'no')]
            else:
                domain = [('is_refer', '=', 'no')]

            # Check for location changes
            previous_location = record._origin.location
            if previous_location != record.location:
                # Reset container size types if location has changed
                record.container_details = [(5, 0, 0)]

            record.refer_container_selection = str(domain)


    @api.depends('container_details')
    def _get_total_containers(self):
        for rec in self:
            rec.total_containers = sum(rec.container_details.mapped('container_qty'))

    @api.depends('container_details')
    def _get_total_balance_container(self):
        for rec in self:
            rec.balance_containers = sum(rec.container_details.mapped('balance'))

    @api.depends('booking_no','validity_datetime')
    def _compute_display_info(self):
        """
        check booking No. and Validity date is valid or not
        :return:
        """
        for record in self:
            get_company = self.env['res.company'].search([('parent_id', '=', False),
                                                          ('active', '=', True)], limit=1)
            company_format = get_company.date_format
            if record.validity_datetime:
                local_tz = pytz.timezone('Asia/Kolkata')
                local_dt = pytz.utc.localize(record.validity_datetime).astimezone(local_tz)
                date_formats = {
                    'DD/MM/YYYY': '%d/%m/%Y',
                    'YYYY/MM/DD': '%Y/%m/%d',
                    'MM/DD/YYYY': '%m/%d/%Y'
                }
                formatted_date = local_dt.strftime(date_formats.get(company_format, '%d/%m/%Y'))
                formatted_time = local_dt.strftime('%I:%M %p')
                record.display_info = f"{record.booking_no} Valid Till: {formatted_date} {formatted_time}"
            else:
                record.display_info = record.booking_no

    @api.constrains('booking_no', 'shipping_line_id', 'active')
    def _check_unique_booking_no_shipping_line(self):
        """
           check booking No. with shipping line is valid or not
        """
        for record in self:
            if record.active:
                domain = [
                    ('id', '!=', record.id),
                    ('booking_no', '=', record.booking_no),
                    ('shipping_line_id', '=', record.shipping_line_id.id),
                    ('active', '=', True)
                ]
                existing_record = self.search(domain)
                if existing_record:
                    raise ValidationError(
                        "The data with the same Vessel Booking number for the carrier already exists.")

    def upload_containers(self):
        """
            wizard action for upload containers
             :return:
        """
        return {
            'name': _('Upload Containers'),
            'type': 'ir.actions.act_window',
            'res_model': 'booking.import.wizard',
            'views': [[False, 'form']],
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
            }
        }

    def download_container_record(self):
        """
        Generate and download an Excel report with container records including 'Name' and 'Unlink Reason'.

        Returns:
            dict: Action dictionary to trigger the download of the generated Excel file.
        """
        # Create an in-memory output file for the new workbook
        output = io.BytesIO()
        # Create a workbook and add a worksheet
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()

        # Write headers
        worksheet.write(0, 0, 'Name')
        worksheet.write(0, 1, 'Unlink Reason')

        # Write data
        row = 1
        for container in self.container_numbers:
            name = container.name or ''
            unlink_reason = container.unlink_reason.reason or ''

            # Ensure unlink_reason is a string
            if not isinstance(unlink_reason, str):
                unlink_reason = str(unlink_reason)

            worksheet.write(row, 0, name)
            worksheet.write(row, 1, unlink_reason)
            row += 1

        # Close the workbook before sending the data
        workbook.close()

        # Get the data from the in-memory output file
        output.seek(0)
        xlsx_data = output.read()
        output.close()

        # Create the attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'container_record.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'res_model': 'vessel.booking',
            'res_id': self.id,
            'mimetype': 'application/vnd.ms-excel'
        })

        # Return the action to download the file
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }

    def download_booking_xlsx_file(self):
        """
            This method used for the download the sample container xlsx file
        """
        total_quantities = {}
        summary_data = {}

        for detail in self.container_details:
            type_size_id = detail.container_size_type.id
            type_size_code = detail.container_size_type.company_size_type_code

            if type_size_code not in total_quantities:
                total_quantities[type_size_code] = {
                    'container_type_size': detail.container_size_type.name,
                    'total_qty': 0,
                    'type_size_id': type_size_id
                }
            total_quantities[type_size_code]['total_qty'] += detail.container_qty
            summary_data[type_size_code] = type_size_id

        # Calculate used quantities from container numbers
        used_quantities = {}
        for container_number in self.container_numbers:
            type_size_code = container_number.name.split('(')[-1].strip(')')
            if type_size_code not in used_quantities:
                used_quantities[type_size_code] = 0
            used_quantities[type_size_code] += 1

        # Calculate remaining quantities and prepare data
        data = []
        for type_size_code, total_data in total_quantities.items():
            used_qty = used_quantities.get(type_size_code, 0)
            remaining_qty = total_data['total_qty'] - used_qty

            if remaining_qty > 0:
                data.extend([{
                    'Container Type Size': total_data['container_type_size'],
                    'Container Number': ''
                }] * remaining_qty)

        # Create a DataFrame
        df = pd.DataFrame(data, columns=['Container Type Size', 'Container Number'])

        # Save to a BytesIO buffer
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)

        # Load the workbook and set column widths
        workbook = load_workbook(buffer)
        worksheet = workbook.active
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 20

        # Prepare the file for download
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Prepare the XLSX file for download
        attachment_name = 'Container_Booking.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': attachment_name,
            'type': 'binary',
            'datas': base64.b64encode(buffer.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Return the file download link
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }

    @api.constrains('booking_date')
    def _check_booking_date(self):
        """
           check booking Date is valid or not
           :return:
        """
        for record in self:
            if record.booking_date and record.booking_date > date.today():
                raise ValidationError(
                    "Please select a valid date. Booking date cannot be greater than the current date.")


    @api.constrains('booking_date', 'validity_datetime')
    def _check_validity_datetime(self):
        """
          check booking Date and Validity Date is valid or not
        """
        for record in self:
            if record.validity_datetime and record.booking_date:
                if record.validity_datetime < fields.Datetime.to_datetime(record.booking_date):
                    raise ValidationError(
                    "Validity date cannot be less than the booking date. Please select a future date.")


    @api.constrains('booking_date', 'cutoff_datetime')
    def _check_cutoff_datetime(self):
        """
        check booking Date and Cutoff Date is valid or not
        """
        for record in self:
            if record.cutoff_datetime and record.booking_date:
                if record.cutoff_datetime < fields.Datetime.to_datetime(record.booking_date):
                    raise ValidationError("Cut-Off date cannot be less than the booking date,"
                                          "please select a future date for Cutoff DateTime.")


    @api.constrains('container_details','booking_no')
    def _check_container_types_present(self):
        for booking in self:
            if not booking.container_details:
                raise ValidationError("Please select at least one container detail.")

    def _check_active_records(self):
        for rec in self:
            if rec.active:
                rec.rec_status = "active"
            else:
                rec.rec_status = "disable"


    def _get_create_record_info(self):
        """
        Assign create record log string to the appropriate field.
        :return: none
        """
        for rec in self:
            if rec.env.user.tz and rec.create_uid:
                tz_create_date = ResUsers.convert_datetime_to_user_timezone(rec.env.user, rec.create_date)
                create_uid_name = rec.create_uid.name
                if tz_create_date:
                    rec.display_create_info = ResUsers.get_user_log_data(rec, tz_create_date, create_uid_name)
            else:
                rec.display_create_info = ''

    def _get_modify_record_info(self):
        """
            Assign update record log string to the appropriate field.
            :return: none
        """
        for rec in self:
            if self.env.user.tz and rec.write_uid:
                tz_write_date = ResUsers.convert_datetime_to_user_timezone(rec.env.user, rec.write_date)
                write_uid_name = rec.write_uid.name
                if tz_write_date:
                    rec.display_modified_info = ResUsers.get_user_log_data(rec, tz_write_date, write_uid_name)
            else:
                rec.display_modified_info = ''


    def upload_button_action(self):
        return True
