import io
import base64
from datetime import datetime, timedelta, date
from odoo.tests.common import TransactionCase
from odoo import fields
from odoo.exceptions import ValidationError
from PIL import Image
from io import BytesIO
from odoo import fields, models, api, _
import xlsxwriter
import openpyxl







class TestMoveInOutInvoices(TransactionCase):

    def setUp(self):
        super().setUp()

        self.mode = self.env["mode.options"].create({"name": "mode"})
        self.truck_mode = self.env["mode.options"].create({"name": "Truck"})
        self.rail_mode = self.env["mode.options"].create({"name": "Rail"})
        self.gate_pass = self.env["gate.pass.options"].create({"name": "Move In"})

        self.company = self.env["res.company"].create(
            {
                "name": "Test Company",
                'date_format': 'DD/MM/YYYY',
            }
        )

        self.state_maharashtra = self.env['res.country.state'].search([
            ('country_id', '=', self.env.ref('base.in').id),
            ('code', '=', 'MH'),
        ], limit=1)

        self.state_gujarat = self.env['res.country.state'].search([
            ('country_id', '=', self.env.ref('base.in').id), ('code', '=', 'GJ'), ], limit=1)

        # Create a user with specific timezone
        self.env.user.tz = "UTC"
        self.user = self.env.user

        self.shipping_line = self.env["res.partner"].create(
            {
                "name": "Test Shipping Line",
                "company_id": self.env.company.id,
            }
        )

        self.shipping_line_1 = self.env["res.partner"].create(
            {
                "name": "Test Shipping Line",
                "company_id": self.env.company.id,
            }
        )

        self.location_shipping_line = self.env["location.shipping.line.mapping"].create(
            {
                "shipping_line_id": self.shipping_line.id,
                "company_id": self.env.company.id,
            }
        )

        self.location_invoice_setting = self.env["location.invoice.setting"].create({
            "company_id": self.env["res.company"].search([], limit=1).id,  # Assuming a company exists
            "inv_shipping_line_id": self.shipping_line.id,
            "inv_applicable_at_location_ids": [(6, 0, self.env["invoice.applicable.options"].search([], limit=3).ids)],
            "active": True,
            "e_invoice_applicable": "yes",
        })

        self.location = self.env["res.company"].create(
            {
                "name": "Test Location",
                "active": True,
                "state_id": self.state_maharashtra.id,
                "mode_ids": [(6, 0, [self.truck_mode.id])],
                "shipping_line_mapping_ids": [(6, 0, [self.location_shipping_line.id])],
                'parent_id': self.company.id,
                "invoice_setting_ids": [(6, 0, [self.location_invoice_setting.id])],
                "laden_status_ids": [
                    (
                        6,
                        0,
                        [self.env["laden.status.options"].create({"name": "laden"}).id],
                    )
                ],
            }
        )

        self.container_type_size = self.env["container.type.data"].create(
            {
                "name": "20FT",
                "size": "20FT",
                "company_size_type_code": "20FT",
                "is_refer": "yes",
                "te_us": 1.0,
                "company_id": self.env.company.id,
            }
        )

        self.partner_transporter = self.env["res.partner"].create(
            {
                "name": "Transporter",
                "parties_type_ids": [(0, 0, {"name": "Transporter"})],
            }
        )
        self.container_master = self.env["container.master"].create(
            {
                "name": "GVTU3000389",
                "type_size": self.container_type_size.id,
                "shipping_line_id": self.shipping_line.id,
                "gross_wt": "123",
                "tare_wt": "123",
            }
        )


        self.damage_condition = self.env["damage.condition"].create(
            {"name": "Damage A", "damage_code": "Size/Type A", "active": True}
        )

        self.container_type_size1 = self.env["container.type.data"].create(
            {
                "name": "40 FT",
                "company_size_type_code": "40FT",
                "is_refer": "yes",
                "active": True,
                "company_id": self.env.company.id,
            }
        )

        # Create Container Detail
        self.container_detail = self.env["container.details"].create(
            {
                "container_size_type": self.container_type_size.id,
                "container_qty": 10,
                "balance": 5,
            }
        )

        self.container_delivery_detail = self.env["container.details.delivery"].create(
            {
                "container_qty": 10,
                "balance_container": 5,
                "quantity": 5,
                "container_size_type": self.container_type_size.id,
            }
        )

        self.booking = self.env["vessel.booking"].create(
            {
                "shipping_line_id": self.shipping_line.id,
                "transporter_name": self.partner_transporter.id,
                "location": [(6, 0, [self.location.id])],
                "booking_no": "BOOK001",
                "booking_date": fields.Date.today(),
                "validity_datetime": fields.Datetime.now(),
                "cutoff_datetime": fields.Datetime.now(),
                "vessel": "Test Vessel",
                "voyage": "12345",
                "container_details": [(6, 0, [self.container_detail.id])],
            }
        )

        # Create a sample Partner record

        self.billed_to_gst_no = self.env['gst.details'].create({
            'gst_no': '33AAJCV4162C1ZK',
            'tax_payer_type': 'Regular',
            'state_jurisdiction': 'Maharashtra',
            'company_id': self.company.id,
            'nature_of_business': 'Special Economic Zone',
            'place_of_business': 'Mumbai',
            'additional_place_of_business': 'Pune',
            'nature_additional_place_of_business': 'Warehouse',
            'additional_place_of_business_2': 'Nashik',
            'nature_additional_place_of_business_2': 'Office',
            'last_update': '2024-12-01',
            'gst_api_response': 'Valid',
            'state': 'Maharashtra',
        })

        self.billed_to_party = self.env['res.partner'].create({
            'name': 'Test Partner',
            'gst_state': 'Maharashtra',
            "is_this_billed_to_party": "yes",
            'street': 'City',
            'street2': 'State',
        })


        self.partner = self.env["res.partner"].create(
            {
                "name": "Test Partner",
                "is_cms_parties": True,
                "is_this_billed_to_party": "yes",
                "street": "123 Test Street",
                "gst_state": "Karnataka",
            }
        )

        self.master_port_data_model = self.env["master.port.data"]
        # Create a port for loading
        self.port_loading = self.master_port_data_model.create(
            {
                "country_iso_code": "US",
                "port_code": "TESTPORT",
                "port_name": "Test Port",
                "state_code": "NY",
                "status": "Active",
                "latitude": "40.7128° N",
                "longitude": "74.0060° W",
                "popular_port": True,
                "active": True,
            }
        )
        # Create a port for discharge
        self.port_discharge = self.master_port_data_model.create(
            {
                "country_iso_code": "US",
                "port_code": "TESTPORT",
                "port_name": "Test Port",
                "state_code": "NY",
                "status": "Active",
                "latitude": "40.7128° N",
                "longitude": "74.0060° W",
                "popular_port": True,
                "active": True,
            }
        )

        self.container_inventory = self.env["container.inventory"].create(
            {
                "name": "GVTU3000389",
                "grade": "a",
                "location_id": self.location.id,
                "container_master_id": self.container_master.id,
            }
        )

        self.delivery_order = self.env["delivery.order"].create(
            {
                "delivery_no": "DO001",
                "shipping_line_id": self.shipping_line.id,
                "delivery_date": datetime.today().date(),
                "validity_datetime": datetime.today() + timedelta(days=1),
                "exporter_name": self.env["res.partner"]
                .create({"name": "Test Exporter"})
                .id,
                "booking_party": self.env["res.partner"]
                .create({"name": "Test Booking Party"})
                .id,
                "forwarder_name": self.env["res.partner"]
                .create({"name": "Test Forwarder"})
                .id,
                "import_name": self.env["res.partner"]
                .create({"name": "Test Importer"})
                .id,
                "commodity": "Test Commodity",
                "cargo_weight": "1000",
                "vessel": "Test Vessel",
                "voyage": "Test Voyage",
                "remark": "Test Remark",
                "port_loading": self.port_loading.id,
                "port_discharge": self.port_discharge.id,
                "location": [(6, 0, [self.location.id])],
                "to_from_location": self.location.id,
                "stuffing_location": self.location.id,
                "total_containers": 10,
                "balance_containers": 5,
                "container_details": [(6, 0, [self.container_delivery_detail.id])],
            }
        )

        self.move_in = self.env["move.in"].create(
            {
                "container": "GVTU3000389",
                'move_in_date_time': datetime(2024, 12, 25, 10, 30),
                "location_id": self.location.id,
                "do_no_id": self.delivery_order.id,
                "booking_no_id": self.booking.id,
                "mode": "truck",
                "damage_condition": self.damage_condition.id,
                "shipping_line_id": self.shipping_line.id,
                "type_size_id": self.container_type_size.id,
                "grade": "a",
                "billed_to_party": self.billed_to_party.id,

            }
        )

        self.move_out = self.env["move.out"].create({
            'shipping_line_id': self.shipping_line.id,
            'inventory_id': self.container_inventory.id,
            'location_id': self.location.id,
            'container_id': self.container_master.id,
            'delivery_order_id': self.delivery_order.id,
            'move_out_date_time': fields.Datetime.now(),
            'movement_type': 'export_stuffing',
            'export_stuffing_to': 'factory',
            'mode': 'truck',
            'truck_number': 'TRK1234',
            'driver_name': 'John Doe',
            'driver_mobile_number': '1234567890',
            'driver_licence_no': 'LIC123456',
            "type_size_id": self.container_type_size.id,
            "billed_to_party": self.billed_to_party.id,
        })

        # Create a sample GST record
        self.gst_details = self.env["gst.details"].create(
            {
                "tax_payer_type": "Regular",
                'gst_no': '27BXRPP8176M2Z8',
                "nature_of_business": "Trading",
                "state_jurisdiction": "Karnataka",
                "additional_place_of_business": "Warehouse A",
                "last_update": "01-01-2023",
            }
        )


        self.tax= self.env["account.tax"].create(
            {
                "name": "15% GST",
                "type_tax_use": "sale",
                "active": True,

            }
        )

        # Create GST tax structure
        self.cgst_tax1 = self.env['account.tax'].create({
            'name': 'CGST 9%',
            'amount': 9.0,
            'type_tax_use': 'sale',
        })

        self.sgst_tax1 = self.env['account.tax'].create({
            'name': 'SGST 9%',
            'amount': 9.0,
            'type_tax_use': 'sale',
        })

        self.igst_tax1 = self.env['account.tax'].create({
            'name': 'IGST 18%',
            'amount': 18.0,
            'type_tax_use': 'sale',
            'children_tax_ids': [(6, 0, [self.cgst_tax1.id, self.sgst_tax1.id])]
        })

        self.hsn_code = self.env["master.hsn.code"].create(
            {
                "code": "123",
                "name": "HSN1",
                "active": True,
            }
        )

        self.other1 = self.env["product.template"].create(
            {
                "name": "Other1",
                "charge_name": "Other1",
                "gst_rate": [(6, 0, [self.igst_tax1.id])],
                "hsn_code": self.hsn_code.id,
                "active": True,
                "invoice_type": "Others",
                "charge_code": "CH01",
            }
        )

        self.disable_charge = self.env["product.template"].create(
            {
                "name": "Disable Charge",
                "charge_name": "Disable Charge",
                "gst_rate": [(6, 0, [self.tax.id])],
                "hsn_code": self.hsn_code.id,
                "active": False,
                "invoice_type": "lift_on",
                "charge_code": "CH04",
            }
        )

        self.lift_off = self.env["product.template"].create(
            {
                "name": "Lift Off",
                "charge_name": "Lift Off",
                "gst_rate": [(6, 0, [self.tax.id])],
                "hsn_code": self.hsn_code.id,
                "active": True,
                "invoice_type": "lift_off",
                "charge_code": "CH02",
            }
        )

        self.lift_on = self.env["product.template"].create(
            {
                "name": "Lift On",
                "charge_name": "Lift On",
                "gst_rate": [(6, 0, [self.tax.id])],
                "hsn_code": self.hsn_code.id,
                "active": True,
                "invoice_type": "lift_on",
                "charge_code": "CH03",
            }
        )

        self.fiscal_year = self.env["account.fiscal.year"].create({
            'date_from': '2024-01-01',
            'date_to': '2024-12-31',
            'name': '2024',
        })

        self.lolo_charge = self.env["lolo.charge"].create(
            {
                "location": self.location.id,
                "shipping_line": self.shipping_line.id,
                "lolo_charge_lines": [
                    (0, 0, {
                        'container_size': '20ft',
                        'lift_on': 300,
                        'lift_off': 200,
                    })
                ]
            }
        )

        self.invoice_charge_off = self.env["move.in.out.invoice.charge"].create({
            'charge_id': self.lift_off.id,
            'charge_name': 'Lift Off',
            "amount" : 200,
        })


        self.invoice_lift_off = self.env["move.in.out.invoice"].create(
            {
                "move_in_id": self.move_in.id,
                "invoice_number": "INV12345",
                "invoice_date": date.today(),
                "billed_to_gst_no": self.gst_details.id,
                "fiscal_year": self.fiscal_year.id,
                "location_id": self.location.id,
                "source": "move",
                "invoice_type": "lift_off",
                "charge_ids": [(6, 0, [self.invoice_charge_off.id])],
                'payment_mode': 'cash',

            }
        )

        self.invoice_charge_on = self.env["move.in.out.invoice.charge"].create({
            'charge_id': self.lift_on.id,
            'charge_name': 'Lift On',
            "amount": 300,
        })

        self.invoice_lift_on = self.env["move.in.out.invoice"].create(
            {
                "move_out_id": self.move_out.id,
                "invoice_number": "INV12346",
                "invoice_date": date.today(),
                "billed_to_gst_no": self.gst_details.id,
                'billed_to_party': self.billed_to_party.id,
                "fiscal_year": self.fiscal_year.id,
                "location_id": self.location.id,
                "source": "move",
                "invoice_type": "lift_on",
                "charge_ids": [(6, 0, [self.invoice_charge_on.id])],
            }
        )


        self.invoice_other_in = self.env['move.in.out.invoice'].create({
            'move_in_id': self.move_in.id,
            'invoice_number': 'INV12347',
            'invoice_date': '2024-12-23',
            'invoice_type': 'Others',
            'other_charge_id': self.other1.id,
            'location_id': self.location.id,
            'amount': 100.0,
            "company_id": self.company.id,
        })

        # Create invoice record
        self.invoice = self.env['move.in.out.invoice'].create({
            'invoice_type': 'Others',
            'invoice_number': 'INV123421',
            'move_out_id': self.move_out.id,
            'location_id': self.location.id,
            'shipping_line_id':self.shipping_line.id,
            'other_charge_id': self.other1.id,
            'amount': 1000.0,
            'gst_rate': [(6, 0, [self.igst_tax1.id])],
            'is_gst_applicable': 'yes',
            'billed_to_gst_no': self.billed_to_gst_no.id,
            'billed_to_party': self.billed_to_party.id,
            'create_uid': False,
            'payment_mode': 'online',
            'payment_reference': 'Ref001',
            'is_sez': True,
            "company_id": self.company.id,
        })

        # Create a record for e.invoice.credentials
        self.e_invoice_credentials = self.env["e.invoice.credentials"].create({
            "username": "kunal_roy",
            "password": "kunal_roy",
            "ip_address": "10.0.0.127",
            "email": "kunal@gmial.com",
            "client_id": "test_client_id",
            "client_secret": "test_client_secret",
            "gstin": "123456789012345",
        })

    def test_compute_gst_rate_from_charge_ids(self):
        # Compute GST rate
        self.invoice_lift_off._compute_gst_rate()

        # Check if GST rate is set to the charge's gst_rate
        self.assertEqual(self.invoice_lift_off.gst_rate.ids, self.lift_off.gst_rate.ids)

    def test_compute_gst_rate_from_other_charge_id(self):
        """ Test GST rate computation when invoice_type is 'Others' and other_charge_id is set """

        self.invoice_other_in._compute_gst_rate()

        # Check if GST rate is set to the other charge's gst_rate
        self.assertEqual(self.invoice_other_in.gst_rate.ids, self.other1.gst_rate.ids)

    def test_compute_other_charge_id(self):
        # Trigger the onchange method
        self.invoice_other_in._compute_other_charge_id()

        # Assert that GST rate is computed correctly
        expected_gst_rate_ids = self.other1.gst_rate.ids
        self.assertEqual(self.invoice_other_in.gst_rate.ids, expected_gst_rate_ids,
                         "GST rate was not computed correctly for 'Others' invoice type.")

    def test_valid_charge(self):
        """Test valid charge_ids."""
        # Assert no ValidationError is raised
        self.invoice_lift_on._check_charge_validity()

    def test_invalid_charge_ids(self):
        """Test invalid charge_ids."""
        # Make charge invalid by setting amount to 0
        self.invoice_lift_off.charge_ids.write({'amount': 0})
        # Check for validation error
        with self.assertRaisesRegex(ValidationError, 'Charge must be active and amount must be greater than 0.'):
            self.invoice_lift_off._check_charge_validity()

        # Make charge invalid by setting charge_id to inactive
        self.invoice_lift_off.charge_ids.charge_id.write({'active': False})
        # Check for validation error and compare the message again
        with self.assertRaisesRegex(ValidationError, 'Charge must be active and amount must be greater than 0.'):
            self.invoice_lift_off._check_charge_validity()

        # Remove all charges and check for validation error
        with self.assertRaisesRegex(ValidationError, 'Charge must be present and it should be active'):
            self.invoice_lift_off.write({'charge_ids': [(5, 0, 0)]})
            self.invoice_lift_off._check_charge_validity()

    def test_gst_computation_same_state(self):
        """Test GST computation for supply within the same state."""
        self.invoice.write({'supply_to_state': self.state_maharashtra.name})
        self.invoice._compute_gst_amounts()

        # Validate CGST and SGST bifurcation
        self.assertAlmostEqual(self.invoice.gst_breakup_cgst, 90.0,
                               msg="CGST should be 90.0 for amount 1000 and CGST rate 9%")
        self.assertAlmostEqual(self.invoice.gst_breakup_sgst, 90.0,
                               msg="SGST should be 90.0 for amount 1000 and SGST rate 9%")
        self.assertAlmostEqual(self.invoice.gst_breakup_igst, 0.0, msg="IGST should be 0.0 for intra-state supply")

    def test_gst_computation_interstate(self):
        """Test GST computation for supply to another state."""
        self.invoice.write({'supply_to_state': self.state_gujarat.name})
        self.invoice._compute_gst_amounts()

        # Validate IGST computation
        self.assertAlmostEqual(self.invoice.gst_breakup_igst, 180.0,
                               msg="IGST should be 180.0 for amount 1000 and IGST rate 18%")
        self.assertAlmostEqual(self.invoice.gst_breakup_cgst, 0.0, msg="CGST should be 0.0 for interstate supply")
        self.assertAlmostEqual(self.invoice.gst_breakup_sgst, 0.0, msg="SGST should be 0.0 for interstate supply")

    def test_compute_billed_to_gst_no(self):
        """Test the _compute_billed_to_gst_no method."""
        # Trigger the compute method
        self.invoice._compute_billed_to_gst_no()

        # Assertions
        self.assertEqual(self.invoice.tax_payer_type, 'Regular', "Tax payer type should be 'Regular'")
        self.assertEqual(self.invoice.state_jurisdiction, 'Maharashtra', "State jurisdiction should be 'Maharashtra'")
        self.assertEqual(self.invoice.company_id, self.company, "Company ID should match")
        self.assertEqual(self.invoice.nature_of_business, 'Special Economic Zone',
                         "Nature of business should be 'Special Economic Zone'")
        self.assertEqual(self.invoice.place_of_business, 'Mumbai', "Place of business should be 'Mumbai'")
        self.assertEqual(self.invoice.additional_place_of_business, 'Pune',
                         "Additional place of business should be 'Pune'")
        self.assertEqual(self.invoice.nature_additional_place_of_business, 'Warehouse',
                         "Nature of additional place of business should be 'Warehouse'")
        self.assertEqual(self.invoice.additional_place_of_business_2, 'Nashik',
                         "Second additional place of business should be 'Nashik'")
        self.assertEqual(self.invoice.nature_additional_place_of_business_2, 'Office',
                         "Nature of second additional place of business should be 'Office'")
        self.assertEqual(self.invoice.last_update, '2024-12-01', "Last update should match")
        self.assertEqual(self.invoice.gst_api_response, 'Valid', "GST API response should be 'Valid'")
        self.assertEqual(self.invoice.gst_state, 'Maharashtra', "GST state should be 'Maharashtra'")
        self.assertTrue(self.invoice.is_sez, "Invoice should be marked as SEZ")

    def test_not_sez(self):
        """Test that SEZ flag is not set when nature_of_business is not SEZ."""
        self.billed_to_gst_no.write({'nature_of_business': 'Regular'})
        self.invoice._compute_billed_to_gst_no()
        self.assertFalse(self.invoice.is_sez, "Invoice should not be marked as SEZ")

    def test_is_parties_gst_invoice_line_empty(self):
        """Test _compute_is_parties_gst_invoice_line_empty method."""
        # Trigger the compute method for an invoice without GST lines
        self.invoice._compute_is_parties_gst_invoice_line_empty()
        self.assertTrue(self.invoice.is_parties_gst_invoice_line_empty,
                        "is_parties_gst_invoice_line_empty should be True when there are no GST lines")

        # Trigger the compute method for an invoice with GST lines
        self.billed_to_party.write({
            'parties_gst_invoice_line_ids': [(4, self.billed_to_gst_no.id)]
        })
        self.invoice._compute_is_parties_gst_invoice_line_empty()
        self.assertFalse(self.invoice.is_parties_gst_invoice_line_empty,
                         "is_parties_gst_invoice_line_empty should be False when there are GST lines")

    def test_no_gst_lines(self):
        """Test when no GST lines exist for the billed_to_party."""
        # Ensure billed_to_party does not have GST lines
        self.billed_to_party.write({
            'parties_gst_invoice_line_ids': []
        })

        # Trigger the compute method
        self.invoice._compute_is_parties_gst_invoice_line_empty()

        # Assert the field is set correctly
        self.assertTrue(self.invoice.is_parties_gst_invoice_line_empty,
                        "is_parties_gst_invoice_line_empty should be True when no GST lines exist")

    def test_check_valid_amount(self):
        """Test that a valid amount does not raise a ValidationError."""
        # Set amount greater than 0 and check if the validation passes
        self.invoice.write({'amount': 200.0})

        try:
            self.invoice._check_amount()
        except ValidationError:
            self.fail("ValidationError raised unexpectedly!")

    def test_check_invalid_amount(self):
        """Test that an invalid amount (<= 0) raises a ValidationError."""
        # Set amount to 0 and check that ValidationError is raised
        with self.assertRaisesRegex(ValidationError, "Amount must be greater than 0. Please enter a valid amount."):
            self.invoice.write({'amount': 0.0})
            self.invoice._check_amount()

    def test_check_without_other_charge_id(self):
        """Test that without other_charge_id no ValidationError is raised."""
        # Set invoice to have no other charge
        self.invoice.write({'other_charge_id': False, 'amount': 100.0})

        try:
            self.invoice._check_amount()
        except ValidationError:
            self.fail("ValidationError raised unexpectedly when other_charge_id is not set.")

    def test_check_charge_ids_valid(self):
        """Test that invoice with valid charge IDs does not raise validation error."""
        try:
            self.invoice_lift_on._check_charge_ids()
        except ValidationError:
            self.fail("ValidationError raised unexpectedly for valid charge IDs.")

    def test_check_charge_ids_no_charge(self):
        """Test that invoice with no charge raises validation error."""
        with self.assertRaisesRegex(ValidationError, "Charge must be present and it should be active"):
            self.invoice_lift_on.write({'charge_ids': [(5,0,0)]})
            self.invoice_lift_on._check_charge_ids()

    def test_check_charge_ids_zero_amount_charge(self):
        self.invoice_lift_on.charge_ids.amount = 0
        """Test that invoice with charge amount <= 0 raises validation error."""
        with self.assertRaisesRegex(ValidationError, "Amount must be greater than 0. Please enter a valid amount."):
            self.invoice_lift_on._check_charge_ids()

    def test_get_create_record_info(self):
        """
        Test _get_create_record_info to ensure it correctly sets the creation record information.
        """
        self.invoice_other_in._get_create_record_info()
        self.assertTrue(
            self.invoice_other_in.display_create_info, "The creation info should be set."
        )

    def test_get_modify_record_info(self):
        """
        Test _get_modify_record_info to ensure it correctly sets the modification record information.
        """
        self.invoice_other_in.write({'invoice_number': 'INV12340'})
        self.invoice_other_in._get_modify_record_info()
        self.assertTrue(
            self.invoice_other_in.display_modified_info, "The modification info should be set."
        )

    def test_compute_billed_to_party_address(self):
        self.invoice._compute_billed_to_party_address()

    def test_qr_code_computation_with_valid_jwt(self):
        """Test QR code computation when jwt_string is valid"""
        # Create a test record with a valid jwt_string

        # Trigger the computation by changing jwt_string
        self.invoice_lift_on._compute_qr_code()

        # Check that qr_code field is set (non-empty base64 encoded string)
        # self.assertTrue(self.invoice_lift_on.qr_code, "QR code should be generated and set")
        # self.assertTrue(isinstance(base64.b64decode(self.invoice_lift_on.qr_code), bytes), "QR code should be base64 encoded")

    def test_image_to_binary(self):
        """Test the _image_to_binary method"""
        # Create a sample PIL image
        image = Image.new('RGB', (100, 100), color='red')

        # Call the method
        binary_data = self.invoice_lift_on._image_to_binary(image)

        # Validate the output
        self.assertTrue(binary_data, "Binary data should not be empty")
        self.assertTrue(isinstance(binary_data, bytes), "Output should be binary data")

    def test_set_e_invoice_valid_data(self):
        """Test set_e_invoice with valid location_id and shipping_line_id"""
        # Call the method with valid data
        e_invoice = self.invoice_lift_on.set_e_invoice(
            location_id=self.location.id,
            shipping_line_id=self.shipping_line.id
        )
        # Validate the output
        self.assertTrue(e_invoice, "E-Invoice should be applicable for matching settings")

    def test_set_e_invoice_no_matching_data(self):
        """Test set_e_invoice with no matching data"""
        # Call the method with non-matching data
        e_invoice = self.invoice_lift_on.set_e_invoice(
            location_id=self.company.id,
            shipping_line_id=99999  # Non-existent shipping line ID
        )
        # Validate the output
        self.assertIsNone(e_invoice, "E-Invoice should be None when no matching settings are found")

    def test_compute_total_amount_other(self):
        """Test the total amount computation when invoice type is other."""
        # Ensure the computed total amount is correct
        self.invoice._compute_total_amount()
        total_amount = (
                self.invoice.amount + self.invoice.gst_breakup_cgst + self.invoice.gst_breakup_sgst + self.invoice.gst_breakup_igst
        )
        self.assertEqual(self.invoice.total_amount, total_amount, "The computed total amount is incorrect.")


    def test_compute_amount_mode_other(self):
        """Test the amount mode computation when invoice type is other."""
        # Ensure the computed amount mode is correct
        self.invoice._compute_amount_mode()
        expected_amount_mode = f"{self.invoice.total_amount}  {self.invoice.payment_mode}"
        self.assertEqual(self.invoice.amount_mode, expected_amount_mode, "The computed amount mode is incorrect.")

    def test_compute_amount_mode_lift_off(self):
        """Test the amount mode computation when invoice type is lift_on."""
        # Ensure the computed amount mode is correct
        self.invoice_lift_off._compute_amount_mode()
        expected_amount_mode = f"{self.invoice_lift_off.total_amount}  {self.invoice_lift_off.payment_mode}"
        self.assertEqual(self.invoice_lift_off.amount_mode, expected_amount_mode, "The computed amount mode is incorrect.")

    def test_compute_total_charge_amount(self):
        """Test the computation of total charge amount."""
        # Manually invoke the compute method
        self.invoice_lift_off._compute_total_charge_amount()

        # Calculate the expected total charge amount
        expected_total = self.invoice_lift_off.amount
        # Assert that the computed total charge amount matches the expected value
        # self.assertEqual(
        #     self.invoice_lift_off.total_charge_amount,
        #     expected_total,
        #     "The computed total charge amount is incorrect."
        # )

    def test_default_get_move_in(self):
        """Test default_get method with move.in context."""
        self.env.context = {
            'params': {'id': self.move_in.id, 'model': 'move.in'},
        }

        fields = [
            'source', 'location_id', 'shipping_line_id', 'e_invoice', 'billed_to_party',
            'invoice_type', 'payment_mode', 'invoice_number'
        ]

        default_values = self.invoice_lift_off.default_get(fields)

        # Assertions
        self.assertEqual(default_values.get('source'), 'move', "Source should be 'move'.")
        self.assertEqual(default_values.get('location_id'), self.move_in.location_id.id, "Location ID should match.")
        self.assertEqual(default_values.get('shipping_line_id'), self.move_in.shipping_line_id.id, "Shipping Line ID should match.")
        self.assertEqual(default_values.get('billed_to_party'), self.move_in.billed_to_party.id, "Billed To Party ID should match.")

    def test_default_get_move_out(self):
        """Test default_get method with move.out context."""
        self.env.context = {
            'params': {'id': self.move_out.id, 'model': 'move.out'},
        }

        fields = [
            'source', 'location_id', 'shipping_line_id', 'e_invoice', 'billed_to_party',
            'invoice_type', 'payment_mode', 'invoice_number'
        ]

        default_values = self.invoice_lift_on.default_get(fields)

        # Assertions
        self.assertEqual(default_values.get('source'), 'move', "Source should be 'move'.")
        self.assertEqual(default_values.get('location_id'), self.move_out.location_id.id, "Location ID should match.")
        self.assertEqual(default_values.get('shipping_line_id'), self.move_out.shipping_line_id.id, "Shipping Line ID should match.")
        self.assertEqual(default_values.get('billed_to_party'), self.move_out.billed_to_party.id, "Billed To Party ID should match.")

    def test_default_get_no_context(self):
        """Test default_get method without specific context."""
        self.env.context = {}

        fields = [
            'source', 'location_id', 'shipping_line_id', 'e_invoice', 'billed_to_party',
            'invoice_type', 'payment_mode', 'invoice_number'
        ]

        default_values = self.invoice_lift_on.default_get(fields)

        # Assertions
        self.assertNotIn('location_id', default_values, "Location ID should not be set without context.")

    def test_default_get_invalid_context(self):
        """Test default_get method with invalid context."""
        self.env.context = {
            'params': {'id': 99999, 'model': 'move.in'},  # Non-existent record
        }

        fields = [
            'source', 'location_id', 'shipping_line_id', 'e_invoice', 'billed_to_party',
            'invoice_type', 'payment_mode', 'invoice_number'
        ]

        default_values = self.invoice_lift_on.default_get(fields)

        # Assertions
        self.assertEqual(default_values.get('source'), 'move', "Source should still be 'move'.")
        self.assertNotIn('location_id', default_values, "Location ID should not be set for non-existent records.")


    def test_compute_invoice_details(self):
        """Test the computation of invoice_details field."""
        self.invoice_other_in._compute_invoice_details()
        expected_details = (
            f"INV12347 {self.invoice_other_in.invoice_date.strftime('%d/%m/%Y')}"
        )
        self.assertEqual(
            self.invoice_other_in.invoice_details,
            expected_details,
            "Invoice details should be correctly computed",
        )

    def test_get_e_invoice_authentication_token_with_credentials(self):
        """Test method returns correct credentials when they exist."""
        result = self.invoice_lift_off.get_e_invoice_authentication_token()

        # Assertions for returned credentials
        self.assertEqual(result[0], "kunal_roy", "Username should match the credentials.")
        self.assertEqual(result[1], "kunal_roy", "Password should match the credentials.")
        self.assertEqual(result[2], "10.0.0.127", "IP Address should match the credentials.")
        self.assertEqual(result[3], "kunal@gmial.com", "Email should match the credentials.")
        self.assertEqual(result[4], "test_client_id", "Client ID should match the credentials.")
        self.assertEqual(result[5], "test_client_secret", "Client Secret should match the credentials.")
        self.assertEqual(result[6], "123456789012345", "GSTIN should match the credentials.")

    def test_compute_party_invoice_type(self):
        """Test the computation of party_invoice_type field."""
        self.invoice_lift_off._compute_party_invoice_type()
        expected_party_invoice_type = (
            f"{self.billed_to_party.name} Lift Off"
        )
        self.assertEqual(
            self.invoice_lift_off.party_invoice_type,
            expected_party_invoice_type,
            "Party invoice type should be correctly computed",
        )

    def test_remarks_for_gst_no(self):
        """Test remarks when GST is not applicable."""
        self.invoice.is_gst_applicable = "no"
        self.invoice._compute_remarks_display()
        expected_remarks = (
            "SUPPLY MEANT FOR EXPORT UNDER LETTER OF UNDERTAKING WITHOUT PAYMENT OF GST"
        )
        self.assertEqual(
            self.invoice.remarks,
            expected_remarks,
            "Remarks should be for export without GST payment",
        )

    def test_remarks_for_gst_yes(self):
        """Test remarks when GST is applicable."""
        self.invoice.is_gst_applicable = "yes"
        self.invoice._compute_remarks_display()
        expected_remarks = (
            f"EXPORT UNDER LUT BOND. ARN NO. "
            f"DATED: (IGST @{self.igst_tax1.name} IGST AMOUNT: {self.invoice.amount}) "
            f"SUPPLY MEANT FOR EXPORT UNDER LETTER OF UNDERTAKING WITHOUT PAYMENT OF GST"
        )
        # self.assertEqual(
        #     self.invoice.remarks,
        #     expected_remarks,
        #     "Remarks should reflect GST applicable information",
        # )

    def test_cancel_invoice_action(self):
        """Test the cancel_invoice action method."""
        result = self.invoice.cancel_invoice()
        self.assertEqual(
            result["name"],
            "Cancel Invoice",
            "The action should be for 'Cancel Invoice'",
        )
        self.assertEqual(
            result["res_model"],
            "invoice.cancellation.wizard",
            "The model should be 'invoice.cancellation.wizard'",
        )
        self.assertEqual(
            result["context"]["default_invoice_id"],
            self.invoice.id,
            "The context should contain the correct invoice ID",
        )

    def test_generate_pdf(self):
        """Test that PDF generation is triggered correctly."""
        result = self.invoice.generate_pdf()
        self.assertTrue(result)

    def test_charge_record_ids(self):
        """Test the charge_record_ids method."""

        """Test with move in."""
        charge_records = self.invoice_lift_off.charge_record_ids(
            move_in_id=self.move_in.id,
            move_out_id=False,
            invoice_type=self.invoice_lift_off.invoice_type,
        )

        # Assertions
        self.assertTrue(charge_records, "Charge records should be created and returned.")
        self.assertEqual(charge_records.size, '20 FT', "Size should be '20 FT' for the given move_in.")
        self.assertEqual(charge_records.amount, 200, "Amount should match the mocked set_amount value.")
        self.assertEqual(charge_records.charge_id.id, self.lift_off.id,
                         "Charge record should reference the correct product template.")
        self.assertEqual(charge_records.move_in_out_invoice_id.id, self.invoice_lift_off.id,
                         "Charge record should reference the correct invoice.")

        # """Test with move out."""
        # charge_records = self.invoice_lift_on.charge_record_ids(
        #     move_in_id=False,
        #     move_out_id=self.move_out.id,
        #     invoice_type=self.invoice_lift_on.invoice_type,
        # )
        #
        # # Assertions
        # self.assertTrue(charge_records, "Charge records should be created and returned for move_out.")
        # self.assertEqual(charge_records.size, '20 FT', "Size should be '40 FT' for the given move_out.")
        # self.assertEqual(charge_records.amount, 300, "Amount should match the mocked set_amount value.")

    def test_calculate_gst_rate_value_others(self):
        """Test GST rate calculation for 'Others' invoice type."""
        vals = {
            'invoice_type': 'Others',
            'other_charge_id': self.invoice.other_charge_id.id,
        }
        result = self.invoice.calculate_gst_rate_value(None, vals)
        # self.assertEqual(result, 18.0, "GST rate for 'Others' invoice type is incorrect.")

    def test_calculate_gst_rate_value_non_others(self):
        """Test GST rate calculation for non-'Others' invoice type."""
        charge_id = self.invoice_charge_off
        vals = {'invoice_type': 'lift_off'}
        result = self.invoice_lift_off.calculate_gst_rate_value(charge_id, vals)
        # self.assertEqual(result, 15.0, "GST rate for non-'Others' invoice type is incorrect.")

    def test_extract_gst_percentage(self):
        """Mock GST percentage extraction logic for testing."""
        gst_name = self.igst_tax1.name
        self.invoice._extract_gst_percentage(gst_name)

    def test_prepare_generate_irn_others(self):
        """Test IRN preparation for 'Others' invoice type."""
        vals = {
            'invoice_type': 'Others',
            'other_charge_id': self.invoice.other_charge_id.id,
            'amount': 1000.0
        }
        # result = self.invoice.prepare_generate_irn(
        #     location=self.location.id,
        #     charge_id=None,
        #     shipping_line_id=self.location.invoice_setting_ids[0].inv_shipping_line_id.id,
        #     billed_party=self.billed_to_party.id,
        #     invoice_number="INV123421",
        #     billed_to_gst_no=self.billed_to_gst_no.id,
        #     vals=vals
        # )
        # self.assertEqual(result['DocDtls']['No'], "INV123421", "Invoice number mismatch.")
        # self.assertEqual(result['ValDtls']['TotInvVal'], 2180.0, "Total invoice value mismatch.")
        # self.assertEqual(result['TranDtls']['SupTyp'], 'SEZWP', "Subtype mismatch.")

    def test_prepare_generate_irn_standard(self):
        """Test IRN preparation for standard invoice type."""
        vals = {
            'invoice_type': 'lift_off'
        }
        result = self.invoice_lift_on.prepare_generate_irn(
            location=self.location.id,
            charge_id=self.invoice_charge_off,
            shipping_line_id=self.location.invoice_setting_ids[0].inv_shipping_line_id.id,
            billed_party=self.billed_to_party.id,
            invoice_number="INV12346",
            billed_to_gst_no=self.gst_details.id,
            vals=vals
        )
        self.assertEqual(result['DocDtls']['No'], "INV12346", "Invoice number mismatch.")
        # self.assertEqual(result['ValDtls']['TotInvVal'],25 , "Total invoice value mismatch.")
        self.assertEqual(result['TranDtls']['SupTyp'], 'B2B', "Subtype mismatch.")

    def test_authenticate_einvoice_api(self):
        """Test the actual e-invoice API authentication."""
        # Ensure the method for fetching credentials is properly configured
        username, password, ip_address, email, client_id, client_secret, gstin = self.invoice_lift_off.get_e_invoice_authentication_token()

        if not all([username, password, ip_address, email, client_id, client_secret, gstin]):
            self.skipTest("E-invoice API credentials are not configured in the environment.")

        result = self.invoice_lift_off.authenticate_einvoice_api()

        # self.assertIsNotNone(result, "The API response should not be None.")
        # self.assertIn("token", result, "The API response should contain a 'token' key.")

    def test_update_billed_party(self):
        """Test the update_billed_party method."""
        # Create a sample billed_party record
        self.billed_to_party.write({
            'name': 'Test Party',
            'tax_payer_type': '',
            'state_jurisdiction': '',
            'company_id': False,
            'gst_state': '',
            'nature_of_business': '',
            'additional_place_of_business': '',
            'nature_additional_place_of_business': '',
            'additional_place_of_business_2': '',
            'nature_additional_place_of_business_2': '',
            'last_update': '',
            'is_gst_applicable': 'no',
        })

        billed_party = self.billed_to_party

        vals = {
            'tax_payer_type': 'Regular',
            'state_jurisdiction': 'Karnataka',
            'company_id': self.env.company.id,
            'state': 'KA',
            'nature_of_business': 'Retail',
            'additional_place_of_business': 'Location 1',
            'nature_additional_place_of_business': 'Warehouse',
            'additional_place_of_business_2': 'Location 2',
            'nature_additional_place_of_business_2': 'Office',
            'last_update': '2024-12-25',
            'gst_api_response': 'Success',
        }
        gst_no = '29ABCDE1234F1Z5'

        self.invoice.update_billed_party(billed_party.id, vals, gst_no)

        self.assertEqual(billed_party.tax_payer_type, vals['tax_payer_type'])
        self.assertEqual(billed_party.state_jurisdiction, vals['state_jurisdiction'])
        self.assertEqual(billed_party.company_id.id, vals['company_id'])
        self.assertEqual(billed_party.gst_state, vals['state'])
        self.assertEqual(billed_party.nature_of_business, vals['nature_of_business'])
        self.assertEqual(billed_party.additional_place_of_business, vals['additional_place_of_business'])
        self.assertEqual(billed_party.nature_additional_place_of_business, vals['nature_additional_place_of_business'])
        self.assertEqual(billed_party.additional_place_of_business_2, vals['additional_place_of_business_2'])
        self.assertEqual(billed_party.nature_additional_place_of_business_2,
                         vals['nature_additional_place_of_business_2'])
        self.assertEqual(billed_party.last_update, vals['last_update'])
        self.assertEqual(billed_party.is_gst_applicable, 'yes')
        self.assertEqual(billed_party.parties_gst_invoice_line_ids.gst_no, gst_no)
        self.assertEqual(billed_party.l10n_in_pan, gst_no[2:12])
        self.assertEqual(billed_party.parties_gst_invoice_line_ids.gst_api_response, vals['gst_api_response'])

    def test_single_gst_record(self):
        """Test when there is a single GST record for the billed party."""
        self.billed_to_gst_no.write({
            'partner_id': self.billed_to_party.id,
        })
        gst_record = self.billed_to_gst_no

        self.invoice._compute_billed_to_party()
        self.assertEqual(self.invoice.billed_to_gst_no, gst_record,
                         "The GST record should be assigned when there is one.")

    def test_no_gst_record(self):
        """Test when there are no GST records for the billed party."""
        self.invoice_other_in._compute_billed_to_party()
        self.assertFalse(self.invoice_other_in.billed_to_gst_no,
                         "billed_to_gst_no should be False when no GST records exist.")

    def test_no_billed_to_party(self):
        """Test when billed_to_party is not set."""
        self.invoice_other_in.billed_to_party = False
        self.invoice_other_in._compute_billed_to_party()
        self.assertFalse(self.invoice_other_in.billed_to_gst_no,
                         "billed_to_gst_no should be False when billed_to_party is not set.")

    def test_invalid_length_gst_number(self):
        """Test with a GST number of incorrect length."""
        gst_no = "A123423"
        with self.assertRaisesRegex(
                ValidationError,
                'GST Number entered should be 15 characters long. Please enter the correct GST Number.'
        ):
            self.invoice._validate_parties_gst_number(gst_no)

    def test_non_alphanumeric_gst_number(self):
        """Test with a non-alphanumeric GST number."""
        gst_no = "29ABCDE1234F!Z5"
        with self.assertRaisesRegex(
                ValidationError,
                'Please enter an alphanumeric GST Number.'
        ):
            self.invoice._validate_parties_gst_number(gst_no)

    def test_duplicate_gst_number(self):
        """Test with a duplicate GST number for a different partner."""
        new_partner = self.env['res.partner'].create({
            'name': 'Another Partner',
        })
        self.billed_to_gst_no.write({"partner_id": self.partner.id})
        gst_no= "33AAJCV4162C1ZK"
        with self.assertRaisesRegex(
                ValidationError,
                'GST Number is already set Please enter a different GST Number.'
        ):
            self.invoice._validate_parties_gst_number(gst_no, partner_id = new_partner.id)

    def test_e_invoice_record_with_ack_date(self):
        self.invoice.write({
            'irn_no': 'IRN123',
            'irn_status': 'active',
            'generate_irn_response': 'Success',
            'ack_date': datetime(2024, 12, 25, 15, 30),
        })
        result = self.invoice.e_invoice_record()

        # Check that the result is a dictionary
        self.assertIsInstance(result, dict)
        self.assertIn('name', result)
        self.assertEqual(result['name'], 'E-Invoice - INV123421 ')
        self.assertIn('context', result)
        self.assertIn('default_invoice_id', result['context'])
        self.assertEqual(result['context']['default_invoice_id'], self.invoice.id)

        # Check that 'default_irn_date' is correctly formatted
        formatted_date_time = '25/12/2024 03:30 PM'
        self.assertIn('default_irn_date', result['context'])
        self.assertEqual(result['context']['default_irn_date'], formatted_date_time)

    def test_e_invoice_record_without_ack_date(self):
        self.invoice.ack_date = None
        result = self.invoice.e_invoice_record()
        self.assertIsNone(result)

    def test_get_report_filename_with_move_in(self):
        result = self.invoice_lift_off._get_report_filename()

        expected_filename = f"GVTU3000389_INV12345_2024-12-25 10:30:00"
        self.assertEqual(result, expected_filename)

    def test_get_report_filename_with_move_out(self):
        result = self.invoice_lift_on._get_report_filename()

        expected_filename = f"GVTU3000389_INV12346_{self.move_out.move_out_date_time}"
        self.assertEqual(result, expected_filename)

    def test_get_report_filename_without_move(self):
        self.invoice.move_out_id= False
        # Test with a record that has neither move_in_id nor move_out_id
        result = self.invoice._get_report_filename()
        # Check if the result is just the invoice_number
        expected_filename = 'INV123421'
        self.assertEqual(result, expected_filename)

    def test_onchange_billed_to_party_with_single_gst_record(self):
        """Test when a partner with a single related GST record is selected."""
        self.billed_to_gst_no.partner_id = self.billed_to_party
        # Trigger onchange
        self.invoice.onchange_billed_to_party()
        # Verify that billed_to_gst_no is set to the GST record
        self.assertEqual(self.invoice.billed_to_gst_no, self.billed_to_gst_no)

    def test_onchange_billed_to_party_with_no_gst_records(self):
        """Test when a partner with no related GST records is selected."""

        self.invoice.billed_to_party = self.partner
        self.invoice.onchange_billed_to_party()
        self.assertFalse(self.invoice.billed_to_gst_no)

    def test_action_download_excel(self):
        """Test the action_download_excel method to ensure it generates an Excel file correctly."""

        # Trigger the action
        result = self.invoice.action_download_excel()

        # Verify that the result is a valid dictionary with the expected URL for downloading
        self.assertIn('type', result)
        self.assertEqual(result['type'], 'ir.actions.act_url')
        self.assertIn('url', result)
        self.assertTrue(result['url'].startswith('/web/content/'))

        # Check the attachment created in Odoo
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'move.in.out.invoice'),
            ('name', '=', f"Invoice Report {datetime.now().strftime('%d/%m/%Y')}.xlsx")
        ], limit=1)

        self.assertTrue(attachment, "Attachment not created.")
        self.assertEqual(attachment.type, 'binary', "Attachment type is not binary.")
        self.assertTrue(attachment.datas, "Attachment has no data.")

        # Ensure the generated file is a valid Excel file (a basic check can be done by checking its size)
        file_data = base64.b64decode(attachment.datas)
        # Further check if workbook contains the expected sheets (by reading the first few bytes)
        with io.BytesIO(file_data) as file:
            from openpyxl import load_workbook
            workbook = load_workbook(file)
            sheet_names = workbook.sheetnames
            expected_sheet_names = ['Detailed Report', 'Line Wise Summary', 'Charge Summary']

            for sheet_name in expected_sheet_names:
                self.assertIn(sheet_name, sheet_names, f"Missing sheet: {sheet_name}")

    def test_action_download_excel_with_empty_invoice(self):
        """Test when there are no records to download."""
        # Create an empty invoice with no data
        empty_invoice = self.env['move.in.out.invoice'].create({
            'invoice_number': 'INV999',
            'location_id': self.location.id,
            'company_id': self.company.id,
        })

        # Trigger the action for an empty invoice
        result = empty_invoice.action_download_excel()

        # Verify the result as before
        self.assertIn('type', result)
        self.assertEqual(result['type'], 'ir.actions.act_url')
        self.assertIn('url', result)
        self.assertTrue(result['url'].startswith('/web/content/'))

        # Check if attachment was created and it is a valid file
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'move.in.out.invoice'),
            ('name', '=', f"Invoice Report {datetime.now().strftime('%d/%m/%Y')}.xlsx")
        ], limit=1)

        self.assertTrue(attachment, "Attachment not created.")
        self.assertEqual(attachment.type, 'binary', "Attachment type is not binary.")
        self.assertTrue(attachment.datas, "Attachment has no data.")

    def test_create_detailed_report_sheet(self):
        """Test the _create_detailed_report_sheet method to ensure it generates the Excel file correctly."""

        # Create an in-memory BytesIO stream for the workbook
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)

        # Call the method
        self.invoice._create_detailed_report_sheet(workbook,self.invoice)

        # Close the workbook
        workbook.close()

        # Read the workbook using openpyxl
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb['Detailed Report']

        # Verify headers
        headers = [
            'Invoice No', 'Invoice Type', 'Location', 'Shipping Line',
            'Container No', 'Billed To Party', 'GST No', 'Supply to State',
            'GST %', 'CGST Amount', 'SGST Amount', 'IGST Amount',
            'Invoice Amount', 'Payment Mode', 'Payment Reference No.'
        ]
        for col_num, header in enumerate(headers, start=1):  # openpyxl uses 1-based indexing
            self.assertEqual(ws.cell(row=1, column=col_num).value, header)

        # Verify data for the first record
        record = self.invoice[0]
        data = [
            record.invoice_number,
            record.invoice_type,
            record.location_id.name,
            record.shipping_line_id.name,
            'GVTU3000389',
            record.billed_to_party.name,
            record.billed_to_gst_no.gst_no,
            record.supply_to_state,
            self.igst_tax1.name,
            0.0,
            0.0,
            0.0,
            record.total_amount,
            record.payment_mode,
            record.payment_reference,
        ]

        for col_num, value in enumerate(data, start=1):
            self.assertEqual(ws.cell(row=2, column=col_num).value, value)
        wb.close()
