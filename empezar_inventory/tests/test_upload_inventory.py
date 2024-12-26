import logging
import io
import base64
from pandas import DataFrame
import pandas as pd
import openpyxl
from odoo.tests.common import TransactionCase
from odoo.exceptions import ValidationError


_logger = logging.getLogger(__name__)


class TestUploadInventory(TransactionCase):

    def setUp(self):
        super().setUp()

        self.test_data = {
            "Container No.": ["12365478921", "12365478922"],
            "Shipping Line": [False, False],
            "Damage Condition": ["Damage A", "Damage B"],
            "Gross Wt.": [False, False],
            "Tare Wt.": [False, False],
            "Size/Type": ["Size/Type A", "Size/Type B"],
            "Production Month/Year": ["05/2024", "07/2024"],
            "Status": ["AE", "AA"],
            "Grade": ["A", "B"],
            "In Date": ["12/01/2023", "10/01/2023"],
            "In Time": ["08:00:00", "10:00:00"],
            "Estimate Date": ["20/01/2023", "12/01/2023"],
            "Approval Date": ["25/01/2023", "14/01/2023"],
            "Estimate Amt": [1000.0, 2000.0],
            "Approved Amount": [800.0, 1800.0],
            "Repair Date": ["26/01/2023", "16/01/2023"],
            "Remarks": ["Test remark 1", "Test remark 2"],
        }

        self.shipping_line1 = self.env["res.partner"].create(
            {"name": "Shipping Line A", "is_shipping_line": True, "active": True}
        )
        self.shipping_line2 = self.env["res.partner"].create(
            {"name": "Shipping Line B", "is_shipping_line": True, "active": True}
        )

        self.container_type_data1 = self.env["container.type.data"].create(
            {"company_size_type_code": "Size/Type A", "active": True}
        )
        self.container_type_data2 = self.env["container.type.data"].create(
            {"company_size_type_code": "Size/Type B", "active": True}
        )

        self.damage_condition1 = self.env["damage.condition"].create(
            {"name": "Damage A", "damage_code": "Size/Type A", "active": True}
        )
        self.damage_condition2 = self.env["damage.condition"].create(
            {"name": "Damage B", "damage_code": "Size/Type B", "active": True}
        )

    def test_set_upload_id(self):
        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "upload_inventory_file": False,  # Placeholder, actual file not needed for this test
                "location_id": self.env.ref("base.main_company").id,
            }
        )
        self.assertTrue(upload_record.upload_id, "Upload ID not set correctly")
        upload_record.upload_id = "custom_upload_id"
        upload_record.set_upload_id()
        self.assertEqual(
            upload_record.upload_id,
            "custom_upload_id",
            "Upload ID should not be modified",
        )

    def test_action_submit(self):
        test_data = self.test_data
        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)

        base64_content = base64.b64encode(excel_file.read())
        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()

        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        self.assertNotIn("Error", uploaded_record_sheet.columns)

    def test_check_shipping_lines_validations_no_active_lines(self):

        self.shipping_line1.active = False
        self.shipping_line2.active = False
        test_data = self.test_data

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))

        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column

        with self.assertRaises(ValidationError) as e:
            upload_record.check_shipping_lines_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        self.assertEqual(e.exception.args[0], "There is no any active shipping lines")

    def test_check_shipping_lines_validations_invalid_lines(self):
        test_data = self.test_data
        test_data["Shipping Line"] = ["Shipping Line A", "Shipping Line AC"]
        df = DataFrame(test_data)

        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0], "> Shipping Line not present in master."
        # )
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][1], "> Shipping Line not present in master."
        # )

    def test_check_company_size_type_code_validations_no_active_record(self):

        test_data = self.test_data
        self.container_type_data1.active = False
        self.container_type_data2.active = False
        test_data["Size/Type"] = ["Size/Type A", "Size/Type AB"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))

        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column

        with self.assertRaises(ValidationError) as e:
            upload_record.check_company_size_type_code_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        self.assertEqual(
            e.exception.args[0],
            "There is no any active records for company size type code.",
        )

        upload_record.action_submit()
        # uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        # uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(uploaded_record_sheet['Error'][0], '> Shipping Line not present in master.')
        # self.assertEqual(uploaded_record_sheet['Error'][1], '> Shipping Line not present in master.')

    def test_check_company_size_type_code_validations_no_record_found(self):

        test_data = self.test_data
        test_data["Size/Type"] = ["Size/Type AC", "Size/Type AB"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0], "> size/type not present in master."
        # )
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][1], "> size/type not present in master."
        # )

    def test_check_container_number_validations_duplicate_number(self):

        test_data = self.test_data
        test_data["Container No."] = ["GVTU3000389", "GVTU3000389"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        self.assertEqual(
            uploaded_record_sheet["Error"][1], "> duplicate container number found."
        )

    def test_check_weight_column_validation_non_numeric(self):
        test_data = self.test_data
        test_data["Gross Wt."] = ["TEST", False]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Only numeric value allow in the Gross Wt..",
        # )

    def test_check_without_weight_column(self):
        test_data = self.test_data
        del test_data["Gross Wt."]
        df = DataFrame(test_data)

        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        file_data = io.BytesIO(uploaded_record)

        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column

        with self.assertRaises(ValidationError) as e:
            upload_record.check_gross_and_tare_weight_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        # self.assertEqual(
        #     e.exception.args[0], 'Column "Gross Wt." not found in the uploaded file.'
        # )

    def test_check_move_in_date_validations_date_formate_one(self):
        get_company = self.env["res.company"].search(
            [("parent_id", "=", False), ("active", "=", True)], limit=1
        )
        if get_company:
            get_company.date_format = "YYYY/MM/DD"

        test_data = self.test_data
        test_data["In Date"] = ["01/12/2023", "01/10/2023"]
        df = DataFrame(test_data)

        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Please enter In Date in YYYY/MM/DD format.Please enter Move in date in format as per company setting.",
        # )
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][1],
        #     ">Please enter In Date in YYYY/MM/DD format.Please enter Move in date in format as per company setting.",
        # )

    def test_check_move_in_date_validations_date_formate_two(self):

        get_company = self.env["res.company"].search(
            [("parent_id", "=", False), ("active", "=", True)], limit=1
        )
        if get_company:
            get_company.date_format = "MM/DD/YYYY"

        test_data = self.test_data
        test_data["In Date"] = ["2023/12/01", "2023/01/10"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Please enter In Date in MM/DD/YYYY format.Please enter Move in date in format as per company setting.",
        # )
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][1],
        #     ">Please enter In Date in MM/DD/YYYY format.Please enter Move in date in format as per company setting.",
        # )

    def test_check_move_in_date_validations_date_formate_without_in_date_cloumn(self):
        test_data = self.test_data
        del test_data["In Date"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)

        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))

        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        # remarks_column_index = None
        # for col in sheet.iter_cols(min_row=1, max_row=1):
        #     for cell in col:
        #         if cell.value == "Remarks":
        #             remarks_column_index = cell.column
        #
        # with self.assertRaises(ValidationError) as e:
        #     upload_record.check_move_in_date_validations(
        #         uploaded_record_sheet, sheet, remarks_column_index
        #     )
        # self.assertEqual(
        #     e.exception.args[0], 'Column "In Date" not found in the uploaded file.'
        # )

    def test_check_damage_validations_no_active_record(self):
        test_data = self.test_data
        self.damage_condition1.active = False
        self.damage_condition2.active = False

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Damage Condition not present in Damage seeder..",
        # )
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][1],
        #     ">Damage Condition not present in Damage seeder..",
        # )

    def test_estimate_approve_and_repair_dates_validations_no_approval_date(self):
        test_data = self.test_data
        del test_data["Approval Date"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))  # df

        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column
        # with self.assertRaises(ValidationError) as e:
        #     upload_record.estimate_approve_and_repair_dates_validations(
        #         uploaded_record_sheet, sheet, remarks_column_index
        #     )
        # self.assertEqual(
        #     e.exception.args[0],
        #     'Column "Approval Date" not found in the uploaded file.',
        # )

    def test_estimate_approve_and_repair_dates_validations_no_estimate_date(self):
        test_data = self.test_data
        del test_data["Estimate Date"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))  # df
        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column

        with self.assertRaises(ValidationError) as e:
            upload_record.estimate_approve_and_repair_dates_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        self.assertEqual(
            e.exception.args[0],
            'Column "Estimate Date" not found in the uploaded file.',
        )

    def test_estimate_approve_and_repair_dates_validations_no_repair_date(self):
        test_data = self.test_data
        del test_data["Repair Date"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))  # df
        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column
        with self.assertRaises(ValidationError) as e:
            upload_record.estimate_approve_and_repair_dates_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        self.assertEqual(
            e.exception.args[0], 'Column "Repair Date" not found in the uploaded file.'
        )

    def test_estimate_approve_and_repair_dates_validations_approval_estimate_date(self):
        test_data = self.test_data
        test_data["Estimate Date"] = ["20/01/2023", "12/01/2023"]
        test_data["Approval Date"] = ["19/01/2023", "14/01/2023"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Approval Date cannot be less than Estimate Date.",
        # )

    def test_estimate_approve_and_repair_dates_validations_estimate_repair_date(self):
        test_data = self.test_data
        test_data["Estimate Date"] = ["20/01/2023", "12/01/2023"]
        test_data["Repair Date"] = ["10/01/2023", "16/01/2023"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     ">Repair Date cannot be less than Estimate Date.\n>Repair Date cannot be less than Approval Date.",
        # )

    def test_check_production_month_year_validations_no_production_date(self):
        test_data = self.test_data
        del test_data["Production Month/Year"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )
        uploaded_record = base64.b64decode(base64_content)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))

        file_data = io.BytesIO(uploaded_record)
        wb = openpyxl.load_workbook(file_data)
        sheet = wb.active

        remarks_column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Remarks":
                    remarks_column_index = cell.column

        with self.assertRaises(ValidationError) as e:
            upload_record.check_production_month_year_validations(
                uploaded_record_sheet, sheet, remarks_column_index
            )
        # self.assertEqual(
        #     e.exception.args[0],
        #     'Column "Production Month/Year" not found in the uploaded file.',
        # )

    def test_validate_production_month_year_format_invalid_format(self):
        test_data = self.test_data
        test_data["Production Month/Year"] = ["05/1920", "2024/00"]

        df = DataFrame(test_data)
        excel_file = io.BytesIO()
        df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        base64_content = base64.b64encode(excel_file.read())

        upload_record = self.env["upload.inventory"].create(
            {
                "name": "Test File",
                "rec_status": "in_progress",
                "upload_inventory_file": base64_content,
                "location_id": self.env.ref("base.main_company").id,
            }
        )

        upload_record.action_submit()
        uploaded_record = base64.b64decode(upload_record.upload_inventory_file)
        uploaded_record_sheet = pd.read_excel(io.BytesIO(uploaded_record))
        # self.assertEqual(
        #     uploaded_record_sheet["Error"][0],
        #     '>Production Month/Year" in MM/YYYY format.Year must be between 1950 to 3999.',
        # )
