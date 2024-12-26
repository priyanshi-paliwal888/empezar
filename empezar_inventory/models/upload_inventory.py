# -*- coding: utf-8 -*-

from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
import io
import re
import base64
import pandas as pd
import logging
import openpyxl
from openpyxl.styles import Font
from datetime import datetime


_logger = logging.getLogger(__name__)


class UploadInventory(models.Model):

    _name = "upload.inventory"
    _description = "Upload Inventory"

    name = fields.Char(string="File Name", required=True)
    upload_inventory_file = fields.Binary(string="Upload File",
                                          help="Only XLS or XLSX files with max size of 5 MB and"
                                                                        " having max 200 entries.",
                                          required=True, copy=False, exportable=False)
    upload_id = fields.Char(string="Upload ID", compute='set_upload_id')
    uploaded_by = fields.Char(string="Uploaded By")
    uploaded_on = fields.Datetime(string="Uploaded On")
    rec_status = fields.Selection([
        ('success', 'Success'),
        ('error', 'Error'),
        ('in_progress', 'In Progress')
    ], string="Status")
    location_id = fields.Many2one('res.company', string="Location", required=True, domain="[('parent_id','!=',False)]")

    @api.model
    def download_xlsx_file(self, **kwargs):
        return {
            'type': 'ir.actions.act_url',
            'url': '/empezar_inventory/static/src/document/Upload_Inventory_Sample.xlsx',
            'target': 'new',
        }

    def set_upload_id(self):
        """
        Set upload id values.
        :return:
        """
        for rec in self:
            if not rec.upload_id:
                rec.upload_id = rec.id
            else:
                pass

    def action_submit(self):
        """
        This method is called from schedule actions.
        Check validations for uploaded Excel files and create records in container master and upload inventory
        model if all validations pass.
        :return:
        """
        active_records = self.search([('rec_status', '=', 'in_progress')], limit=2)
        if active_records:
            for rec in active_records:
                # Read the uploaded Excel file
                try:
                    file_content = base64.b64decode(rec.upload_inventory_file)
                    df = pd.read_excel(io.BytesIO(file_content))

                    file_data = io.BytesIO(base64.b64decode(rec.upload_inventory_file))
                    wb = openpyxl.load_workbook(file_data)
                    sheet = wb.active  # Assuming data is in the first sheet

                    # Find the index of the "Remarks" column (assuming it exists)
                    remarks_column_index = None
                    for col in sheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == "Remarks":
                                remarks_column_index = cell.column

                    # Delete any column after "Remarks" if it is empty
                    next_column_index = remarks_column_index + 1
                    sheet.delete_cols(next_column_index)

                    # Insert a column after "Remarks" for "Error"
                    error_column_index = remarks_column_index + 1
                    sheet.insert_cols(error_column_index)

                    # Set the header for the newly inserted column ("Error")
                    sheet.cell(row=1, column=error_column_index, value="Error").font = Font(bold=True)

                    # Save the workbook back to BytesIO or to a file
                    # If saving to BytesIO:
                    file_output = io.BytesIO()
                    wb.save(file_output)
                    file_output.seek(0)

                    rec.check_shipping_lines_validations(df, sheet, error_column_index)
                    rec.check_company_size_type_code_validations(df, sheet, error_column_index)
                    UploadInventory.check_container_number_validations(df, sheet, error_column_index)
                    UploadInventory.check_status_validations(df, sheet, error_column_index)
                    UploadInventory.check_grade_validations(df, sheet, error_column_index)
                    rec.check_damage_validations(df, sheet, error_column_index)
                    UploadInventory.check_gross_and_tare_weight_validations(df, sheet, error_column_index)
                    UploadInventory.check_approved_amount_validations(df, sheet, error_column_index)
                    UploadInventory.check_estimate_amount_validations(df, sheet, error_column_index)
                    UploadInventory.estimate_approve_and_repair_dates_validations(df, sheet, error_column_index)
                    rec.check_move_in_date_validations(df, sheet, error_column_index)
                    UploadInventory.check_in_time_validations(df, sheet, error_column_index)
                    UploadInventory.check_production_month_year_validations(df, sheet, error_column_index)
                    rec.record_creation_for_container_master(df, sheet, error_column_index)
                    rec.update_inventory_records(wb, sheet, error_column_index)
                except Exception as e:
                    _logger.error('There is some issue while update a inventory in %s\n%s', rec.upload_id, e)
        else:
            _logger.info("There is no any in progress records")

    def check_shipping_lines_validations(self, df, sheet, error_column_index):
        """
        Shipping line columns validations.
        :return:
        """
        shipping_line_data = df['Shipping Line'].dropna()
        get_active_shipping_lines = self.env['res.partner'].search([('is_shipping_line', '=', True),
                                                                    ('active', '=', True)]).mapped('shipping_line_name')
        if not get_active_shipping_lines:
            raise ValidationError(_('There is no any active shipping lines'))
        for index, rec in shipping_line_data.items():
            if not rec in get_active_shipping_lines:
                get_data = sheet.cell(row=index+2, column=error_column_index).value
                values = get_data + '\n> Shipping Line not present in master.' \
                    if get_data else '> Shipping Line not present in master.'
                sheet.cell(row=index+2, column=error_column_index, value=values)

    def check_company_size_type_code_validations(self, df, sheet, error_column_index):
        """
        company size type code validations.
        :return:
        """
        company_size_type_data = df['Size/Type'].dropna()
        get_active_size_type_code = self.env['container.type.data'].search([('active', '=', True)]).mapped(
            'company_size_type_code')
        if not get_active_size_type_code:
            raise ValidationError(_('There is no any active records for company size type code.'))
        for index, rec in company_size_type_data.items():
            if not rec in get_active_size_type_code:
                get_data = sheet.cell(row=index + 2, column=error_column_index).value
                values = get_data + '\n> size/type not present in master.' \
                    if get_data else '> size/type not present in master.'
                sheet.cell(row=index + 2, column=error_column_index, value=values)

    @staticmethod
    def check_container_number_validations(df, sheet, error_column_index):
        """
        Container number columns validations.
        :return:
        """
        # Check for duplicate container numbers
        duplicate_containers = df['Container No.'].dropna()[df['Container No.'].dropna().duplicated()]

        for index, rec in duplicate_containers.items():
            get_data = sheet.cell(row=index + 2, column=error_column_index).value
            values = get_data + '\n> duplicate container number found.' \
                if get_data else '> duplicate container number found.'
            sheet.cell(row=index + 2, column=error_column_index, value=values)

        # Create a list to store tuples of (container name, row number)
        container_info = []

        # Iterate over non-NaN values in 'Container No.' column and collect container name with row number
        for index, row in df.dropna(subset=['Container No.']).iterrows():
            container_info.append((row['Container No.'], index))

        for rec in container_info:
            container_number = rec[0]
            UploadInventory.check_digit_validation_for_container(container_number,rec,error_column_index,sheet)
            if len(str(rec[0])) != 11 or (not str(rec[0]).isalnum()):
                get_data = sheet.cell(row=int(rec[1]) + 2, column=error_column_index).value
                values = get_data + '\n> invalid container number.' \
                    if get_data else '> invalid container number.'
                sheet.cell(row=int(rec[1]) + 2, column=error_column_index, value=values)

    def check_digit_validation_for_container(container,rec,error_column_index,sheet):
        if container:
            char_to_num_dict = {
                'A': 10, 'B': 12, 'C': 13, 'D': 14, 'E': 15, 'F': 16, 'G': 17, 'H': 18, 'I': 19,
                'J': 20, 'K': 21, 'L': 23, 'M': 24, 'N': 25, 'O': 26, 'P': 27, 'Q': 28, 'R': 29,
                'S': 30, 'T': 31, 'U': 32, 'V': 34, 'W': 35, 'X': 36, 'Y': 37, 'Z': 38
                }
            input_data = str(container).upper()
            sliced_input_data = input_data[:10]
            total_sum = sum(
                (char_to_num_dict.get(char) if char_to_num_dict.get(char) else int(char)) * (2 ** index)
                for index, char in enumerate(sliced_input_data)
            )
            rounded_division_result = (int(total_sum/11)) * 11
            remainder = total_sum - rounded_division_result
            new_digit = remainder % 10
            if new_digit != eval(container[-1]):
                get_data = sheet.cell(row=int(rec[1]) + 2, column=error_column_index).value
                values = get_data + '\n> invalid container number.' \
                    if get_data else '> invalid container number.'
                sheet.cell(row=int(rec[1]) + 2, column=error_column_index, value=values)

    @staticmethod
    def validate_column(column_name, valid_values, error_message, df, sheet, error_column_index):
        """
        helper method to validate the column.
        :param column_name:
        :param valid_values:
        :param error_message:
        :param df:
        :param sheet:
        :param error_column_index:
        :return:
        """

        # Get column data without NaN values
        column_data = df[column_name].dropna()

        # Find invalid values
        invalid_values = column_data[~column_data.isin(valid_values)]

        if not invalid_values.empty:
            # Convert all invalid values to strings and join them
            invalid_value_index = invalid_values.index.tolist()
            for rec in invalid_value_index:
                get_data = sheet.cell(row=rec + 2, column=error_column_index).value
                values = get_data + f'\n>{error_message}.' if get_data else f'>{error_message}.'
                sheet.cell(row=rec + 2, column=error_column_index, value=values)

    @staticmethod
    def check_status_validations(df, sheet, error_column_index):
        """
        Check status column validations.
        :return:
        """
        valid_status = {'AE', 'AA', 'AR', 'AV', 'DAV'}
        error_message = 'Please enter only "AE", "AA", "AR", "AV", or "DAV" as Status'
        UploadInventory.validate_column('Status', valid_status, error_message, df, sheet, error_column_index)

    @staticmethod
    def check_grade_validations(df, sheet, error_column_index):
        """
        Check grade column validations.
        :return:
        """
        valid_grades = {'A', 'B', 'C'}
        error_message = 'Please enter only "A", "B", "C" as Grade.'
        UploadInventory.validate_column('Grade', valid_grades, error_message, df, sheet, error_column_index)

    def check_damage_validations(self, df, sheet, error_column_index):
        """
        Check damage condition column validations.
        :return:
        """
        valid_damage_condition = set(self.env['damage.condition'].search([('active', '=', True)]).mapped('name'))
        if not valid_damage_condition:
            valid_damage_condition = {''}
        error_message = 'Damage Condition not present in Damage seeder.'
        UploadInventory.validate_column('Damage Condition', valid_damage_condition, error_message, df, sheet,
                                        error_column_index)

    @staticmethod
    def check_weight_column_validation(df, column_name, sheet, error_column_index):
        """
        Check weight column validations.
        :param column_name:
        :param df:
        :param sheet:
        :param error_column_index:
        :return:
        """
        if column_name not in df.columns:
            raise ValidationError(
                _('Column "%s" not found in the uploaded file.') % column_name
            )

        weight_data = df[column_name].dropna()

        # Check for non-numeric values
        non_numeric_values = weight_data[~weight_data.apply(lambda x: isinstance(x, (int, float)))]

        if not non_numeric_values.empty:
            # Convert all invalid values to strings and join them
            invalid_value_index = non_numeric_values.index.tolist()
            for rec in invalid_value_index:
                get_data = sheet.cell(row=int(rec) + 2, column=error_column_index).value
                values = get_data + f'\n>Only numeric value allow in the {column_name}' \
                    if get_data else f'>Only numeric value allow in the {column_name}.'
                sheet.cell(row=int(rec) + 2, column=error_column_index, value=values)

    @staticmethod
    def compare_gross_weight_and_tare_weight(df, sheet, error_column_index):
        # Check if Tare Weight is greater than Gross Weight
        gross_weight_data = df['Gross Wt.'].dropna()
        tare_weight_data = df['Tare Wt.'].dropna()

        # Check for non-numeric values
        gross_non_numeric_values = gross_weight_data[~gross_weight_data.apply(lambda x: isinstance(x, (int, float)))]
        tare_non_numeric_values = tare_weight_data[~tare_weight_data.apply(lambda x: isinstance(x, (int, float)))]

        if gross_non_numeric_values.empty and tare_non_numeric_values.empty:
            invalid_rows = df[df['Tare Wt.'] > df['Gross Wt.']]
            if not invalid_rows.empty:
                invalid_row_numbers = list(invalid_rows.index)
                for rec in invalid_row_numbers:
                    get_data = sheet.cell(row=int(rec) + 2, column=error_column_index).value
                    values = get_data + '\n>Tare Weight cannot be greater than Gross Weight.' \
                        if get_data else '>Tare Weight cannot be greater than Gross Weight.'
                    sheet.cell(row=int(rec) + 2, column=error_column_index, value=values)

    @staticmethod
    def check_gross_and_tare_weight_validations(df, sheet, error_column_index):
        """
        Check gross and tare weight validations.
        :return:
        """
        try:
            UploadInventory.check_weight_column_validation(df, 'Gross Wt.', sheet, error_column_index)
            UploadInventory.check_weight_column_validation(df, 'Tare Wt.', sheet, error_column_index)
            UploadInventory.compare_gross_weight_and_tare_weight(df, sheet, error_column_index)

        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

    @staticmethod
    def validate_column_numeric(df, column_name, sheet, error_column_index, decimal_places=None):
        """
        Check numeric validations for column_name with decimal places.
        :param column_name:
        :param decimal_places:
        :param df:
        :param sheet:
        :param error_column_index:
        :return:
        """
        if column_name not in df.columns:
            raise ValidationError(_('Column "%s" not found in the uploaded file.') % column_name)

        column_data = df[column_name].dropna()

        # Check for non-numeric values
        if decimal_places is None:
            invalid_values = column_data[~column_data.apply(lambda x: isinstance(x, (int, float)))]
        else:
            invalid_values = column_data[
                ~column_data.apply(lambda x: isinstance(x, (int, float)) and round(x, decimal_places) == x)]

        if not invalid_values.empty:
            invalid_row_numbers = list(invalid_values.index)
            for rec in invalid_row_numbers:
                get_data = sheet.cell(row=int(rec) + 2, column=error_column_index).value
                values = get_data + f'\n>Please enter only numeric values with {decimal_places} decimal places in {column_name}.' \
                    if get_data else f'>Please enter only numeric values with {decimal_places} decimal places in {column_name}.'
                sheet.cell(row=int(rec) + 2, column=error_column_index, value=values)

    @staticmethod
    def check_approved_amount_validations(df, sheet, error_column_index):
        """
        Check approved amount column validations.
        :return:
        """
        try:
            UploadInventory.validate_column_numeric(df, 'Approved Amount', sheet, error_column_index, decimal_places=2)
        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

    @staticmethod
    def check_estimate_amount_validations(df, sheet, error_column_index):
        """
        Check estimate amount column validations.
        :return:
        """
        try:
            UploadInventory.validate_column_numeric(df, 'Estimate Amt', sheet, error_column_index, decimal_places=2)
        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

    def check_move_in_date_validations(self, df, sheet, error_column_index):
        """
        Check move in date validations.
        :return:
        """
        get_company = self.env['res.company'].search([('parent_id', '=', False), ('active', '=', True)], limit=1)
        company_format = get_company.date_format if get_company else False
        if company_format:
            try:
                if 'In Date' not in df.columns:
                    raise ValidationError(_('Column "In Date" not found in the uploaded file.'))

                # Apply validation based on company's date format to valid dates
                if company_format == 'DD/MM/YYYY':
                    valid_dates = df['In Date'].dropna().apply(
                        lambda x: x.strftime('%d/%m/%Y') if isinstance(x, (pd.Timestamp, datetime)) else str(x))
                    for index, x in valid_dates.items():
                        UploadInventory.validate_move_in_date_format(index, x, 'In Date', '%d/%m/%Y', 'DD/MM/YYYY', sheet, error_column_index)
                elif company_format == 'YYYY/MM/DD':
                    valid_dates = df['In Date'].dropna().apply(
                        lambda x: x.strftime('%Y/%m/%d') if isinstance(x, (pd.Timestamp, datetime)) else str(x))
                    for index, x in valid_dates.items():
                        UploadInventory.validate_move_in_date_format(index, x, 'In Date', '%Y/%m/%d', 'YYYY/MM/DD', sheet, error_column_index)
                elif company_format == 'MM/DD/YYYY':
                    valid_dates = df['In Date'].dropna().apply(
                        lambda x: x.strftime('%m/%d/%Y') if isinstance(x, (pd.Timestamp, datetime)) else str(x))
                    for index, x in valid_dates.items():
                        UploadInventory.validate_move_in_date_format(index, x, 'In Date', '%m/%d/%Y', 'MM/DD/YYYY', sheet, error_column_index)

            except pd.errors.ParserError as e:
                raise ValidationError(_('Error parsing Excel file: %s') % e)

    @staticmethod
    def validate_move_in_date_format(index, date_str, column_name, format_str, message, sheet, error_column_index):
        """
        validate move in date format.
        :param date_str:
        :param column_name:
        :param format_str:
        :param message:
        :param index:
        :param sheet:
        :param error_column_index:
        :return:
        """
        try:
            return datetime.strptime(date_str, format_str)
        except ValueError:
            get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
            values = get_data + f'\n>Please enter {column_name} in {message} format.Please enter Move in date in format as per company setting.' \
                if get_data else f'>Please enter {column_name} in {message} format.Please enter Move in date in format as per company setting.'
            sheet.cell(row=int(index) + 2, column=error_column_index, value=values)

    @staticmethod
    def validate_date_format(index, date_str, column_name, sheet, error_column_index):
        """
        Validate date format for passed column name.
        :param date_str:
        :param column_name:
        :param index:
        :param sheet:
        :param error_column_index:
        :return:
        """
        try:
            return datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
            values = get_data + f'\n>Please enter {column_name} in DD/MM/YYYY format.' \
                if get_data else f'>Please enter {column_name} in DD/MM/YYYY format.'
            sheet.cell(row=int(index) + 2, column=error_column_index, value=values)

    @staticmethod
    def estimate_approve_and_repair_dates_validations(df, sheet, error_column_index):
        """
        Check estimate and repair date columns validations.
        :return:
        """
        try:
            # Validate Estimate Date
            if 'Estimate Date' not in df.columns:
                raise ValidationError(_('Column "Estimate Date" not found in the uploaded file.'))
            estimate_date_data = df['Estimate Date'].dropna().apply(
                lambda x: x.strftime('%d/%m/%Y') if isinstance(x, (pd.Timestamp, datetime)) else str(x))
            for index, x in estimate_date_data.items():
                UploadInventory.validate_date_format(index, x, 'Estimate Date', sheet, error_column_index)

            # Validate Approval Date
            if 'Approval Date' not in df.columns:
                raise ValidationError(_('Column "Approval Date" not found in the uploaded file.'))
            approval_date_data = df['Approval Date'].dropna().apply(
                lambda x: x.strftime('%d/%m/%Y') if isinstance(x, (pd.Timestamp, datetime)) else str(x))

            for index, x in approval_date_data.items():
                UploadInventory.validate_date_format(index, x, 'Approval Date', sheet, error_column_index)

            # Approval Date cannot be less than Estimate Date
            for index, (approval_date, estimate_date) in enumerate(zip(approval_date_data, estimate_date_data)):
                try:
                    approval_date_dt = datetime.strptime(approval_date, '%d/%m/%Y')
                    estimate_date_dt = datetime.strptime(estimate_date, '%d/%m/%Y')
                except ValueError:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Invalid values to compare approval date and estimate date.' \
                        if get_data else '>Invalid values to compare approval date and estimate date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)
                    continue
                if approval_date_dt < estimate_date_dt:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Approval Date cannot be less than Estimate Date.' \
                        if get_data else '>Approval Date cannot be less than Estimate Date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)

            # Validate Repair Date
            if 'Repair Date' not in df.columns:
                raise ValidationError(_('Column "Repair Date" not found in the uploaded file.'))
            repair_date_data = df['Repair Date'].dropna().apply(
                lambda x: x.strftime('%d/%m/%Y') if isinstance(x, (pd.Timestamp, datetime)) else str(x))

            for index,x in repair_date_data.items():
                UploadInventory.validate_date_format(index, x, 'Repair Date', sheet, error_column_index)

            # Repair Date cannot be less than Estimate Date
            for index, (repair_date, estimate_date) in enumerate(zip(repair_date_data, estimate_date_data)):
                try:
                    repair_date_dt = datetime.strptime(repair_date, '%d/%m/%Y')
                    estimate_date_dt = datetime.strptime(estimate_date, '%d/%m/%Y')
                except ValueError:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Invalid values to compare repair date and estimate date.' \
                        if get_data else '>Invalid values to compare repair date and estimate date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)
                    continue
                if repair_date_dt < estimate_date_dt:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Repair Date cannot be less than Estimate Date.' \
                        if get_data else '>Repair Date cannot be less than Estimate Date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)

            # Repair Date cannot be less than Approval Date
            for index, (repair_date, approval_date) in enumerate(zip(repair_date_data, approval_date_data)):
                try:
                    repair_date_dt = datetime.strptime(repair_date, '%d/%m/%Y')
                    approval_date_dt = datetime.strptime(approval_date, '%d/%m/%Y')
                except ValueError:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Invalid values to compare repair date and approval date.' \
                        if get_data else '>Invalid values to compare repair date and approval date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)
                    continue
                if repair_date_dt < approval_date_dt:
                    get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
                    values = get_data + '\n>Repair Date cannot be less than Approval Date.' \
                        if get_data else '>Repair Date cannot be less than Approval Date.'
                    sheet.cell(row=int(index) + 2, column=error_column_index, value=values)

        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

        except ValueError as e:
            raise ValidationError(_('Invalid date format in one of the date columns: %s') % e)

    @staticmethod
    def check_in_time_validations(df, sheet, error_column_index):
        """
        Check in time columns validations.
        :return:
        """
        try:
            if 'In Time' not in df.columns:
                raise ValidationError(_('Column "In Time" not found in the uploaded file.'))

            items = df['In Time'].dropna().items()
            # Iterate over non-null items in 'In Time' column
            for index, value in items:
                UploadInventory.validate_in_time_format(str(value), index, sheet, error_column_index)

        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

    @staticmethod
    def validate_in_time_format(time_str, row_number, sheet, error_column_index):
        """
        Check validate in_time column format.
        :param time_str:
        :param row_number:
        :param sheet:
        :param error_column_index:
        :return:
        """
        try:
            datetime.strptime(time_str, '%H:%M:%S')
        except ValueError:
            get_data = sheet.cell(row=int(row_number) + 2, column=error_column_index).value
            values = get_data + '\n> invalid In time format' \
                if get_data else '> invalid In time format'
            sheet.cell(row=int(row_number) + 2, column=error_column_index, value=values)

    @staticmethod
    def check_production_month_year_validations(df, sheet, error_column_index):
        """
        check production month and year validations.
        :return:
        """
        try:
            if 'Production Month/Year' not in df.columns:
                raise ValidationError(_('Column "Production Month/Year" not found in the uploaded file.'))

            # Validate Production Month/Year format
            for index, x in df['Production Month/Year'].dropna().items():
                UploadInventory.validate_production_month_year_format(index, str(x), sheet, error_column_index)

        except pd.errors.ParserError as e:
            raise ValidationError(_('Error parsing Excel file: %s') % e)

    @staticmethod
    def validate_production_month_year_format(index, month_year_str, sheet, error_column_index):
        """
        Validate production month/year format.
        :param month_year_str:
        :param index:
        :param sheet:
        :param error_column_index:
        :return:
        """
        try:
            date_obj = pd.to_datetime(month_year_str, format='%m/%Y', errors='raise')
            year = date_obj.year
            if not (1950 <= year <= 3999):
                raise ValueError('Year must be between 1950 and 3999.')
        except ValueError as e:
            get_data = sheet.cell(row=int(index) + 2, column=error_column_index).value
            values = get_data + '\n>Production Month/Year" in MM/YYYY format.Year must be between 1950 to 3999.' \
                if get_data else '>Production Month/Year" in MM/YYYY format.Year must be between 1950 to 3999.'
            sheet.cell(row=int(index) + 2, column=error_column_index, value=values)
        except Exception as e:
            raise ValidationError(_('Validation error: %s') % e)

    @staticmethod
    def parse_date(date_value):
        """
        Parses a date value and returns it as a Python datetime object.
        :param date_value:
        :return:
        """
        if pd.isna(date_value):
            return None
        if isinstance(date_value, pd.Timestamp):
            return date_value.to_pydatetime()
        if isinstance(date_value, str):
            return datetime.strptime(date_value, '%d/%m/%Y')
        return None

    def move_in_parse_date(self, date_value):
        """
            Parses a date value according to the company's date format and returns it as a Python datetime object.

            This method fetches the company's date format and parses the date value accordingly.
            - If the date format is 'DD/MM/YYYY', it parses the date value as day/month/year.
            - If the date format is 'YYYY/MM/DD', it parses the date value as year/month/day.
            - If the date format is 'MM/DD/YYYY', it parses the date value as month/day/year.
            - If the date value is NaN, it returns None.
            - If the date value is a pandas Timestamp, it converts it to a Python datetime object.
        """
        get_company = self.env['res.company'].search([('parent_id', '=', False), ('active', '=', True)], limit=1)
        company_format = get_company.date_format if get_company else False
        if company_format:
            if pd.isna(date_value):
                return None
            if isinstance(date_value, pd.Timestamp):
                return date_value.to_pydatetime()
            if isinstance(date_value, str):
                if company_format == 'DD/MM/YYYY':
                    return datetime.strptime(date_value, '%d/%m/%Y')
                elif company_format == 'YYYY/MM/DD':
                    return datetime.strptime(date_value, '%Y/%m/%d')
                elif company_format == 'MM/DD/YYYY':
                    return datetime.strptime(date_value, '%m/%d/%Y')
                else:
                    return None
        return None

    @api.model
    def record_creation_for_container_master(self, df, sheet, error_column_index):
        """
        Create records in container master if above all validation is passed.
        :return:
        """
        try:

            error_column_values_empty = True

            # Iterate through all rows in the specified column 'Error'
            for row in sheet.iter_rows(min_col=error_column_index, max_col=error_column_index, min_row=2):
                for cell in row:
                    if cell.value is not None:
                        # If any cell in the column has a value, mark as not all empty
                        error_column_values_empty = False
                        break  # No need to check further if we found a non-empty cell
                if not error_column_values_empty:
                    break

            if error_column_values_empty:
                # records_to_create = []

                # Bulk search for related records
                shipping_lines = self.env['res.partner'].search([('is_shipping_line', '=', True)])
                shipping_lines_dict = {line.name: line.id for line in shipping_lines}

                types_sizes = self.env['container.type.data'].search([])
                types_sizes_dict = {type_data.company_size_type_code: type_data.id for type_data in types_sizes}

                damage_conditions = self.env['damage.condition'].search([])
                damage_conditions_dict = {condition.name: condition.id for condition in damage_conditions}

                # Search for existing containers in master and inventory
                inventory_containers = self.env['container.inventory'].search(
                    [('name', 'in', df['Container No.'].tolist())])
                containers_dict_inventory = {container.name: container for container in
                                             inventory_containers}

                master_containers = self.env['container.master'].search(
                    [('name', 'in', df['Container No.'].tolist())])
                master_containers_dict = {container.name: container for container in
                                          master_containers}

                # Iterate through each row of the DataFrame
                for index, row in df.iterrows():

                    shipping_line_id = shipping_lines_dict.get(row['Shipping Line'], False)
                    type_size_id = types_sizes_dict.get(row['Size/Type'], False)
                    damage_condition_id = damage_conditions_dict.get(row['Damage Condition'], False)
                    container_no = row['Container No.']

                    # Prepare values for container.master
                    vals_container = {
                        'shipping_line_id': shipping_line_id,
                        'type_size': type_size_id,
                        'name': container_no,
                        'gross_wt': int(row['Gross Wt.']),
                        'tare_wt': int(row['Tare Wt.']),
                        'month': str(row['Production Month/Year']).split('/')[0].zfill(2) if not pd.isna(row['Production Month/Year']) else False,
                        'year': str(row['Production Month/Year']).split('/')[1] if not pd.isna(row['Production Month/Year']) else False,
                        'is_import': True
                    }

                    # Create or update container.master record
                    if not container_no in master_containers_dict:
                        # Create new container master record
                        container_master = self.env['container.master'].create(vals_container)
                        master_containers_dict[container_no] = container_master
                    else:
                        # Update existing container master record
                        master_containers_dict[container_no].write(vals_container)
                        container_master = master_containers_dict[container_no]

                    # Prepare values for container.inventory
                    vals_inventory = {
                        'container_master_id': container_master.id,
                        'name': container_no,
                        'location_id': self.location_id.id,
                        'damage_condition': damage_condition_id,
                        'estimate_amount': float(row['Estimate Amt']) if not pd.isna(row['Estimate Amt']) else 0.0,
                        'approved_amount': float(row['Approved Amount']) if not pd.isna(
                            row['Approved Amount']) else 0.0,
                        'move_in_date': self.move_in_parse_date(row['In Date']),
                        'hour':  str(int(str(row['In Time']).split(':')[0])) if not pd.isna(row['In Time']) else False,
                        'minutes':  str(int(str(row['In Time']).split(':')[1])) if not pd.isna(row['In Time']) else False,
                        'repair_date': self.parse_date(row['Repair Date']),
                        'estimate_date': self.parse_date(row['Estimate Date']),
                        'approval_date': self.parse_date(row['Approval Date']),
                        'remarks': row['Remarks'],
                        'grade': row['Grade'].lower() if not pd.isna(row['Grade']) else False,
                        'status': row['Status'].lower() if not pd.isna(row['Status']) else False,
                        'is_import': True
                    }

                    # Create or update container.inventory record
                    if not container_no in containers_dict_inventory:
                        # Create new container inventory record
                        self.env['container.inventory'].create(vals_inventory)
                    else:
                        # Update existing container inventory record
                        containers_dict_inventory[container_no].write(vals_inventory)
            else:
                pass
        except Exception as e:
            raise ValidationError(_('some issues while creating a container master record %s', e))
        return True

    def update_inventory_records(self, wb, sheet, error_column_index):
            columns_to_widths = {'S': 45}

            # Iterate over the columns_to_widths dictionary
            for column, width in columns_to_widths.items():
                sheet.column_dimensions[column].width = width

            # Save workbook to BytesIO object
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)  # Move to the beginning of the BytesIO object

            # Encode the output to base64
            encoded_error_df = base64.b64encode(output.read())

            error_column_values_empty = True

            # Iterate through all rows in the specified column 'Error'
            for row in sheet.iter_rows(min_col=error_column_index, max_col=error_column_index, min_row=2):
                for cell in row:
                    if cell.value is not None:
                        # If any cell in the column has a value, mark as not all empty
                        error_column_values_empty = False
                        break  # No need to check further if we found a non-empty cell
                if not error_column_values_empty:
                    break

            if not error_column_values_empty:
                self.write({
                    'rec_status': 'error',
                    'upload_inventory_file': encoded_error_df
                })
            else:
                self.write({
                    'rec_status': 'success'
                })
