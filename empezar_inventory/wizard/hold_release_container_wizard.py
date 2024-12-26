import logging
import io
import os
import base64
import pandas as pd
import openpyxl
from odoo.exceptions import ValidationError
from odoo import fields, models, _

_logger = logging.getLogger(__name__)


class HoldReleaseContainerWizard(models.TransientModel):
    _name = "hold.release.container.wizard"
    _description = "Hold Release Container Wizard"

    file_name = fields.Char(string="Name")
    location_id = fields.Many2one('res.company', string="Location",
                                  required=True,
                                  domain="[('parent_id','!=',False)]")
    hold_reason_id = fields.Many2one('hold.reason', string="Hold Reason", required=True)
    remarks = fields.Char(string="Remarks", size=132)
    upload_file = fields.Binary(string="Upload File",
                                help="Only XLS or XLSX files with max size of 5 MB and"
                                     " having max 200 entries.",
                                required=True, copy=False, exportable=False)

    def action_submit(self):
        """ Check validations for uploaded excel files and create records in container
            master and hold/Release Containers
            model if all validations pass.
            :return:
        """
        self.check_file_type_and_size_validations()
        self.check_header_files_validations()
        self.check_require_columns_data()
        self.check_container_validations()
        self.check_data_length_validations()
        self.create_hold_release_container_record()

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

    def get_data_sheet(self):
        """ Get the data from the excel file
        :return:  list of data
        """
        data = []
        try:
            file_data = io.BytesIO(base64.b64decode(self.upload_file))
            wb = openpyxl.load_workbook(file_data)
            sheet = wb.active
            if sheet:
                data.append(sheet.max_row)
        except Exception as e:
            _logger.error("Error occurred while getting data from sheet: %s", str(e))
        return data

    def check_header_files_validations(self):
        """ compare sample file and uploaded file header if different then
            it will raise validation error.
            :return:
        """
        # Get the base path of the module
        module_base_path = os.path.dirname(os.path.dirname(__file__))

        # Construct the file path relative to the module's directory

        file_relative_path = 'static/src/document/hold_containers_sample.xlsx'
        sample_file_path = os.path.join(module_base_path, file_relative_path)

        # Check if the sample file exists
        if not os.path.exists(sample_file_path):
            raise ValidationError(_('The specified sample Excel file does not exist.'))

        # Read the sample Excel file to get its columns
        sample_df = pd.read_excel(sample_file_path)
        sample_file_columns = sample_df.columns.tolist()

        # Check if an upload file is provided
        if not self.upload_file:
            raise ValidationError(_('No update container status file uploaded.'))

        # Read the uploaded Excel file
        file_content = base64.b64decode(self.upload_file)
        # file_content = pd.read_excel(self.upload_file)
        df = pd.read_excel(io.BytesIO(file_content))
        upload_file_columns = df.columns.tolist()

        # Compare headers
        if sample_file_columns != upload_file_columns:
            raise ValidationError(_('Invalid columns headers in the file uploaded'))

        return True

    def check_container_validations(self):
        """
        Container No. columns validations.
        :return:
        """
        if self.upload_file:
            file_content = base64.b64decode(self.upload_file)
            df = pd.read_excel(io.BytesIO(file_content))

            duplicate_containers = df['Container No.'].dropna()[df['Container No.'].dropna().duplicated()].tolist()

            if duplicate_containers:
                raise ValidationError(_('%s Duplicate container number found.',
                                        duplicate_containers))
            container_no_data = df['Container No.'].dropna().tolist()
            hold_container = self.env['hold.release.containers'].search([]).mapped('inventory_id').mapped('name')
            container_inventory = self.env['container.inventory'].search([]).mapped('name')
            get_active_container = self.env['container.inventory'].search(
                [('location_id', '=', self.location_id.id)]).mapped('name')
            if not get_active_container:
                raise ValidationError(_('There is no any active Container'))
            for rec in container_no_data:
                rec = rec.strip(" ")
                if rec in hold_container:
                    raise ValidationError(_('Container number already on hold. Please release the container first.'))
                elif rec not in container_inventory:
                    raise ValidationError(_('Container Number is not present in the Inventory.'))
                elif not rec in get_active_container:
                    raise ValidationError(_('%s Container Number not found with Selected location', rec))

    def check_require_columns_data(self):
        """
        Validate that specified columns do not contain empty values.
        :return:
        """
        if self.upload_file:
            file_content = base64.b64decode(self.upload_file)
            df = pd.read_excel(io.BytesIO(file_content))
            require_columns = ['Container No.']
            for column in require_columns:
                if df[column].isnull().any() or df[column].eq('').any():
                    raise ValidationError(f"Column '{column}' should not contain empty values.")
            # Validate that specified columns do not contain empty or zero values.

    def get_containers(self):
        """
        Get the data from the excel file
        :return: list of container numbers
        """
        container_numbers = []
        try:
            file_data = io.BytesIO(base64.b64decode(self.upload_file))
            wb = openpyxl.load_workbook(file_data)
            sheet = wb.active
            if sheet:
                for row in sheet.iter_rows(values_only=True):
                    container_no = row[0]  # Assuming the container number is in the first column
                    if container_no and isinstance(container_no,
                                                   str) and container_no.strip() and container_no != 'Container No.':
                        container_numbers.append(container_no.strip())
        except Exception as e:
            _logger.error("Error occurred while getting containers from Excel file: %s", str(e))
        return container_numbers

    def check_file_type_and_size_validations(self):
        """
        Check file type and file size validations.
        :return:
        """

        if self.upload_file:
            # Get the file name and check the extension
            file_name = self.file_name
            if not file_name.lower().endswith(('.xls', '.xlsx')):
                raise ValidationError(
                    _("File Type selected is not allowed.Please upload files of types-.xls /.xlsx."))
            if len(self.upload_file) > 5242880:  # 5MB in bytes
                raise ValidationError(_("The maximum file size allowed should be 5 MB."))

    def check_data_length_validations(self):
        """
        Verify the row size of uploaded Excel file.
        :return:
        """
        if self.upload_file:
            # Read data from the uploaded Excel file
            data = self.get_data_sheet()
            if data and data[0] > 201:
                raise ValidationError(_('Records entered are more than 200. Kindly'
                                        ' reupload with 200 or fewer records.'))

    def create_hold_release_container_record(self):
        """
        Create or update record in upload Hold Release container record
        while successfully passed all validations.
        """
        try:
            container_numbers = self.get_containers()
            for container_no in container_numbers:
                container = self.env['container.inventory'].search([('name', '=', container_no)], limit=1)
                container.write({'hold_release_status': 'hold'})
                if container:
                    vals = {
                        'location_id': self.location_id.id,
                        'hold_reason_id': self.hold_reason_id.id,
                        'hold_date': fields.Datetime.now(),
                        'inventory_id': container.id,
                        'remarks': self.remarks
                    }
                    existing_record = self.env['hold.release.containers'].search([('inventory_id', '=', container.id)])
                    if existing_record:
                        existing_record.write(vals)
                    else:
                        self.env['hold.release.containers'].create(vals)
                else:
                    raise ValidationError(_('Container with name %s not found in container')
                                          % container_no)
        except Exception as e:
            _logger.error("Error occurred while creating Hold Release Containers record: %s",str(e))
            raise ValidationError(_('Some issues while Hold Release Containers data %s', str(e)))


class HoldContainerWizard(models.TransientModel):
    _name = "release.container.wizard"
    _description = "Release Container Wizard"
    _rec_name = "display_name"

    release_container_id = fields.Many2one('hold.release.containers', string="Release Container", required=True)
    display_name = fields.Char("Name", required=True)

    def action_release(self):
        """ Action method to release a held container and update related records.
            Returns:
                dict or None: Action dictionary for client action or None if no action is performed.
        """
        if self.release_container_id:
            container = self.env['container.inventory'].search([('name', '=', self.display_name)], limit=1)
            if container:
                container.write({'hold_release_status': 'release'})
            self.release_container_id.unlink()